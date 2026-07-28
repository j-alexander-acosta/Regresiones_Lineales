import streamlit as st
import pandas as pd
import numpy as np
from sklearn.linear_model import LinearRegression
from sklearn.metrics import r2_score
import scipy.stats as stats
import plotly.graph_objects as plgo
import matplotlib.pyplot as plt
import io
from fpdf import FPDF
import importlib
import translations
importlib.reload(translations)
from translations import TRANSLATIONS

# --- HELPER FUNCTIONS FOR STREAMLIT DATA EDITOR CALLBACKS ---
def update_session_df(df_key, editor_key):
    if editor_key not in st.session_state:
        return
    editor_state = st.session_state[editor_key]
    df = st.session_state[df_key]
    
    # Apply edits
    edited_rows = editor_state.get("edited_rows", {})
    for row_idx_str, col_edits in edited_rows.items():
        row_idx = int(row_idx_str)
        for col_name, val in col_edits.items():
            df.at[row_idx, col_name] = val
            
    # Apply additions
    added_rows = editor_state.get("added_rows", [])
    if added_rows:
        added_df = pd.DataFrame(added_rows)
        for col in df.columns:
            if col not in added_df.columns:
                added_df[col] = np.nan
        added_df = added_df[df.columns]
        df = pd.concat([df, added_df], ignore_index=True)
        
    # Apply deletions
    deleted_rows = editor_state.get("deleted_rows", [])
    if deleted_rows:
        df = df.drop(deleted_rows).reset_index(drop=True)
        
    st.session_state[df_key] = df

def update_mc_true_df():
    if "mc_editor" not in st.session_state:
        return
    editor_state = st.session_state["mc_editor"]
    df = st.session_state.mc_true_df
    
    # Apply edits
    edited_rows = editor_state.get("edited_rows", {})
    for row_idx_str, col_edits in edited_rows.items():
        row_idx = int(row_idx_str)
        for col_name, val in col_edits.items():
            if col_name == "True mbm_bmb":
                df.at[row_idx, "True X (mbm_bmb)"] = val
            elif col_name == "True MwM_wMw":
                df.at[row_idx, "True Y (MwM_wMw)"] = val
                
    # Apply additions
    added_rows = editor_state.get("added_rows", [])
    if added_rows:
        added_records = []
        for row in added_rows:
            added_records.append({
                "True X (mbm_bmb)": row.get("True mbm_bmb", np.nan),
                "True Y (MwM_wMw)": row.get("True MwM_wMw", np.nan)
            })
        added_df = pd.DataFrame(added_records)
        df = pd.concat([df, added_df], ignore_index=True)
        
    # Apply deletions
    deleted_rows = editor_state.get("deleted_rows", [])
    if deleted_rows:
        df = df.drop(deleted_rows).reset_index(drop=True)
        
    st.session_state.mc_true_df = df


# --- CLASE HELPER PARA ERRORES EN LAS VARIABLES (CARROLL & RUPPERT, 1996) ---
class LinearErrorsInVariables:
    """
    Clase para el cálculo de regresiones lineales considerando errores de medición.
    Calcula OLS, Método de Momentos (corregido por atenuación) y Regresión de Deming.
    """
    def __init__(self, W, Y, sigma_u, sigma_e):
        self.W = np.asarray(W, dtype=float)
        self.Y = np.asarray(Y, dtype=float)
        self.sigma_u = float(sigma_u)
        self.sigma_e = float(sigma_e)
        self.n = len(self.W)
        
        # A. Estimadores Base de la Muestra
        self.mean_w = np.mean(self.W) if self.n > 0 else 0.0
        self.mean_y = np.mean(self.Y) if self.n > 0 else 0.0
        self.s2_w = np.var(self.W, ddof=1) if self.n > 1 else 0.0
        self.s2_y = np.var(self.Y, ddof=1) if self.n > 1 else 0.0
        
        # Covarianza muestral
        if self.n > 1:
            cov_mat = np.cov(self.W, self.Y)
            self.s_wy = cov_mat[0, 1]
        else:
            self.s_wy = 0.0
            
        # B. Mínimos Cuadrados Ordinarios (OLS)
        self.beta1_ols = self.s_wy / self.s2_w if self.s2_w != 0.0 else 0.0
        self.beta0_ols = self.mean_y - self.beta1_ols * self.mean_w
        
        # C. Método de Momentos (MoM) Corregido por Atenuación
        # Varianza estimada de la variable verdadera X (s2_x = s2_w - sigma_u^2)
        s2_x_est = self.s2_w - self.sigma_u**2
        if s2_x_est <= 0.0001:
            s2_x_est = 0.0001  # Límite inferior para evitar división por cero o pendientes extremas
            
        self.beta1_mom = self.s_wy / s2_x_est
        self.beta0_mom = self.mean_y - self.beta1_mom * self.mean_w
        # Factor de confiabilidad (reliability ratio)
        self.lambda_ratio = s2_x_est / self.s2_w if self.s2_w != 0.0 else 1.0
        
        # D. Regresión Ortogonal (Deming)
        self.eta = (self.sigma_e**2) / (self.sigma_u**2) if self.sigma_u != 0.0 else 1.0
        if self.s_wy != 0.0:
            beta1_num = (self.s2_y - self.eta * self.s2_w) + np.sqrt((self.s2_y - self.eta * self.s2_w)**2 + 4.0 * self.eta * self.s_wy**2)
            beta1_den = 2.0 * self.s_wy
            self.beta1_dem = beta1_num / beta1_den
        else:
            self.beta1_dem = 0.0
        self.beta0_dem = self.mean_y - self.beta1_dem * self.mean_w
        
    def compute_all_metrics(self):
        # OLS
        y_pred_ols = self.beta1_ols * self.W + self.beta0_ols
        sse_ols = np.sum((self.Y - y_pred_ols)**2) if self.n > 0 else 0.0
        s2_e_ols = sse_ols / (self.n - 2) if self.n > 2 else 0.0
        se_beta1_ols = np.sqrt(s2_e_ols / ((self.n - 1) * self.s2_w)) if self.s2_w != 0.0 and self.n > 2 else 0.0
        se_beta0_ols = se_beta1_ols * np.sqrt(np.sum(self.W**2) / self.n) if self.n > 0 else 0.0
        rmse_ols = np.sqrt(sse_ols / self.n) if self.n > 0 else 0.0
        r2_ols = 1.0 - (sse_ols / ((self.n - 1) * self.s2_y)) if self.s2_y != 0.0 and self.n > 1 else 0.0
        
        # MoM (Atenuación)
        y_pred_mom = self.beta1_mom * self.W + self.beta0_mom
        sse_mom = np.sum((self.Y - y_pred_mom)**2) if self.n > 0 else 0.0
        # SE aproximado por el delta method simplificado (inflación del SE de OLS por lambda)
        se_beta1_mom = se_beta1_ols / self.lambda_ratio if self.n > 2 else 0.0
        se_beta0_mom = se_beta1_mom * np.sqrt(np.sum(self.W**2) / self.n) if self.n > 0 else 0.0
        rmse_mom = np.sqrt(sse_mom / self.n) if self.n > 0 else 0.0
        r2_mom = 1.0 - (sse_mom / ((self.n - 1) * self.s2_y)) if self.s2_y != 0.0 and self.n > 1 else 0.0
        
        # Deming
        y_pred_dem = self.beta1_dem * self.W + self.beta0_dem
        sse_dem = np.sum((self.Y - y_pred_dem)**2) if self.n > 0 else 0.0
        s2_e_pseudo_dem = sse_dem / (self.n - 2) if self.n > 2 else 0.0
        se_beta1_dem = np.sqrt(s2_e_pseudo_dem / ((self.n - 1) * self.s2_w)) if self.s2_w != 0.0 and self.n > 2 else 0.0
        se_beta0_dem = se_beta1_dem * np.sqrt(np.sum(self.W**2) / self.n) if self.n > 0 else 0.0
        rmse_dem = np.sqrt(sse_dem / self.n) if self.n > 0 else 0.0
        r2_dem = 1.0 - (sse_dem / ((self.n - 1) * self.s2_y)) if self.s2_y != 0.0 and self.n > 1 else 0.0
        
        # Proyecciones Ortogonales verdaderas para Deming
        X_t = (self.beta1_dem * (self.Y - self.beta0_dem) + self.eta * self.W) / (self.eta + self.beta1_dem**2) if (self.eta + self.beta1_dem**2) != 0.0 else self.W
        Y_t = self.beta0_dem + self.beta1_dem * X_t
        s2_e_dem_ort = (1.0 / (self.n - 2)) * np.sum(((self.Y - self.beta0_dem - self.beta1_dem * self.W)**2) / (self.beta1_dem**2 + self.eta)) if self.n > 2 else 0.0
        sigma_dem_ort = np.sqrt(s2_e_dem_ort)
        
        return {
            "ols": {
                "slope": self.beta1_ols,
                "intercept": self.beta0_ols,
                "se_slope": se_beta1_ols,
                "se_intercept": se_beta0_ols,
                "rmse": rmse_ols,
                "r2": r2_ols,
                "y_pred": y_pred_ols
            },
            "mom": {
                "slope": self.beta1_mom,
                "intercept": self.beta0_mom,
                "se_slope": se_beta1_mom,
                "se_intercept": se_beta0_mom,
                "rmse": rmse_mom,
                "r2": r2_mom,
                "y_pred": y_pred_mom,
                "lambda": self.lambda_ratio
            },
            "deming": {
                "slope": self.beta1_dem,
                "intercept": self.beta0_dem,
                "se_slope": se_beta1_dem,
                "se_intercept": se_beta0_dem,
                "rmse": rmse_dem,
                "r2": r2_dem,
                "y_pred": y_pred_dem,
                "X_t": X_t,
                "Y_t": Y_t,
                "s2_e_ort": s2_e_dem_ort,
                "sigma_ort": sigma_dem_ort
            }
        }

# --- IDIOMA / LANGUAGE SELECTOR ---
if 'lang' not in st.session_state:
    st.session_state.lang = 'es'

t = TRANSLATIONS[st.session_state.lang]

# --- CONFIGURACIÓN DE PÁGINA ---
st.set_page_config(page_title=t["page_title"], page_icon="📈", layout="wide")

# Estilos CSS
st.markdown("""
<style>
    .reportview-container .main .block-container{
        padding-top: 2rem;
    }
    .latex-container {
        background-color: #f8f9fa;
        padding: 15px;
        border-radius: 5px;
        border-left: 5px solid #1f77b4;
        margin-bottom: 15px;
    }
    /* Ocultar el botón de Deploy */
    .stDeployButton, .stAppDeployButton {
        display: none !important;
    }
</style>
""", unsafe_allow_html=True)

st.title(t["st_title"])
st.markdown(t["st_desc"])

# --- BARRA LATERAL: ENTRADA DE DATOS ---
st.sidebar.markdown("### 🌐 Idioma / Language")
col_lang1, col_lang2 = st.sidebar.columns(2)
with col_lang1:
    if st.button("Español", use_container_width=True, type="primary" if st.session_state.lang == 'es' else "secondary"):
        st.session_state.lang = 'es'
        st.rerun()
with col_lang2:
    if st.button("English", use_container_width=True, type="primary" if st.session_state.lang == 'en' else "secondary"):
        st.session_state.lang = 'en'
        st.rerun()

st.sidebar.markdown("---")
analysis_type = st.sidebar.selectbox(t["sb_analysis_type"], [
    t["sb_simple_regression"], 
    t["sb_multiple_regression"],
    t["sb_probability_distributions"],
    t["sb_monte_carlo_regression"],
    t["sb_carroll_ruppert_table"],
    t["sb_markov_chain"],
    t["sb_seismic_bssa"]
])
is_mlr = (analysis_type == t["sb_multiple_regression"])
is_prob = (analysis_type == t["sb_probability_distributions"])
is_mc_reg = (analysis_type == t["sb_monte_carlo_regression"])
is_cr_table = (analysis_type == t["sb_carroll_ruppert_table"])
is_markov_chain = (analysis_type == t["sb_markov_chain"])
is_seismic_bssa = (analysis_type == t["sb_seismic_bssa"])

if is_cr_table:
    st.header(t["cr_title"])
    st.markdown(t["cr_desc"])
    
    st.markdown("""
    <style>
    .latex-container {
        background-color: #f8f9fa;
        padding: 15px;
        border-radius: 5px;
        border-left: 5px solid #1f77b4;
        margin-bottom: 20px;
    }
    </style>
    """, unsafe_allow_html=True)
    
    st.markdown("<div class='latex-container'>", unsafe_allow_html=True)
    st.markdown("<h4 style='margin-top:0;'>Fórmulas Matemáticas Utilizadas / Mathematical Formulas Used:</h4>", unsafe_allow_html=True)
    
    st.markdown("1. **Relación de Varianzas (Sin error de ecuación) / Variance Ratio (Without equation error):**")
    st.latex(r"\eta_{\text{no\_eq}} = \left(\frac{\sigma_y}{\sigma_x}\right)^2")
    
    st.markdown("2. **Pendiente del Método de Momentos / Method of Moments (MoM) Slope:**")
    st.latex(r"\beta_1^{\text{MoM}} = \frac{s_{wy}}{s_w^2 - \sigma_x^2}")
    
    st.markdown("3. **Varianza de Residuos MoM / MoM Residuals Variance:**")
    st.latex(r"s_{\text{res},\text{MoM}}^2 = \frac{n-1}{n-2} \left[ s_y^2 - 2 \beta_1^{\text{MoM}} s_{wy} + (\beta_1^{\text{MoM}})^2 s_w^2 \right]")
    
    st.markdown("4. **Varianza del Error de Ecuación / Equation Error Variance ($\\sigma_q^2$):**")
    st.latex(r"\sigma_q^2 = s_{\text{res},\text{MoM}}^2 - (\beta_1^{\text{MoM}})^2 \sigma_x^2 - \sigma_y^2")
    
    st.markdown("5. **Relación de Varianzas (Con error de ecuación) / Variance Ratio (With equation error):**")
    st.latex(r"\eta_{\text{with\_eq}} = \frac{\sigma_q^2 + \sigma_y^2}{\sigma_x^2}")
    
    st.markdown("6. **Pendiente de Regresión Ortogonal (GOR) / Orthogonal Regression (GOR) Slope:**")
    st.latex(r"\beta_{1,\text{OR}} = \frac{(s_y^2 - \eta s_w^2) + \sqrt{(s_y^2 - \eta s_w^2)^2 + 4 \eta s_{wy}^2}}{2 s_{wy}}")
    
    st.markdown("</div>", unsafe_allow_html=True)
    
    st.sidebar.markdown(f"### {t['sb_header_data']}")
    cr_data_source = st.sidebar.radio(t["sb_data_source"], (t["sb_manual"], t["sb_upload"]), key="cr_ds")
    
    cr_df = None
    if cr_data_source == t["sb_manual"]:
        if "cr_manual_df_v2" not in st.session_state:
            # Preload dataset exactly matching the covariance in the user's screenshot
            desired_mean = np.array([5.0, 5.0])
            desired_cov = np.array([[0.408359, 0.371890], [0.371890, 0.425499]])
            np.random.seed(42)
            Z = np.random.normal(0, 1, (100, 2))
            Z = Z - np.mean(Z, axis=0)
            S_Z = np.cov(Z, rowvar=False)
            L_Z = np.linalg.cholesky(S_Z)
            Z_white = np.dot(Z, np.linalg.inv(L_Z).T)
            L = np.linalg.cholesky(desired_cov)
            X_final = np.dot(Z_white, L.T) + desired_mean
            st.session_state.cr_manual_df_v2 = pd.DataFrame({
                "X": np.round(X_final[:, 0], 6),
                "Y": np.round(X_final[:, 1], 6)
            })
        
        st.markdown(f"### {t['cr_data_editor_title']}")
        cr_df = st.data_editor(
            st.session_state.cr_manual_df_v2,
            num_rows="dynamic",
            use_container_width=True,
            column_config={
                "X": st.column_config.NumberColumn("X", format="%.6f"),
                "Y": st.column_config.NumberColumn("Y", format="%.6f")
            },
            key="cr_manual_editor_v2",
            on_change=update_session_df,
            args=("cr_manual_df_v2", "cr_manual_editor_v2")
        )
    else:
        uploaded_file = st.sidebar.file_uploader(t["sb_upload_desc"], type=["csv", "xlsx"], key="cr_uploader")
        if uploaded_file is not None:
            try:
                if uploaded_file.name.endswith('.csv'):
                    df_raw = pd.read_csv(uploaded_file)
                else:
                    df_raw = pd.read_excel(uploaded_file)
                
                st.sidebar.markdown(t["sb_select_cols"])
                x_col = st.sidebar.selectbox(t["sb_col_x"], df_raw.columns, key="cr_x_col")
                y_col = st.sidebar.selectbox(t["sb_col_y"], df_raw.columns, key="cr_y_col")
                
                cr_df = df_raw[[x_col, y_col]].rename(columns={x_col: "X", y_col: "Y"}).dropna()
            except Exception as e:
                st.sidebar.error(t["sb_err_file"].format(e))
        else:
            st.info(t["info_data_points"] if "info_data_points" in t else "Por favor sube un archivo para comenzar.")
            
    st.sidebar.markdown("---")
    st.sidebar.markdown("### Parámetros de Incertidumbre")
    
    default_sig_x_str = "0.200000, 0.100000"
    sig_x_input_str = st.sidebar.text_input(t["cr_sigma_x_list"], value=default_sig_x_str, key="cr_sig_x_list_v2")
    
    default_sig_y_str = "0.024000, 0.045000, 0.070000, 0.100000, 0.120000, 0.150000, 0.180000, 0.200000"
    sig_y_input_str = st.sidebar.text_input(t["cr_sigma_y_list"], value=default_sig_y_str, key="cr_sig_y_list_v2")
    
    try:
        sigma_x_list = [float(x.strip()) for x in sig_x_input_str.split(",") if x.strip()]
    except Exception:
        st.sidebar.error("Error al procesar sigma_x. Asegúrate de separar los números con comas.")
        sigma_x_list = [0.2, 0.1]
        
    try:
        sigma_y_list = [float(y.strip()) for y in sig_y_input_str.split(",") if y.strip()]
    except Exception:
        st.sidebar.error("Error al procesar sigma_y. Asegúrate de separar los números con comas.")
        sigma_y_list = [0.024, 0.045, 0.07, 0.1, 0.12, 0.15, 0.18, 0.2]
        
    if cr_df is not None and len(cr_df) >= 3:
        n_points = len(cr_df)
        X_vals = cr_df["X"].values
        Y_vals = cr_df["Y"].values
        
        sw2 = np.var(X_vals, ddof=1)
        sy2 = np.var(Y_vals, ddof=1)
        swy = np.cov(X_vals, Y_vals)[0, 1]
        
        st.markdown(f"### {t['cr_statistics_hdr']}")
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric(t["cr_n_points"], n_points)
        with col2:
            st.metric(t["cr_sw2"], f"{sw2:.6f}")
        with col3:
            st.metric(t["cr_sy2"], f"{sy2:.6f}")
        with col4:
            st.metric(t["cr_swy"], f"{swy:.6f}")
            
        ols_slope = swy / sw2 if sw2 != 0 else 0.0
        
        cr_rows = []
        for sx in sigma_x_list:
            for sy in sigma_y_list:
                eta_no_eq = (sy / sx)**2 if sx != 0 else np.nan
                
                denom_mm = sw2 - sx**2
                if denom_mm > 1e-6:
                    slope_mm = swy / denom_mm
                else:
                    slope_mm = np.nan
                    
                if not np.isnan(slope_mm):
                    s_res_mom_2 = ((n_points - 1) / (n_points - 2)) * (sy2 - 2 * slope_mm * swy + (slope_mm**2) * sw2)
                    sigma_q2 = s_res_mom_2 - (slope_mm**2) * (sx**2) - sy**2
                else:
                    sigma_q2 = np.nan
                    
                if not np.isnan(sigma_q2) and sx != 0:
                    eta_with_eq = (sigma_q2 + sy**2) / (sx**2)
                else:
                    eta_with_eq = np.nan
                    
                if not np.isnan(eta_no_eq) and swy != 0:
                    num = (sy2 - eta_no_eq * sw2) + np.sqrt((sy2 - eta_no_eq * sw2)**2 + 4 * eta_no_eq * swy**2)
                    slope_or_no_eq = num / (2 * swy)
                else:
                    slope_or_no_eq = np.nan
                    
                if not np.isnan(eta_with_eq) and eta_with_eq > 0 and swy != 0:
                    num = (sy2 - eta_with_eq * sw2) + np.sqrt((sy2 - eta_with_eq * sw2)**2 + 4 * eta_with_eq * swy**2)
                    slope_or_with_eq = num / (2 * swy)
                else:
                    slope_or_with_eq = np.nan
                    
                cr_rows.append({
                    "sigma_x": sx,
                    "sigma_y": sy,
                    "eta_no_eq": eta_no_eq,
                    "sigma_q^2": sigma_q2,
                    "eta_with_eq": eta_with_eq,
                    "slope_OLS": ols_slope,
                    "slope_OR(no_eq)": slope_or_no_eq,
                    "slope_MM": slope_mm,
                    "slope_OR(with_eq)": slope_or_with_eq
                })
                
        res_df = pd.DataFrame(cr_rows)
        
        st.markdown(f"### {t['cr_table_title']}")
        st.dataframe(
            res_df,
            column_config={
                "sigma_x": st.column_config.NumberColumn("sigma_x", format="%.6f"),
                "sigma_y": st.column_config.NumberColumn("sigma_y", format="%.6f"),
                "eta_no_eq": st.column_config.NumberColumn("eta_no_eq", format="%.6f"),
                "sigma_q^2": st.column_config.NumberColumn("sigma_q^2", format="%.6f"),
                "eta_with_eq": st.column_config.NumberColumn("eta_with_eq", format="%.6f"),
                "slope_OLS": st.column_config.NumberColumn("slope_OLS", format="%.6f"),
                "slope_OR(no_eq)": st.column_config.NumberColumn("slope_OR(no_eq)", format="%.6f"),
                "slope_MM": st.column_config.NumberColumn("slope_MM", format="%.6f"),
                "slope_OR(with_eq)": st.column_config.NumberColumn("slope_OR(with_eq)", format="%.6f"),
            },
            use_container_width=True
        )
        
        col_dl1, col_dl2 = st.columns(2)
        with col_dl1:
            csv_cr = res_df.to_csv(index=False).encode('utf-8')
            st.download_button(
                label=t["cr_download_csv"],
                data=csv_cr,
                file_name="carroll_ruppert_comparison.csv",
                mime="text/csv",
                key="cr_download_csv_btn"
            )
        with col_dl2:
            output_excel = io.BytesIO()
            with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
                res_df.to_excel(writer, sheet_name="Comparison", index=False)
            excel_data = output_excel.getvalue()
            st.download_button(
                label=t["cr_download_excel"],
                data=excel_data,
                file_name="carroll_ruppert_comparison.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="cr_download_xlsx_btn"
            )
        
        # --- SECCIÓN DE GRÁFICO COMPARATIVO ---
        st.markdown("---")
        st.subheader(t["cr_plot_title"])
        
        # Select error combination to plot
        combinations_labels = [f"σ_x = {row['sigma_x']:.4f}, σ_y = {row['sigma_y']:.4f}" for index, row in res_df.iterrows()]
        selected_comb_label = st.selectbox(t["cr_plot_select"], combinations_labels, key="cr_comb_select")
        selected_index = combinations_labels.index(selected_comb_label)
        selected_row = res_df.iloc[selected_index]
        
        # Slopes
        s_ols = selected_row["slope_OLS"]
        s_or_no_eq = selected_row["slope_OR(no_eq)"]
        s_mm = selected_row["slope_MM"]
        s_or_with_eq = selected_row["slope_OR(with_eq)"]
        
        # Intercepts: beta0 = mean_y - beta1 * mean_x
        mean_x = np.mean(X_vals)
        mean_y = np.mean(Y_vals)
        
        c_ols = mean_y - s_ols * mean_x
        c_or_no_eq = mean_y - s_or_no_eq * mean_x if not np.isnan(s_or_no_eq) else np.nan
        c_mm = mean_y - s_mm * mean_x if not np.isnan(s_mm) else np.nan
        c_or_with_eq = mean_y - s_or_with_eq * mean_x if not np.isnan(s_or_with_eq) else np.nan
        
        # Plotly Figure
        fig_plotly = plgo.Figure()
        
        # Observed Data points
        fig_plotly.add_trace(plgo.Scatter(
            x=X_vals, y=Y_vals, mode='markers', name=t["chart_obs"],
            marker=dict(color='black', size=8),
            hovertemplate='X: %{x}<br>Y: %{y}<extra></extra>'
        ))
        
        # X line span
        x_line = np.linspace(min(X_vals), max(X_vals), 100)
        
        # 1. OLS Line
        fig_plotly.add_trace(plgo.Scatter(
            x=x_line, y=s_ols * x_line + c_ols, mode='lines', name=t["cr_plot_ols"],
            line=dict(color='blue', width=2),
            hovertemplate='X: %{x:.4f}<br>Y: %{y:.4f}<extra></extra>'
        ))
        
        # 2. OR (no_eq) Line
        if not np.isnan(s_or_no_eq):
            fig_plotly.add_trace(plgo.Scatter(
                x=x_line, y=s_or_no_eq * x_line + c_or_no_eq, mode='lines', name=t["cr_plot_or_no_eq"],
                line=dict(color='orange', width=2, dash='dash'),
                hovertemplate='X: %{x:.4f}<br>Y: %{y:.4f}<extra></extra>'
            ))
            
        # 3. MM Line
        if not np.isnan(s_mm):
            fig_plotly.add_trace(plgo.Scatter(
                x=x_line, y=s_mm * x_line + c_mm, mode='lines', name=t["cr_plot_mm"],
                line=dict(color='magenta', width=2, dash='dashdot'),
                hovertemplate='X: %{x:.4f}<br>Y: %{y:.4f}<extra></extra>'
            ))
            
        # 4. OR (with_eq) Line
        if not np.isnan(s_or_with_eq):
            fig_plotly.add_trace(plgo.Scatter(
                x=x_line, y=s_or_with_eq * x_line + c_or_with_eq, mode='lines', name=t["cr_plot_or_with_eq"],
                line=dict(color='red', width=2.5, dash='dot'),
                hovertemplate='X: %{x:.4f}<br>Y: %{y:.4f}<extra></extra>'
            ))
            
        fig_plotly.update_layout(
            title=t["cr_plot_title"],
            xaxis_title=t["cr_plot_x"],
            yaxis_title=t["cr_plot_y"],
            hovermode='closest',
            template='plotly_white',
            legend=dict(yanchor="top", y=0.99, xanchor="left", x=0.01)
        )
        
        st.plotly_chart(fig_plotly, use_container_width=True)
            
    else:
        st.warning("Se requieren al menos 3 puntos de datos para calcular la tabla de Carroll & Ruppert (debido al cálculo de varianza de residuos con n - 2 grados de libertad).")
        
    st.markdown("---")
    st.markdown(
        "<p style='text-align: center; color: gray; font-size: 14px;'>"
        "Desarrollado y mantenido por <b>Alexander Acosta</b> "
        "(<a href='https://github.com/j-alexander-acosta' target='_blank' style='color: #1f77b4; text-decoration: none;'>@j-alexander-acosta</a>)"
        "</p>", 
        unsafe_allow_html=True
    )
    st.stop()

if is_markov_chain:
    st.sidebar.markdown(f"### {t['sb_header_data']}")
    
    model_type = st.sidebar.radio(
        t["mkv_model_mode_label"],
        [t["mkv_mode_standard"], t["mkv_mode_hmm"], t["mkv_mode_game"], t["mkv_mode_business"]],
        key="mkv_model_mode"
    )
    
    if model_type == t["mkv_mode_game"]:
        import json
        import streamlit.components.v1 as components
        
        st.header(t["game_title"])
        st.markdown(t["game_desc"])
        
        js_translations = {
            "game_title": t["game_title"],
            "game_desc": t["game_desc"],
            "brand_a": t["game_brand_a"],
            "brand_b": t["game_brand_b"],
            "brand_c": t["game_brand_c"],
            "row_warning": t["game_row_warning"],
            "row_success": t["game_row_success"],
            "validate_btn": t["game_validate_btn"],
            "success_msg": t["game_success_msg"],
            "steady_state_title": t["game_steady_state_title"],
            "load_example": t["game_load_example"],
            "clear_matrix": t["game_clear_matrix"]
        }
        
        try:
            with open("markov_game.html", "r", encoding="utf-8") as f:
                html_template = f.read()
            html_content = html_template.replace("/* TRANSLATIONS_PLACEHOLDER */", f"const t = {json.dumps(js_translations)};")
            components.html(html_content, height=880, scrolling=True)
        except Exception as e:
            st.error(f"Error cargando el juego / Error loading game: {e}")
            
        st.markdown("---")
        st.markdown(
            "<p style='text-align: center; color: gray; font-size: 14px;'>"
            "Desarrollado y mantenido por <b>Alexander Acosta</b> "
            "(<a href='https://github.com/j-alexander-acosta' target='_blank' style='color: #1f77b4; text-decoration: none;'>@j-alexander-acosta</a>)"
            "</p>", 
            unsafe_allow_html=True
        )
        st.stop()
        
    if model_type == t["mkv_mode_business"]:
        st.header(t["biz_title"])
        st.markdown(t["biz_desc"])
        
        # --- CONFIGURACIÓN INICIAL ---
        biz_states = [t["game_brand_a"], t["game_brand_b"], t["game_brand_c"]]
        
        # Inicializar sesión para matriz de transición de negocios
        if "biz_tpm" not in st.session_state:
            init_vals = [
                [0.6, 0.3, 0.1],
                [0.3, 0.5, 0.2],
                [0.4, 0.3, 0.3]
            ]
            st.session_state.biz_tpm = pd.DataFrame(init_vals, index=biz_states, columns=biz_states)
            
        # 1. Panel de Entradas (Inputs)
        st.subheader(t["biz_init_vector"])
        col_v1, col_v2, col_v3 = st.columns(3)
        with col_v1:
            init_a = st.number_input(t["game_brand_a"], min_value=0, value=200, step=10, key="biz_v0_a")
        with col_v2:
            init_b = st.number_input(t["game_brand_b"], min_value=0, value=300, step=10, key="biz_v0_b")
        with col_v3:
            init_c = st.number_input(t["game_brand_c"], min_value=0, value=500, step=10, key="biz_v0_c")
            
        V0 = np.array([init_a, init_b, init_c], dtype=float)
        
        st.subheader(t["biz_tpm"])
        tpm_df = st.data_editor(
            st.session_state.biz_tpm,
            use_container_width=True,
            key="biz_tpm_editor",
            on_change=update_session_df,
            args=("biz_tpm", "biz_tpm_editor")
        )
        
        # Validación
        row_sums = tpm_df.sum(axis=1)
        invalid_rows = [r for r, s in zip(biz_states, row_sums) if not np.isclose(s, 1.0, atol=1e-4)]
        if invalid_rows:
            st.warning(t["mkv_sum_warning"] + ", ".join(invalid_rows))
            if st.button(t["mkv_normalize_btn"], key="biz_norm_btn"):
                normalized_vals = tpm_df.values.copy()
                for idx in range(len(normalized_vals)):
                    row_sum = np.sum(normalized_vals[idx])
                    if row_sum > 0:
                        normalized_vals[idx] = normalized_vals[idx] / row_sum
                    else:
                        normalized_vals[idx] = np.ones(3) / 3
                st.session_state.biz_tpm = pd.DataFrame(normalized_vals, index=biz_states, columns=biz_states)
                st.success(t["mkv_normalize_success"])
                st.rerun()
                
        P = tpm_df.values.copy()
        for idx in range(len(P)):
            r_sum = np.sum(P[idx])
            if r_sum > 0:
                P[idx] = P[idx] / r_sum
            else:
                P[idx] = np.ones(3) / 3
                
        # --- CÁLCULOS ---
        history_V = [V0]
        curr_V = V0.copy()
        for t_step in range(1, 11):
            curr_V = np.dot(curr_V, P)
            history_V.append(curr_V)
        history_V = np.array(history_V)
        
        M_powers = [P]
        curr_M = P.copy()
        for _ in range(2, 7):
            curr_M = np.dot(curr_M, P)
            M_powers.append(curr_M)
            
        # --- RENDER PANELS (Layout de 2 Columnas) ---
        col_left, col_right = st.columns([1, 1.2])
        
        # Columna Izquierda: Potencias de Matriz (Mk)
        with col_left:
            st.markdown(f"### 🔀 {t['biz_matrix_powers']}")
            for k in range(1, 7):
                k_mat = M_powers[k-1]
                st.markdown(f"**Matriz $M_{k}$**" + (f" (Matriz Original $M_1$)" if k == 1 else f" ($M_{k} = M_1^{k}$)"))
                
                # Tabla HTML compacta estilizada
                html_table = "<table style='width:100%; border-collapse: collapse; margin-bottom: 20px; font-size:13px; text-align:center; border: 1px solid rgba(128,128,128,0.2);'>"
                html_table += "<tr style='background-color:rgba(128,128,128,0.08); font-weight:bold;'>"
                html_table += "<th style='padding:6px; border:1px solid rgba(128,128,128,0.2);'></th>"
                for s in biz_states:
                    html_table += f"<th style='padding:6px; border:1px solid rgba(128,128,128,0.2);'>{s.split(' ')[1] if ' ' in s else s}</th>"
                html_table += "</tr>"
                
                for idx, row_name in enumerate(biz_states):
                    html_table += f"<tr>"
                    html_table += f"<td style='padding:6px; border:1px solid rgba(128,128,128,0.2); font-weight:bold; background-color:rgba(128,128,128,0.02);'>{row_name}</td>"
                    for val in k_mat[idx]:
                        html_table += f"<td style='padding:6px; border:1px solid rgba(128,128,128,0.2);'>{val:.3f}</td>"
                    html_table += "</tr>"
                html_table += "</table>"
                st.html(html_table)
                
        # Columna Derecha: Evolución y Estabilización (Vt)
        with col_right:
            st.markdown(f"### 📈 {t['biz_chart_title']}")
            
            # DataFrame para evolución
            evo_data = []
            for t_idx in range(11):
                vt = history_V[t_idx]
                evo_data.append({
                    t["biz_col_time"]: t_idx,
                    t["game_brand_a"]: round(vt[0]),
                    t["game_brand_b"]: round(vt[1]),
                    t["game_brand_c"]: round(vt[2])
                })
            
            evo_df = pd.DataFrame(evo_data)
            
            # Gráfico interactivo Plotly
            import plotly.graph_objects as plgo
            fig_evo = plgo.Figure()
            colors = ['#ef4444', '#3b82f6', '#f59e0b']
            brands = [t["game_brand_a"], t["game_brand_b"], t["game_brand_c"]]
            
            for idx, b_name in enumerate(brands):
                fig_evo.add_trace(plgo.Scatter(
                    x=evo_df[t["biz_col_time"]],
                    y=evo_df[b_name],
                    mode='lines+markers',
                    name=b_name,
                    line=dict(color=colors[idx], width=3),
                    marker=dict(size=8),
                    hovertemplate='Paso %{x}<br>' + b_name + ': %{y:.0f}<extra></extra>'
                ))
                
            fig_evo.update_layout(
                xaxis=dict(tickmode='linear', tick0=0, dtick=1),
                xaxis_title=t["biz_chart_xaxis"],
                yaxis_title=t["biz_chart_yaxis"],
                template='plotly_white',
                legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
                margin=dict(l=40, r=40, t=20, b=40)
            )
            st.plotly_chart(fig_evo, use_container_width=True)
            
            # Tabla de evolución Vt con fórmulas explicativas
            st.markdown("**Tabla de Evolución Temporal ($V_t$)**")
            
            html_v_table = "<table style='width:100%; border-collapse: collapse; font-size:13px; text-align:center; border: 1px solid rgba(128,128,128,0.2);'>"
            html_v_table += "<tr style='background-color:rgba(128,128,128,0.08); font-weight:bold;'>"
            html_v_table += f"<th style='padding:6px; border:1px solid rgba(128,128,128,0.2);'>{t['biz_col_time']}</th>"
            html_v_table += f"<th style='padding:6px; border:1px solid rgba(128,128,128,0.2); color:#ef4444;'>{t['game_brand_a']}</th>"
            html_v_table += f"<th style='padding:6px; border:1px solid rgba(128,128,128,0.2); color:#3b82f6;'>{t['game_brand_b']}</th>"
            html_v_table += f"<th style='padding:6px; border:1px solid rgba(128,128,128,0.2); color:#f59e0b;'>{t['game_brand_c']}</th>"
            html_v_table += f"<th style='padding:6px; border:1px solid rgba(128,128,128,0.2);'>Ecuación / Equation</th>"
            html_v_table += "</tr>"
            
            for t_idx in range(11):
                vt = history_V[t_idx]
                formula_lbl = ""
                if t_idx == 0:
                    formula_lbl = "V₀"
                elif t_idx == 1:
                    formula_lbl = "V₁ = V₀ * M₁"
                else:
                    formula_lbl = f"V_{t_idx} = V_{t_idx-1} * M₁"
                    
                html_v_table += f"<tr>"
                html_v_table += f"<td style='padding:6px; border:1px solid rgba(128,128,128,0.2); font-weight:bold;'>{t_idx}</td>"
                html_v_table += f"<td style='padding:6px; border:1px solid rgba(128,128,128,0.2);'>{vt[0]:.0f}</td>"
                html_v_table += f"<td style='padding:6px; border:1px solid rgba(128,128,128,0.2);'>{vt[1]:.0f}</td>"
                html_v_table += f"<td style='padding:6px; border:1px solid rgba(128,128,128,0.2);'>{vt[2]:.0f}</td>"
                html_v_table += f"<td style='padding:6px; border:1px solid rgba(128,128,128,0.2); font-style:italic; font-family:monospace;'>{formula_lbl}</td>"
                html_v_table += "</tr>"
            html_v_table += "</table>"
            st.html(html_v_table)
            
        st.markdown("---")
        st.markdown(
            "<p style='text-align: center; color: gray; font-size: 14px;'>"
            "Desarrollado y mantenido por <b>Alexander Acosta</b> "
            "(<a href='https://github.com/j-alexander-acosta' target='_blank' style='color: #1f77b4; text-decoration: none;'>@j-alexander-acosta</a>)"
            "</p>", 
            unsafe_allow_html=True
        )
        st.stop()
        
    states_input = st.sidebar.text_input(t["mkv_states_input"], value=t["mkv_default_states"], key="mkv_states_inp")
    states = [s.strip() for s in states_input.split(",") if s.strip()]
    
    if len(states) < 2:
        st.warning("Se requieren al menos 2 estados / At least 2 states are required.")
        st.stop()
        
    if model_type == t["mkv_mode_hmm"]:
        st.header(t["hmm_title"])
        st.markdown(t["hmm_desc"])
        
        # Observations input
        obs_input = st.sidebar.text_input(t["hmm_obs_input"], value=t["hmm_default_obs"], key="hmm_obs_inp")
        observations = [o.strip() for o in obs_input.split(",") if o.strip()]
        if len(observations) < 2:
            st.warning("Se requieren al menos 2 observaciones / At least 2 observations are required.")
            st.stop()
            
        # HMM matrices initialization
        if ("hmm_tpm" not in st.session_state 
                or st.session_state.get("hmm_prev_states") != states 
                or st.session_state.get("hmm_prev_obs") != observations):
            n_states = len(states)
            n_obs = len(observations)
            
            # Transition Matrix (TPM)
            if n_states == 3:
                init_vals_tpm = [
                    [0.4, 0.4, 0.2],
                    [0.2, 0.5, 0.3],
                    [0.2, 0.3, 0.5]
                ]
            else:
                init_vals_tpm = np.ones((n_states, n_states)) / n_states
                
            # Emission Matrix (EPM)
            if n_states == 3 and n_obs == 3:
                init_vals_epm = [
                    [0.3, 0.6, 0.1],  # Soleado -> Paraguas, Normal, Impermeable
                    [0.7, 0.2, 0.1],  # Nublado -> Paraguas, Normal, Impermeable
                    [0.5, 0.1, 0.4]   # Lluvioso -> Paraguas, Normal, Impermeable
                ]
            else:
                init_vals_epm = np.ones((n_states, n_obs)) / n_obs
                
            st.session_state.hmm_tpm = pd.DataFrame(init_vals_tpm, index=states, columns=states)
            st.session_state.hmm_epm = pd.DataFrame(init_vals_epm, index=states, columns=observations)
            st.session_state.hmm_prev_states = states
            st.session_state.hmm_prev_obs = observations
            
        if "hmm_init_probs" not in st.session_state or len(st.session_state.hmm_init_probs.columns) != len(states):
            if len(states) == 3:
                init_p_vals = [[0.4, 0.4, 0.2]]
            else:
                init_p_vals = [[1.0] + [0.0] * (len(states) - 1)]
            st.session_state.hmm_init_probs = pd.DataFrame(
                init_p_vals,
                columns=states,
                index=["Probabilidad / Probability"]
            )
            
        # Editors and tabs
        hmm_tabs = st.tabs([t["hmm_tab_params"], t["hmm_tab_evolution"], t["hmm_tab_path"]])
        
        with hmm_tabs[0]:
            st.subheader(t["mkv_tpm_title"])
            tpm_df = st.data_editor(
                st.session_state.hmm_tpm,
                use_container_width=True,
                key="hmm_tpm_editor",
                on_change=update_session_df,
                args=("hmm_tpm", "hmm_tpm_editor")
            )
            
            row_sums = tpm_df.sum(axis=1)
            invalid_rows = [r for r, s in zip(states, row_sums) if not np.isclose(s, 1.0, atol=1e-4)]
            if invalid_rows:
                st.warning(t["mkv_sum_warning"] + ", ".join(invalid_rows))
                if st.button(t["mkv_normalize_btn"], key="hmm_norm_tpm_btn"):
                    normalized_vals = tpm_df.values.copy()
                    for idx in range(len(normalized_vals)):
                        r_sum = np.sum(normalized_vals[idx])
                        if r_sum > 0:
                            normalized_vals[idx] /= r_sum
                        else:
                            normalized_vals[idx] = np.ones(len(states)) / len(states)
                    st.session_state.hmm_tpm = pd.DataFrame(normalized_vals, index=states, columns=states)
                    st.success(t["mkv_normalize_success"])
                    st.rerun()
                    
            st.subheader(t["hmm_epm_title"])
            epm_df = st.data_editor(
                st.session_state.hmm_epm,
                use_container_width=True,
                key="hmm_epm_editor",
                on_change=update_session_df,
                args=("hmm_epm", "hmm_epm_editor")
            )
            
            epm_row_sums = epm_df.sum(axis=1)
            invalid_epm_rows = [r for r, s in zip(states, epm_row_sums) if not np.isclose(s, 1.0, atol=1e-4)]
            if invalid_epm_rows:
                st.warning(t["mkv_sum_warning"] + ", ".join(invalid_epm_rows))
                if st.button(t["mkv_normalize_btn"], key="hmm_norm_epm_btn"):
                    normalized_vals = epm_df.values.copy()
                    for idx in range(len(normalized_vals)):
                        r_sum = np.sum(normalized_vals[idx])
                        if r_sum > 0:
                            normalized_vals[idx] /= r_sum
                        else:
                            normalized_vals[idx] = np.ones(len(observations)) / len(observations)
                    st.session_state.hmm_epm = pd.DataFrame(normalized_vals, index=states, columns=observations)
                    st.success(t["mkv_normalize_success"])
                    st.rerun()
                    
            st.subheader(t["mkv_init_probs"])
            init_probs_df = st.data_editor(
                st.session_state.hmm_init_probs,
                use_container_width=True,
                key="hmm_init_editor",
                on_change=update_session_df,
                args=("hmm_init_probs", "hmm_init_editor")
            )
            
            pi_0 = init_probs_df.values[0].copy()
            pi_0_sum = np.sum(pi_0)
            if pi_0_sum > 0:
                pi_0 = pi_0 / pi_0_sum
            else:
                pi_0 = np.zeros(len(states))
                pi_0[0] = 1.0
                
        with hmm_tabs[1]:
            st.subheader(t["mkv_evolution_chart_title"])
            N_steps = st.slider(t["mkv_steps_slider"], min_value=1, max_value=50, value=20, key="hmm_steps_sl")
            
            P = tpm_df.values.copy()
            for idx in range(len(P)):
                r_sum = np.sum(P[idx])
                if r_sum > 0: P[idx] /= r_sum
                else: P[idx] = np.ones(len(states)) / len(states)
                
            E = epm_df.values.copy()
            for idx in range(len(E)):
                r_sum = np.sum(E[idx])
                if r_sum > 0: E[idx] /= r_sum
                else: E[idx] = np.ones(len(observations)) / len(observations)
                
            hidden_history = [pi_0]
            obs_history = [np.dot(pi_0, E)]
            curr_pi = pi_0.copy()
            for _ in range(N_steps):
                curr_pi = np.dot(curr_pi, P)
                hidden_history.append(curr_pi)
                obs_history.append(np.dot(curr_pi, E))
                
            hidden_history = np.array(hidden_history)
            obs_history = np.array(obs_history)
            t_axis = np.arange(N_steps + 1)
            
            col_evo1, col_evo2 = st.columns(2)
            colors = ['#1f77b4', '#ff7f0e', '#2ca02c', '#d62728', '#9467bd', '#8c564b', '#e377c2', '#7f7f7f', '#bcbd22', '#17becf']
            
            with col_evo1:
                st.markdown("**Estados Ocultos (Hidden States)**")
                fig_hid = plgo.Figure()
                for idx, name in enumerate(states):
                    color = colors[idx % len(colors)]
                    fig_hid.add_trace(plgo.Scatter(
                        x=t_axis, y=hidden_history[:, idx], mode='lines+markers', name=name,
                        line=dict(color=color, width=2),
                        hovertemplate='Paso %{x}<br>Prob: %{y:.4f}<extra></extra>'
                    ))
                fig_hid.update_layout(
                    xaxis_title="Paso / Step (t)",
                    yaxis_title="Probabilidad / Probability",
                    template='plotly_white',
                    legend=dict(yanchor="top", y=0.99, xanchor="right", x=0.99)
                )
                st.plotly_chart(fig_hid, use_container_width=True)
                
            with col_evo2:
                st.markdown("**Observaciones (Observations)**")
                fig_obs = plgo.Figure()
                for idx, name in enumerate(observations):
                    color = colors[(idx + len(states)) % len(colors)]
                    fig_obs.add_trace(plgo.Scatter(
                        x=t_axis, y=obs_history[:, idx], mode='lines+markers', name=name,
                        line=dict(color=color, width=2),
                        hovertemplate='Paso %{x}<br>Prob: %{y:.4f}<extra></extra>'
                    ))
                fig_obs.update_layout(
                    xaxis_title="Paso / Step (t)",
                    yaxis_title="Probabilidad / Probability",
                    template='plotly_white',
                    legend=dict(yanchor="top", y=0.99, xanchor="right", x=0.99)
                )
                st.plotly_chart(fig_obs, use_container_width=True)
                
        with hmm_tabs[2]:
            st.subheader(t["hmm_path_title"])
            st.markdown(t["hmm_path_desc"])
            
            path_len = st.slider(t["mkv_seq_length"], min_value=2, max_value=10, value=4, key="hmm_path_len_sl")
            
            st.markdown(f"**{t['hmm_path_hidden_seq']}**")
            cols_states = st.columns(path_len)
            selected_states = []
            
            default_states_es = ["Nublado", "Soleado", "Lluvioso", "Nublado"]
            default_states_en = ["Cloudy", "Sunny", "Rainy", "Cloudy"]
            default_obs_es = ["Paraguas", "Normal", "Paraguas", "Impermeable"]
            default_obs_en = ["Umbrella", "Normal", "Umbrella", "Raincoat"]
            
            for i in range(path_len):
                default_state_index = 0
                if path_len == 4 and len(states) >= 3:
                    target_state = default_states_es[i] if st.session_state.lang == 'es' else default_states_en[i]
                    matching = [idx for idx, s in enumerate(states) if target_state.lower() in s.lower()]
                    if matching:
                        default_state_index = matching[0]
                    else:
                        default_state_index = min(i % len(states), len(states) - 1)
                else:
                    default_state_index = min(i % len(states), len(states) - 1)
                    
                st_val = cols_states[i].selectbox(
                    f"{t['mkv_seq_step'].format(i+1)} (E{i+1})",
                    states,
                    index=default_state_index,
                    key=f"hmm_state_{i}"
                )
                selected_states.append(st_val)
                
            st.markdown(f"**{t['hmm_path_obs_seq']}**")
            cols_obs = st.columns(path_len)
            selected_obs = []
            for i in range(path_len):
                default_obs_index = 0
                if path_len == 4 and len(observations) >= 3:
                    target_obs = default_obs_es[i] if st.session_state.lang == 'es' else default_obs_en[i]
                    matching_obs = [idx for idx, o in enumerate(observations) if target_obs.lower() in o.lower()]
                    if matching_obs:
                        default_obs_index = matching_obs[0]
                    else:
                        default_obs_index = min(i % len(observations), len(observations) - 1)
                else:
                    default_obs_index = min(i % len(observations), len(observations) - 1)
                    
                ob_val = cols_obs[i].selectbox(
                    f"{t['mkv_seq_step'].format(i+1)} (O{i+1})",
                    observations,
                    index=default_obs_index,
                    key=f"hmm_obs_{i}"
                )
                selected_obs.append(ob_val)
                
            P_trans = tpm_df.values.copy()
            P_em = epm_df.values.copy()
            
            for idx in range(len(P_trans)):
                r_sum = np.sum(P_trans[idx])
                if r_sum > 0: P_trans[idx] /= r_sum
                else: P_trans[idx] = np.ones(len(states)) / len(states)
            for idx in range(len(P_em)):
                r_sum = np.sum(P_em[idx])
                if r_sum > 0: P_em[idx] /= r_sum
                else: P_em[idx] = np.ones(len(observations)) / len(observations)
                
            first_state = selected_states[0]
            first_state_idx = states.index(first_state)
            init_p = pi_0[first_state_idx]
            
            trans_steps = []
            trans_probs = []
            for i in range(path_len - 1):
                from_s = selected_states[i]
                to_s = selected_states[i+1]
                from_idx = states.index(from_s)
                to_idx = states.index(to_s)
                prob = P_trans[from_idx, to_idx]
                trans_steps.append((from_s, to_s))
                trans_probs.append(prob)
                
            total_trans_prob = init_p * np.prod(trans_probs) if trans_probs else init_p
            
            em_steps = []
            em_probs = []
            for i in range(path_len):
                st_val = selected_states[i]
                ob_val = selected_obs[i]
                st_idx = states.index(st_val)
                ob_idx = observations.index(ob_val)
                prob = P_em[st_idx, ob_idx]
                em_steps.append((st_val, ob_val))
                em_probs.append(prob)
                
            total_em_prob = np.prod(em_probs) if em_probs else 1.0
            joint_prob = total_trans_prob * total_em_prob
            
            def get_state_emoji(state_name):
                name = state_name.lower()
                if "soleado" in name or "sunny" in name:
                    return "☀️"
                elif "nublado" in name or "cloudy" in name:
                    return "☁️"
                elif "lluvioso" in name or "rainy" in name:
                    return "🌧️"
                return "🔮"

            def get_obs_emoji(obs_name):
                name = obs_name.lower()
                if "paraguas" in name or "umbrella" in name:
                    return "☔"
                elif "normal" in name:
                    return "👕"
                elif "impermeable" in name or "raincoat" in name:
                    return "🧥"
                return "📦"
                
            def get_state_color(state_name):
                name = state_name.lower()
                if "soleado" in name or "sunny" in name:
                    return "linear-gradient(135deg, #ffb300, #ff6f00)"
                elif "nublado" in name or "cloudy" in name:
                    return "linear-gradient(135deg, #90a4ae, #455a64)"
                elif "lluvioso" in name or "rainy" in name:
                    return "linear-gradient(135deg, #4fc3f7, #0288d1)"
                else:
                    colors = [
                        "linear-gradient(135deg, #26a69a, #00695c)",
                        "linear-gradient(135deg, #ab47bc, #6a1b9a)",
                        "linear-gradient(135deg, #ec407a, #ad1457)",
                        "linear-gradient(135deg, #78909c, #37474f)",
                        "linear-gradient(135deg, #5c6bc0, #283593)"
                    ]
                    idx = sum(ord(c) for c in state_name) % len(colors)
                    return colors[idx]
                    
            st.markdown("### 🗺️ " + t["hmm_diagram_title"])
            
            html_content = """
            <div style="display: flex; align-items: center; justify-content: center; flex-wrap: wrap; gap: 15px; padding: 25px; background-color: #f8f9fa; border-radius: 12px; border: 1px solid #e9ecef; margin-bottom: 25px; box-shadow: inset 0 1px 3px rgba(0,0,0,0.05);">
            """
            
            for idx in range(path_len):
                st_val = selected_states[idx]
                ob_val = selected_obs[idx]
                st_emoji = get_state_emoji(st_val)
                ob_emoji = get_obs_emoji(ob_val)
                st_color = get_state_color(st_val)
                em_prob = em_probs[idx]
                
                html_content += f"""
                <div style="display: flex; flex-direction: column; align-items: center; gap: 8px;">
                    <!-- Observation Box (Top) -->
                    <div style="background: linear-gradient(135deg, #17a2b8, #117a8b); color: white; padding: 10px 15px; border-radius: 8px; font-weight: bold; text-align: center; box-shadow: 0 4px 6px rgba(0,0,0,0.1); min-width: 100px;">
                        <span style="font-size: 1.25em; display: block;">{ob_emoji}</span>
                        <span style="font-size: 0.95em; display: block; margin-top: 2px;">{ob_val}</span>
                        <span style="font-size: 0.7em; font-weight: normal; opacity: 0.85; display: block; margin-top: 2px;">O{idx+1}</span>
                    </div>
                    
                    <!-- Vertical Emission Arrow -->
                    <div style="display: flex; flex-direction: column; align-items: center; justify-content: center; margin: 4px 0;">
                        <span style="font-weight: 800; color: #dc3545; font-size: 0.85em; background-color: #f8d7da; padding: 1px 5px; border-radius: 8px; border: 1px dashed #f5c6cb; margin-bottom: 2px;">{em_prob:.4f}</span>
                        <span style="color: #dc3545; font-size: 1.3em; line-height: 1;">▲</span>
                    </div>
                    
                    <!-- Hidden State Box (Bottom) -->
                    <div style="background: {st_color}; color: white; padding: 10px 15px; border-radius: 8px; font-weight: bold; text-align: center; box-shadow: 0 4px 6px rgba(0,0,0,0.1); min-width: 100px;">
                        <span style="font-size: 1.25em; display: block;">{st_emoji}</span>
                        <span style="font-size: 0.95em; display: block; margin-top: 2px;">{st_val}</span>
                        <span style="font-size: 0.7em; font-weight: normal; opacity: 0.85; display: block; margin-top: 2px;">E{idx+1}</span>
                """
                
                if idx == 0:
                    html_content += f"""
                        <div style="border-top: 1px solid rgba(255,255,255,0.3); font-size: 0.7em; font-weight: normal; margin-top: 6px; padding-top: 4px;">P(init) = {init_p:.4f}</div>
                    """
                html_content += "</div></div>"
                
                if idx < path_len - 1:
                    trans_prob = trans_probs[idx]
                    html_content += f"""
                    <div style="display: flex; flex-direction: column; align-items: center; justify-content: center; min-width: 50px; align-self: flex-end; margin-bottom: 20px;">
                        <span style="font-weight: 800; color: #1f77b4; font-size: 0.85em; background-color: #e3f2fd; padding: 1px 5px; border-radius: 8px; border: 1px dashed #90caf9; margin-bottom: 2px;">{trans_prob:.4f}</span>
                        <span style="font-size: 1.4em; color: #90a4ae; line-height: 1;">➔</span>
                    </div>
                    """
                    
            html_content += "</div>"
            st.html(html_content)
            
            col_calc1, col_calc2 = st.columns(2)
            
            with col_calc1:
                st.markdown(f"### 🧮 {t['hmm_path_calc']}")
                
                st.markdown(f"#### **{t['hmm_path_emission_formula']}**")
                em_terms_latex = " \\cdot ".join([f"P(O_{{{i+1}}} \\mid E_{{{i+1}}})" for i in range(path_len)])
                em_vals_latex = " \\cdot ".join([f"{prob:.4f}" for prob in em_probs])
                
                st.markdown(f"$$\\text{{Emisiones}} = {em_terms_latex}$$")
                st.markdown(f"$$\\text{{Emisiones}} = {em_vals_latex} = \\mathbf{{{total_em_prob:.6f}}}$$")
                
                st.markdown(f"#### **{t['hmm_path_transition_formula']}**")
                tr_terms_latex = "P(E_1) " + "".join([f" \\cdot P(E_{{{i+2}}} \\mid E_{{{i+1}}})" for i in range(path_len - 1)])
                tr_vals_latex = f"{init_p:.4f}" + "".join([f" \\cdot {prob:.4f}" for prob in trans_probs])
                
                st.markdown(f"$$\\text{{Transiciones}} = {tr_terms_latex}$$")
                st.markdown(f"$$\\text{{Transiciones}} = {tr_vals_latex} = \\mathbf{{{total_trans_prob:.6f}}}$$")
                
            with col_calc2:
                st.markdown(f"### 📊 {t['hmm_path_joint_formula']}")
                st.markdown("$$\\text{Probabilidad Conjunta } P(O, E) = \\text{Emisiones} \\times \\text{Transiciones}$$")
                st.markdown(f"$$P(O, E) = {total_em_prob:.6f} \\times {total_trans_prob:.6f}$$")
                st.markdown(f"$$P(O, E) = \\mathbf{{{joint_prob:.8f}}}$$")
                
                st.success(f"**{t['hmm_path_result']}**\n### {joint_prob:.8f}")
                
        st.markdown("---")
        st.markdown(
            "<p style='text-align: center; color: gray; font-size: 14px;'>"
            "Desarrollado y mantenido por <b>Alexander Acosta</b> "
            "(<a href='https://github.com/j-alexander-acosta' target='_blank' style='color: #1f77b4; text-decoration: none;'>@j-alexander-acosta</a>)"
            "</p>", 
            unsafe_allow_html=True
        )
        st.stop()
        
    st.header(t["mkv_title"])
    st.markdown(t["mkv_desc"])
        
    if "mkv_tpm" not in st.session_state or st.session_state.get("mkv_prev_states") != states:
        n_states = len(states)
        if n_states == 3:
            init_vals = [
                [0.4, 0.4, 0.2],
                [0.2, 0.5, 0.3],
                [0.2, 0.3, 0.5]
            ]
        else:
            init_vals = np.ones((n_states, n_states)) / n_states
        st.session_state.mkv_tpm = pd.DataFrame(init_vals, index=states, columns=states)
        st.session_state.mkv_prev_states = states
        
    if "mkv_init_probs" not in st.session_state or len(st.session_state.mkv_init_probs.columns) != len(states):
        if len(states) == 3:
            init_p_vals = [[0.4, 0.4, 0.2]]
        else:
            init_p_vals = [[1.0] + [0.0] * (len(states) - 1)]
        st.session_state.mkv_init_probs = pd.DataFrame(
            init_p_vals,
            columns=states,
            index=["Probabilidad / Probability"]
        )

    st.subheader(t["mkv_tpm_title"])
    tpm_df = st.data_editor(
        st.session_state.mkv_tpm,
        use_container_width=True,
        key="mkv_tpm_editor",
        on_change=update_session_df,
        args=("mkv_tpm", "mkv_tpm_editor")
    )
    
    row_sums = tpm_df.sum(axis=1)
    invalid_rows = [r for r, s in zip(states, row_sums) if not np.isclose(s, 1.0, atol=1e-4)]
    if invalid_rows:
        st.warning(t["mkv_sum_warning"] + ", ".join(invalid_rows))
        if st.button(t["mkv_normalize_btn"], key="mkv_norm_btn"):
            normalized_vals = tpm_df.values.copy()
            for idx in range(len(normalized_vals)):
                row_sum = np.sum(normalized_vals[idx])
                if row_sum > 0:
                    normalized_vals[idx] = normalized_vals[idx] / row_sum
                else:
                    normalized_vals[idx] = np.ones(len(states)) / len(states)
            st.session_state.mkv_tpm = pd.DataFrame(normalized_vals, index=states, columns=states)
            st.success(t["mkv_normalize_success"])
            st.rerun()
            
    P = tpm_df.values.copy()
    for idx in range(len(P)):
        row_sum = np.sum(P[idx])
        if row_sum > 0:
            P[idx] = P[idx] / row_sum
        else:
            P[idx] = np.ones(len(states)) / len(states)
            
    st.subheader(t["mkv_init_probs"])
    init_probs_df = st.data_editor(
        st.session_state.mkv_init_probs,
        use_container_width=True,
        key="mkv_init_editor",
        on_change=update_session_df,
        args=("mkv_init_probs", "mkv_init_editor")
    )
    
    pi_0 = init_probs_df.values[0].copy()
    pi_0_sum = np.sum(pi_0)
    if pi_0_sum > 0:
        pi_0 = pi_0 / pi_0_sum
    else:
        pi_0 = np.zeros(len(states))
        pi_0[0] = 1.0

    mkv_tabs = st.tabs([t["mkv_tab_evolution"], t["mkv_tab_steady"], t["mkv_tab_simulation"], t["mkv_tab_sequence"]])
    
    with mkv_tabs[0]:
        st.subheader(t["mkv_evolution_chart_title"])
        N_steps = st.slider(t["mkv_steps_slider"], min_value=1, max_value=50, value=20, key="mkv_steps_sl")
        
        history = [pi_0]
        curr_pi = pi_0.copy()
        for _ in range(N_steps):
            curr_pi = np.dot(curr_pi, P)
            history.append(curr_pi)
            
        history = np.array(history)
        
        fig_evo = plgo.Figure()
        t_axis = np.arange(N_steps + 1)
        colors = ['#1f77b4', '#ff7f0e', '#2ca02c', '#d62728', '#9467bd', '#8c564b', '#e377c2', '#7f7f7f', '#bcbd22', '#17becf']
        
        for idx, name in enumerate(states):
            color = colors[idx % len(colors)]
            fig_evo.add_trace(plgo.Scatter(
                x=t_axis, y=history[:, idx], mode='lines+markers', name=name,
                line=dict(color=color, width=2),
                hovertemplate='Paso / Step %{x}<br>Prob: %{y:.4f}<extra></extra>'
            ))
            
        fig_evo.update_layout(
            title=t["mkv_evolution_chart_title"],
            xaxis_title="Paso / Step (t)",
            yaxis_title="Probabilidad / Probability",
            hovermode='closest',
            template='plotly_white',
            legend=dict(yanchor="top", y=0.99, xanchor="right", x=0.99)
        )
        st.plotly_chart(fig_evo, use_container_width=True)
        
        st.subheader(t["mkv_tpm_power_title"].replace("$P^N$", f"$P^{{{N_steps}}}$"))
        P_N = np.linalg.matrix_power(P, N_steps)
        P_N_df = pd.DataFrame(P_N, index=states, columns=states)
        st.dataframe(
            P_N_df,
            column_config={c: st.column_config.NumberColumn(c, format="%.6f") for c in states},
            use_container_width=True
        )
        
    with mkv_tabs[1]:
        st.subheader(t["mkv_steady_title"])
        st.markdown(t["mkv_steady_desc"])
        
        n_st = P.shape[0]
        A_st = P.T - np.eye(n_st)
        A_st[-1, :] = 1
        b_st = np.zeros(n_st)
        b_st[-1] = 1
        
        try:
            steady_dist = np.linalg.solve(A_st, b_st)
            steady_dist = np.clip(steady_dist, 0.0, 1.0)
            steady_dist = steady_dist / np.sum(steady_dist)
        except np.linalg.LinAlgError:
            P_pow = np.linalg.matrix_power(P, 100)
            steady_dist = P_pow[0, :]
            
        steady_df = pd.DataFrame({
            "Estado / State": states,
            "Probabilidad Estacionaria / Stationary Probability": steady_dist
        })
        
        col_st1, col_st2 = st.columns([1, 1])
        with col_st1:
            st.dataframe(
                steady_df,
                column_config={
                    "Estado / State": st.column_config.TextColumn("Estado / State"),
                    "Probabilidad Estacionaria / Stationary Probability": st.column_config.NumberColumn("Probabilidad Estacionaria / Stationary Probability", format="%.6f")
                },
                use_container_width=True
            )
        with col_st2:
            fig_steady = plgo.Figure()
            fig_steady.add_trace(plgo.Bar(
                x=states, y=steady_dist,
                marker_color='#2ca02c',
                hovertemplate='Estado: %{x}<br>Prob: %{y:.6f}<extra></extra>'
            ))
            fig_steady.update_layout(
                title=t["mkv_steady_title"],
                xaxis_title="Estado / State",
                yaxis_title="Probabilidad / Probability",
                template='plotly_white'
            )
            st.plotly_chart(fig_steady, use_container_width=True)
            
    with mkv_tabs[2]:
        st.subheader(t["mkv_sim_title"])
        st.markdown(t["mkv_sim_desc"])
        
        col_sim1, col_sim2 = st.columns(2)
        with col_sim1:
            sim_start = st.selectbox(t["mkv_sim_start_state"], states, key="mkv_sim_start")
        with col_sim2:
            sim_steps = st.slider(t["mkv_sim_steps"], min_value=10, max_value=1000, value=200, step=10, key="mkv_sim_steps_sl")
            
        if st.button(t["mkv_sim_run_btn"], key="mkv_run_btn"):
            current_state_idx = states.index(sim_start)
            visited_indices = [current_state_idx]
            for _ in range(sim_steps):
                next_idx = np.random.choice(len(states), p=P[current_state_idx])
                visited_indices.append(next_idx)
                current_state_idx = next_idx
            visited_states = [states[idx] for idx in visited_indices]
            
            st.session_state.mkv_sim_path = visited_states
            st.session_state.mkv_sim_indices = visited_indices
            
        if "mkv_sim_path" in st.session_state:
            visited_states = st.session_state.mkv_sim_path
            visited_indices = st.session_state.mkv_sim_indices
            
            st.markdown(f"**{t['mkv_sim_path_title']}**")
            path_str = " ➔ ".join(visited_states[:60]) + (" ➔ ..." if len(visited_states) > 60 else "")
            st.code(path_str, language="")
            
            counts = np.bincount(visited_indices, minlength=len(states))
            empirical_dist = counts / len(visited_indices)
            
            fig_comp = plgo.Figure()
            fig_comp.add_trace(plgo.Bar(
                x=states, y=empirical_dist, name=t["mkv_sim_empirical"],
                marker_color='#1f77b4',
                hovertemplate='Estado: %{x}<br>Empírica: %{y:.4f}<extra></extra>'
            ))
            fig_comp.add_trace(plgo.Bar(
                x=states, y=steady_dist, name=t["mkv_sim_theoretical"],
                marker_color='#2ca02c',
                hovertemplate='Estado: %{x}<br>Teórica: %{y:.4f}<extra></extra>'
            ))
            fig_comp.update_layout(
                title=t["mkv_sim_dist_title"],
                barmode='group',
                xaxis_title="Estado / State",
                yaxis_title="Proporción / Proportion",
                template='plotly_white'
            )
            st.plotly_chart(fig_comp, use_container_width=True)
            
    with mkv_tabs[3]:
        st.subheader(t["mkv_seq_title"])
        st.markdown(t["mkv_seq_desc"])
        
        # Sequence length slider
        seq_len = st.slider(t["mkv_seq_length"], min_value=2, max_value=10, value=4, key="mkv_seq_len_sl")
        
        # Display selectboxes in columns to choose the state at each step
        st.write("")
        cols = st.columns(seq_len)
        selected_seq = []
        for i in range(seq_len):
            # Try to select default states matching the slide if available
            default_index = 0
            if seq_len == 4 and len(states) >= 3:
                # Slide sequence: Cloudy (Nublado), Sunny (Soleado), Cloudy (Nublado), Rainy (Lluvioso)
                slide_seq_es = ["Nublado", "Soleado", "Nublado", "Lluvioso"]
                slide_seq_en = ["Cloudy", "Sunny", "Cloudy", "Rainy"]
                
                target_state = slide_seq_es[i] if st.session_state.lang == 'es' else slide_seq_en[i]
                # Find if target_state is in states
                matching_indices = [idx for idx, s in enumerate(states) if target_state.lower() in s.lower()]
                if matching_indices:
                    default_index = matching_indices[0]
                else:
                    default_index = min(i % len(states), len(states) - 1)
            else:
                default_index = min(i % len(states), len(states) - 1)
                
            state_val = cols[i].selectbox(
                f"{t['mkv_seq_step'].format(i+1)} (E{i+1})",
                states,
                index=default_index,
                key=f"mkv_seq_step_val_{i}"
            )
            selected_seq.append(state_val)
            
        # Perform calculations
        # 1. Initial Probability P(E1)
        first_state = selected_seq[0]
        first_state_idx = states.index(first_state)
        # get initial probability from the vector
        p_init = pi_0[first_state_idx]
        
        # 2. Transition Probabilities
        transitions = []
        trans_probs = []
        for i in range(seq_len - 1):
            from_state = selected_seq[i]
            to_state = selected_seq[i+1]
            from_idx = states.index(from_state)
            to_idx = states.index(to_state)
            prob = P[from_idx, to_idx]
            transitions.append((from_state, to_state))
            trans_probs.append(prob)
            
        # 3. Total Probabilities
        # Conditional probability: P(E2, E3, ... | E1) = product of transitions
        cond_prob = np.prod(trans_probs) if trans_probs else 1.0
        # Joint probability: P(E1, E2, E3, ...) = P(E1) * P(E2 | E1) * ...
        joint_prob = p_init * cond_prob
        
        def get_state_color(state_name):
            name = state_name.lower()
            if "soleado" in name or "sunny" in name:
                return "linear-gradient(135deg, #ffb300, #ff6f00)"  # Warm amber/orange
            elif "nublado" in name or "cloudy" in name:
                return "linear-gradient(135deg, #90a4ae, #455a64)"  # Cool blue-grey
            elif "lluvioso" in name or "rainy" in name:
                return "linear-gradient(135deg, #4fc3f7, #0288d1)"  # Rainy blue
            else:
                colors = [
                    "linear-gradient(135deg, #26a69a, #00695c)", # Teal
                    "linear-gradient(135deg, #ab47bc, #6a1b9a)", # Purple
                    "linear-gradient(135deg, #ec407a, #ad1457)", # Pink
                    "linear-gradient(135deg, #78909c, #37474f)", # Grey
                    "linear-gradient(135deg, #5c6bc0, #283593)"  # Indigo
                ]
                idx = sum(ord(c) for c in state_name) % len(colors)
                return colors[idx]
                
        # Visual Sequence Diagram
        st.markdown("### 🗺️ " + t["mkv_seq_diagram_title"])
        
        # Build HTML for sequence diagram
        html_content = """
        <div style="display: flex; align-items: center; justify-content: center; flex-wrap: wrap; gap: 15px; padding: 25px; background-color: #f8f9fa; border-radius: 12px; border: 1px solid #e9ecef; margin-bottom: 25px; box-shadow: inset 0 1px 3px rgba(0,0,0,0.05);">
        """
        
        for idx, state in enumerate(selected_seq):
            color = get_state_color(state)
            # Add state box
            html_content += f"""
            <div style="background: {color}; color: white; padding: 12px 20px; border-radius: 10px; font-weight: bold; text-align: center; box-shadow: 0 4px 8px rgba(0,0,0,0.1); min-width: 100px;">
                <span style="font-size: 1.1em; display: block;">{state}</span>
                <span style="font-size: 0.8em; font-weight: normal; opacity: 0.9; margin-top: 4px; display: block;">E{idx+1}</span>
            """
            if idx == 0:
                html_content += f"""
                <div style="border-top: 1px solid rgba(255,255,255,0.3); font-size: 0.75em; font-weight: normal; margin-top: 6px; padding-top: 4px;">P(E1) = {p_init:.4f}</div>
                """
            html_content += "</div>"
            
            # Add transition arrow if not the last state
            if idx < seq_len - 1:
                prob_val = trans_probs[idx]
                html_content += f"""
                <div style="display: flex; flex-direction: column; align-items: center; justify-content: center; min-width: 60px;">
                    <span style="font-weight: 800; color: #1f77b4; font-size: 1.05em; background-color: #e3f2fd; padding: 2px 8px; border-radius: 12px; border: 1px dashed #90caf9; margin-bottom: 2px;">{prob_val:.4f}</span>
                    <span style="font-size: 1.8em; color: #90a4ae; line-height: 1;">➔</span>
                </div>
                """
                
        html_content += "</div>"
        st.html(html_content)
        
        # Details and formulas
        col_calc1, col_calc2 = st.columns(2)
        
        with col_calc1:
            st.markdown(f"### 🧮 {t['mkv_seq_calc']}")
            
            # Step by step details in a nice list
            st.markdown(f"1. **{t['mkv_seq_init_prob']} $E_1$ ({first_state}):**")
            st.markdown(f"   $$P(E_1) = {p_init:.4f}$$")
            
            st.markdown(f"2. **{t['mkv_seq_trans_prob']}s:**")
            for idx, (from_s, to_s) in enumerate(transitions):
                st.markdown(f"   *   Paso {idx+1} ➔ {idx+2} ($E_{idx+1} \\rightarrow E_{idx+2}$):")
                st.markdown(f"       $$P(E_{idx+2} \\mid E_{idx+1}) = P(\\text{{{to_s}}} \\mid \\text{{{from_s}}}) = {trans_probs[idx]:.4f}$$")
                
        with col_calc2:
            st.markdown("### 📊 Fórmulas y Resultados")
            
            # Conditional Probability LaTeX and Value
            st.markdown(f"#### 1. {t['mkv_seq_formula_cond']}")
            # Formula term: P(E2|E1) * P(E3|E2) * ...
            cond_terms_latex = " \\cdot ".join([f"P(E_{{{i+2}}} \\mid E_{{{i+1}}})" for i in range(seq_len - 1)])
            cond_vals_latex = " \\cdot ".join([f"{val:.4f}" for val in trans_probs])
            
            st.markdown(f"$$P(E_2, \\dots, E_{{{seq_len}}} \\mid E_1) = {cond_terms_latex}$$")
            st.markdown(f"$$P(E_2, \\dots, E_{{{seq_len}}} \\mid E_1) = {cond_vals_latex} = \\mathbf{{{cond_prob:.6f}}}$$")
            
            # Joint Probability LaTeX and Value
            st.markdown(f"#### 2. {t['mkv_seq_formula_joint']}")
            joint_terms_latex = f"P(E_1) \\cdot {cond_terms_latex}"
            joint_vals_latex = f"{p_init:.4f} \\cdot {cond_vals_latex}"
            
            st.markdown(f"$$P(E_1, \\dots, E_{{{seq_len}}}) = {joint_terms_latex}$$")
            st.markdown(f"$$P(E_1, \\dots, E_{{{seq_len}}}) = {joint_vals_latex} = \\mathbf{{{joint_prob:.6f}}}$$")
            
            # Highlighted result cards
            st.info(f"**{t['mkv_seq_result_cond']}:**\n### {cond_prob:.6f}")
            st.success(f"**{t['mkv_seq_result_joint']}:**\n### {joint_prob:.6f}")
            
    st.markdown("---")
    st.markdown(
        "<p style='text-align: center; color: gray; font-size: 14px;'>"
        "Desarrollado y mantenido por <b>Alexander Acosta</b> "
        "(<a href='https://github.com/j-alexander-acosta' target='_blank' style='color: #1f77b4; text-decoration: none;'>@j-alexander-acosta</a>)"
        "</p>", 
        unsafe_allow_html=True
    )
    st.stop()

if is_seismic_bssa:
    st.header(t["bssa_title"])
    st.markdown(t["bssa_desc"])
    
    st.sidebar.markdown(f"### {t['sb_header_data']}")
    preset = st.sidebar.selectbox(t["bssa_preset_select"], [
        t["bssa_preset_chile"],
        t["bssa_preset_nz77"],
        t["bssa_preset_sc3"],
        t["bssa_preset_iside"],
        t["bssa_preset_manual"]
    ], key="bssa_preset_sel")
    
    def generate_exact_cov_data(mean, cov, n, seed=42):
        np.random.seed(seed)
        Z = np.random.normal(0, 1, (n, 2))
        Z = Z - np.mean(Z, axis=0)
        S_Z = np.cov(Z, rowvar=False)
        L_Z = np.linalg.cholesky(S_Z)
        Z_white = np.dot(Z, np.linalg.inv(L_Z).T)
        L = np.linalg.cholesky(cov)
        X_final = np.dot(Z_white, L.T) + mean
        return pd.DataFrame({
            "X": np.round(X_final[:, 0], 6),
            "Y": np.round(X_final[:, 1], 6)
        })
        
    bssa_df = None
    
    if preset == t["bssa_preset_nz77"]:
        cov_nz77 = [[0.109333, 0.123875], [0.123875, 0.170000]]
        bssa_df = generate_exact_cov_data(np.array([5.0, 5.0]), np.array(cov_nz77), 200, seed=42)
    elif preset == t["bssa_preset_sc3"]:
        cov_sc3 = [[0.182967, 0.173819], [0.173819, 0.209757]]
        bssa_df = generate_exact_cov_data(np.array([5.0, 5.0]), np.array(cov_sc3), 213, seed=42)
    elif preset == t["bssa_preset_iside"]:
        cov_iside = [[0.266100, 0.245344], [0.245344, 0.257344]]
        bssa_df = generate_exact_cov_data(np.array([5.0, 5.0]), np.array(cov_iside), 599, seed=42)
    elif preset == t["bssa_preset_chile"]:
        cov_chile = [[0.135600, 0.118900], [0.118900, 0.145000]]
        bssa_df = generate_exact_cov_data(np.array([5.0, 5.0]), np.array(cov_chile), 350, seed=42)
    else:
        if "bssa_manual_df" not in st.session_state:
            cov_nz77 = [[0.109333, 0.123875], [0.123875, 0.170000]]
            st.session_state.bssa_manual_df = generate_exact_cov_data(np.array([5.0, 5.0]), np.array(cov_nz77), 200, seed=42)
        
        st.markdown(f"### {t['bssa_data_editor_title']}")
        bssa_df = st.data_editor(
            st.session_state.bssa_manual_df,
            num_rows="dynamic",
            use_container_width=True,
            column_config={
                "X": st.column_config.NumberColumn("X (ML)", format="%.4f"),
                "Y": st.column_config.NumberColumn("Y (Mw)", format="%.4f")
            },
            key="bssa_manual_editor",
            on_change=update_session_df,
            args=("bssa_manual_df", "bssa_manual_editor")
        )

    if preset != t["bssa_preset_manual"] and bssa_df is not None:
        st.markdown(f"### {t['bssa_data_editor_title']}")
        st.dataframe(bssa_df, use_container_width=True, height=250)

    if bssa_df is not None and len(bssa_df) >= 3:
        n = len(bssa_df)
        X_vals = bssa_df["X"].values
        Y_vals = bssa_df["Y"].values
        
        sxx = np.var(X_vals, ddof=1)
        syy = np.var(Y_vals, ddof=1)
        sxy = np.cov(X_vals, Y_vals)[0, 1]
        
        st.markdown(f"### {t['cr_statistics_hdr']}")
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric(t["cr_n_points"], n)
        with col2:
            st.metric(t["cr_sw2"], f"{sxx:.6f}")
        with col3:
            st.metric(t["cr_sy2"], f"{syy:.6f}")
        with col4:
            st.metric(t["cr_swy"], f"{sxy:.6f}")
            
        ols_slope = sxy / sxx if sxx != 0 else 0.0
        ols_se = np.sqrt((syy - (sxy**2)/sxx)/((n - 2)*sxx)) if sxx != 0 and n > 2 else 0.0
        
        st.markdown(f"### {t['bssa_plot_title']}")
        
        ref_sx = 0.07
        ref_sy = 0.1
        ref_eta = (ref_sy / ref_sx)**2
        
        mean_x = np.mean(X_vals)
        mean_y = np.mean(Y_vals)
        c_ols = mean_y - ols_slope * mean_x
        
        denom_mm = sxx - ref_sx**2
        if denom_mm > 1e-6:
            ref_slope_mm = sxy / denom_mm
            c_mm = mean_y - ref_slope_mm * mean_x
        else:
            ref_slope_mm = np.nan
            c_mm = np.nan
            
        num_eiv = syy - ref_eta * sxx + np.sqrt((syy - ref_eta * sxx)**2 + 4 * ref_eta * sxy**2)
        ref_slope_eiv = num_eiv / (2 * sxy) if sxy != 0 else np.nan
        c_eiv = mean_y - ref_slope_eiv * mean_x if not np.isnan(ref_slope_eiv) else np.nan
        
        fig_scatter = plgo.Figure()
        fig_scatter.add_trace(plgo.Scatter(
            x=X_vals, y=Y_vals, mode='markers', name=t["chart_obs"],
            marker=dict(color='black', size=6),
            hovertemplate='ML (X): %{x}<br>Mw (Y): %{y}<extra></extra>'
        ))
        
        x_line = np.linspace(min(X_vals), max(X_vals), 100)
        fig_scatter.add_trace(plgo.Scatter(
            x=x_line, y=ols_slope * x_line + c_ols, mode='lines',
            name=f"OLS (Slope: {ols_slope:.3f})",
            line=dict(color='green', width=2),
            hovertemplate='X: %{x:.2f}<br>Y: %{y:.2f}<extra></extra>'
        ))
        
        if not np.isnan(ref_slope_mm):
            fig_scatter.add_trace(plgo.Scatter(
                x=x_line, y=ref_slope_mm * x_line + c_mm, mode='lines',
                name=f"MM (σx={ref_sx}, Slope: {ref_slope_mm:.3f})",
                line=dict(color='gray', width=2, dash='dash'),
                hovertemplate='X: %{x:.2f}<br>Y: %{y:.2f}<extra></extra>'
            ))
            
        if not np.isnan(ref_slope_eiv):
            fig_scatter.add_trace(plgo.Scatter(
                x=x_line, y=ref_slope_eiv * x_line + c_eiv, mode='lines',
                name=f"EIV (σx={ref_sx}, σy={ref_sy}, Slope: {ref_slope_eiv:.3f})",
                line=dict(color='blue', width=2, dash='dot'),
                hovertemplate='X: %{x:.2f}<br>Y: %{y:.2f}<extra></extra>'
            ))
            
        fig_scatter.update_layout(
            xaxis_title=t["bssa_plot_x"],
            yaxis_title=t["bssa_plot_y"],
            template='plotly_white',
            hovermode='closest'
        )
        st.plotly_chart(fig_scatter, use_container_width=True)

        st.markdown(f"### {t['bssa_slope_vs_error_title']}")
        
        grid_sx = np.linspace(0.01, 0.20, 50)
        ols_curve = np.full(len(grid_sx), ols_slope)
        mm_curve = []
        eiv_01_curve = []
        eiv_02_curve = []
        
        for sx in grid_sx:
            denom = sxx - sx**2
            if denom > 1e-6:
                mm_val = sxy / denom
            else:
                mm_val = np.nan
            mm_curve.append(mm_val)
            
            eta_01 = (0.1**2) / (sx**2)
            num_01 = syy - eta_01 * sxx + np.sqrt((syy - eta_01 * sxx)**2 + 4 * eta_01 * sxy**2)
            eiv_01 = num_01 / (2 * sxy) if sxy != 0 else np.nan
            eiv_01_curve.append(eiv_01)
            
            eta_02 = (0.2**2) / (sx**2)
            num_02 = syy - eta_02 * sxx + np.sqrt((syy - eta_02 * sxx)**2 + 4 * eta_02 * sxy**2)
            eiv_02 = num_02 / (2 * sxy) if sxy != 0 else np.nan
            eiv_02_curve.append(eiv_02)
            
        fig_slope_vs_err = plgo.Figure()
        
        fig_slope_vs_err.add_trace(plgo.Scatter(
            x=grid_sx, y=ols_curve, mode='lines',
            name="β₁ (OLS)",
            line=dict(color='green', width=2),
            hovertemplate='σ_x: %{x:.4f}<br>β₁ (OLS): %{y:.4f}<extra></extra>'
        ))
        
        fig_slope_vs_err.add_trace(plgo.Scatter(
            x=grid_sx, y=mm_curve, mode='lines',
            name="β₁ (MM)",
            line=dict(color='gray', width=2),
            hovertemplate='σ_x: %{x:.4f}<br>β₁ (MM): %{y:.4f}<extra></extra>'
        ))
        
        fig_slope_vs_err.add_trace(plgo.Scatter(
            x=grid_sx, y=eiv_01_curve, mode='lines',
            name="β₁ (EIV, σ_y = 0.1)",
            line=dict(color='red', width=2),
            hovertemplate='σ_x: %{x:.4f}<br>β₁ (EIV σy=0.1): %{y:.4f}<extra></extra>'
        ))
        
        fig_slope_vs_err.add_trace(plgo.Scatter(
            x=grid_sx, y=eiv_02_curve, mode='lines',
            name="β₁ (EIV, σ_y = 0.2)",
            line=dict(color='blue', width=2),
            hovertemplate='σ_x: %{x:.4f}<br>β₁ (EIV σy=0.2): %{y:.4f}<extra></extra>'
        ))
        
        fig_slope_vs_err.update_layout(
            xaxis_title="σ_ML (σ_x)",
            yaxis_title="Slope (β₁)",
            template='plotly_white',
            hovermode='closest'
        )
        st.plotly_chart(fig_slope_vs_err, use_container_width=True)
        
        st.markdown(f"### {t['bssa_table_title']}")
        
        tbl_sx_list = [0.024, 0.045, 0.07, 0.10, 0.12, 0.15, 0.18, 0.20]
        tbl_sy_list = [0.2, 0.1]
        
        rows = []
        for sy in tbl_sy_list:
            for sx in tbl_sx_list:
                ols_slope_val = ols_slope
                ols_se_val = ols_se
                
                denom_mm = sxx - sx**2
                if denom_mm > 1e-6:
                    mm_slope_val = sxy / denom_mm
                    s_res_mom_2 = ((n - 1) / (n - 2)) * (syy - 2 * mm_slope_val * sxy + (mm_slope_val**2) * sxx)
                    mm_se_val = np.sqrt((sxx * s_res_mom_2 + (mm_slope_val**2) * (sx**4)) / (((sxx - sx**2)**2) * (n - 1)))
                else:
                    mm_slope_val = np.nan
                    mm_se_val = np.nan
                    s_res_mom_2 = np.nan
                    
                eta = (sy**2) / (sx**2)
                num_eiv = syy - eta * sxx + np.sqrt((syy - eta * sxx)**2 + 4 * eta * sxy**2)
                eiv_slope_val = num_eiv / (2 * sxy) if sxy != 0 else np.nan
                
                if not np.isnan(eiv_slope_val):
                    sxx_val = (np.sqrt((syy - eta * sxx)**2 + 4 * eta * sxy**2) - (syy - eta * sxx)) / (2 * eta)
                    suu_val = (syy + eta * sxx - np.sqrt((syy - eta * sxx)**2 + 4 * eta * sxy**2)) / (2 * eta)
                    s_res_eiv_2 = ((n - 1) * (eta + eiv_slope_val**2) * suu_val) / (n - 2)
                    
                    denom_se_eiv = sxx_val**2 * (n - 1)
                    if denom_se_eiv > 0:
                        val_under_sqrt = sxx * s_res_eiv_2 - (eiv_slope_val**2) * suu_val
                        if val_under_sqrt > 0:
                            eiv_se_val = np.sqrt(val_under_sqrt / denom_se_eiv)
                        else:
                            eiv_se_val = 0.0
                    else:
                        eiv_se_val = np.nan
                else:
                    eiv_se_val = np.nan
                    
                if not np.isnan(mm_slope_val) and not np.isnan(s_res_mom_2):
                    sigma_q2 = s_res_mom_2 - sy**2 - (mm_slope_val**2) * (sx**2)
                else:
                    sigma_q2 = np.nan
                    
                if not np.isnan(sigma_q2) and sigma_q2 > 0:
                    eta_corr = (sigma_q2 + sy**2) / (sx**2)
                    num_eiv_corr = syy - eta_corr * sxx + np.sqrt((syy - eta_corr * sxx)**2 + 4 * eta_corr * sxy**2)
                    eiv_corr_slope = num_eiv_corr / (2 * sxy) if sxy != 0 else np.nan
                    
                    if not np.isnan(eiv_corr_slope):
                        sxx_corr = (np.sqrt((syy - eta_corr * sxx)**2 + 4 * eta_corr * sxx)**2 - (syy - eta_corr * sxx)) / (2 * eta_corr)
                        suu_corr = (syy + eta_corr * sxx - np.sqrt((syy - eta_corr * sxx)**2 + 4 * eta_corr * sxy**2)) / (2 * eta_corr)
                        s_res_eiv_corr = ((n - 1) * (eta_corr + eiv_corr_slope**2) * suu_corr) / (n - 2)
                        denom_se_corr = sxx_corr**2 * (n - 1)
                        if denom_se_corr > 0:
                            val_under_sqrt_corr = sxx * s_res_eiv_corr - (eiv_corr_slope**2) * suu_corr
                            if val_under_sqrt_corr > 0:
                                eiv_corr_se = np.sqrt(val_under_sqrt_corr / denom_se_corr)
                            else:
                                eiv_corr_se = 0.0
                        else:
                            eiv_corr_se = np.nan
                    else:
                        eiv_corr_slope = np.nan
                        eiv_corr_se = np.nan
                else:
                    eiv_corr_slope = np.nan
                    eiv_corr_se = np.nan
                    
                rows.append({
                    "sigma_y": sy,
                    "sigma_x": sx,
                    "beta_OLS": ols_slope_val,
                    "se_OLS": ols_se_val,
                    "beta_MM": mm_slope_val,
                    "se_MM": mm_se_val,
                    "beta_EIV": eiv_slope_val,
                    "se_EIV": eiv_se_val,
                    "sigma_q^2": sigma_q2,
                    "beta_EIVcorr": eiv_corr_slope,
                    "se_EIVcorr": eiv_corr_se
                })
                
        tbl_df = pd.DataFrame(rows)
        st.dataframe(
            tbl_df,
            column_config={
                "sigma_y": st.column_config.NumberColumn("σ_y (Mw)", format="%.4f"),
                "sigma_x": st.column_config.NumberColumn("σ_x (ML)", format="%.4f"),
                "beta_OLS": st.column_config.NumberColumn("β₁ OLS", format="%.3f"),
                "se_OLS": st.column_config.NumberColumn("SE OLS", format="%.3f"),
                "beta_MM": st.column_config.NumberColumn("β₁ MM", format="%.3f"),
                "se_MM": st.column_config.NumberColumn("SE MM", format="%.3f"),
                "beta_EIV": st.column_config.NumberColumn("β₁ EIV", format="%.3f"),
                "se_EIV": st.column_config.NumberColumn("SE EIV", format="%.3f"),
                "sigma_q^2": st.column_config.NumberColumn("σ_q²", format="%.3f"),
                "beta_EIVcorr": st.column_config.NumberColumn("β₁ EIVcorr", format="%.3f"),
                "se_EIVcorr": st.column_config.NumberColumn("SE EIVcorr", format="%.3f"),
            },
            use_container_width=True

        )
        
        col_dl1, col_dl2 = st.columns(2)
        with col_dl1:
            csv_bssa = tbl_df.to_csv(index=False).encode('utf-8')
            st.download_button(
                label=t["cr_download_csv"],
                data=csv_bssa,
                file_name="bssa_slopes_comparison.csv",
                mime="text/csv",
                key="bssa_download_csv_btn"
            )
        with col_dl2:
            output_excel = io.BytesIO()
            with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
                tbl_df.to_excel(writer, sheet_name="BSSA Table", index=False)
            excel_data = output_excel.getvalue()
            st.download_button(
                label=t["cr_download_excel"],
                data=excel_data,
                file_name="bssa_slopes_comparison.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="bssa_download_xlsx_btn"
            )
            
    else:
        st.warning("Se requieren al menos 3 puntos de datos para este análisis.")
        
    st.markdown("---")
    st.markdown(
        "<p style='text-align: center; color: gray; font-size: 14px;'>"
        "Desarrollado y mantenido por <b>Alexander Acosta</b> "
        "(<a href='https://github.com/j-alexander-acosta' target='_blank' style='color: #1f77b4; text-decoration: none;'>@j-alexander-acosta</a>)"
        "</p>", 
        unsafe_allow_html=True
    )
    st.stop()

if is_mc_reg:
    st.header(t["mc_reg_title"])
    st.markdown(t["mc_reg_desc"])
    
    st.sidebar.markdown(f"### {t['mc_reg_sidebar_header']}")
    mc_sigma_x = st.sidebar.number_input(t["mc_reg_sigma_x"], value=0.20, min_value=0.0001, step=0.05, format="%.4f", key="mc_reg_sig_x")
    mc_sigma_y = st.sidebar.number_input(t["mc_reg_sigma_y"], value=0.15, min_value=0.0001, step=0.05, format="%.4f", key="mc_reg_sig_y")
    
    default_true_x = [4.0 + 0.05 * i for i in range(40)]
    default_true_y = [
        4.10, 4.14, 4.19, 4.23, 4.27, 4.31, 4.36, 4.40, 4.44, 4.48,
        4.53, 4.57, 4.61, 4.65, 4.70, 4.74, 4.78, 4.82, 4.87, 4.91,
        4.95, 4.99, 5.04, 5.08, 5.12, 5.16, 5.21, 5.25, 5.29, 5.33,
        5.38, 5.42, 5.46, 5.50, 5.55, 5.59, 5.63, 5.67, 5.72, 5.76
    ]
    
    if "mc_true_df" not in st.session_state:
        st.session_state.mc_true_df = pd.DataFrame({
            "True X (mbm_bmb)": default_true_x,
            "True Y (MwM_wMw)": default_true_y
        })
        
    if "mc_errors" not in st.session_state:
        np.random.seed(42)
        n_rows = len(st.session_state.mc_true_df)
        err_x = np.random.normal(0, mc_sigma_x, n_rows)
        err_y = np.random.normal(0, mc_sigma_y, n_rows)
        st.session_state.mc_errors = {
            "err_x": err_x,
            "err_y": err_y
        }
        
    if st.sidebar.button(t["mc_reg_generate_btn"], key="mc_reg_gen_btn"):
        n_rows = len(st.session_state.mc_true_df)
        err_x = np.random.normal(0, mc_sigma_x, n_rows)
        err_y = np.random.normal(0, mc_sigma_y, n_rows)
        st.session_state.mc_errors = {
            "err_x": err_x,
            "err_y": err_y
        }
        st.rerun()
        
    n_rows = len(st.session_state.mc_true_df)
    current_err_len = len(st.session_state.mc_errors["err_x"])
    if current_err_len < n_rows:
        extra_x = np.random.normal(0, mc_sigma_x, n_rows - current_err_len)
        extra_y = np.random.normal(0, mc_sigma_y, n_rows - current_err_len)
        st.session_state.mc_errors["err_x"] = np.concatenate([st.session_state.mc_errors["err_x"], extra_x])
        st.session_state.mc_errors["err_y"] = np.concatenate([st.session_state.mc_errors["err_y"], extra_y])
    elif current_err_len > n_rows:
        st.session_state.mc_errors["err_x"] = st.session_state.mc_errors["err_x"][:n_rows]
        st.session_state.mc_errors["err_y"] = st.session_state.mc_errors["err_y"][:n_rows]
        
    df_display = pd.DataFrame({
        "True mbm_bmb": st.session_state.mc_true_df["True X (mbm_bmb)"],
        "True MwM_wMw": st.session_state.mc_true_df["True Y (MwM_wMw)"],
        "mbm_bmb Error": st.session_state.mc_errors["err_x"],
        "MwM_wMw Error": st.session_state.mc_errors["err_y"]
    })
    
    df_display["OBSERVED MB"] = df_display["True mbm_bmb"] + df_display["mbm_bmb Error"]
    df_display["OBSERVED MW"] = df_display["True MwM_wMw"] + df_display["MwM_wMw Error"]
    
    df_display.insert(0, "No.", np.arange(1, len(df_display) + 1))
    
    st.subheader(t["mc_reg_table_title"])
    st.markdown(t["mc_reg_table_desc"])
    
    edited_df = st.data_editor(
        df_display,
        num_rows="dynamic",
        use_container_width=True,
        column_config={
            "No.": st.column_config.NumberColumn(disabled=True),
            "True mbm_bmb": st.column_config.NumberColumn(format="%.2f"),
            "True MwM_wMw": st.column_config.NumberColumn(format="%.2f"),
            "mbm_bmb Error": st.column_config.NumberColumn(disabled=True, format="%.6f"),
            "MwM_wMw Error": st.column_config.NumberColumn(disabled=True, format="%.6f"),
            "OBSERVED MB": st.column_config.NumberColumn(disabled=True, format="%.6f"),
            "OBSERVED MW": st.column_config.NumberColumn(disabled=True, format="%.6f"),
        },
        key="mc_editor",
        on_change=update_mc_true_df
    )
    
    X = edited_df["OBSERVED MB"].values
    Y = edited_df["OBSERVED MW"].values
    n = len(edited_df)
    
    if n >= 2:
        # Calcular usando la nueva clase helper
        model_mc = LinearErrorsInVariables(X, Y, mc_sigma_x, mc_sigma_y)
        metrics_mc = model_mc.compute_all_metrics()
        
        m_slr = metrics_mc["ols"]["slope"]
        c_slr = metrics_mc["ols"]["intercept"]
        y_pred_slr = metrics_mc["ols"]["y_pred"]
        se_m_slr = metrics_mc["ols"]["se_slope"]
        se_c_slr = metrics_mc["ols"]["se_intercept"]
        rmse_slr = metrics_mc["ols"]["rmse"]
        r2_slr = metrics_mc["ols"]["r2"]
        
        m_gor = metrics_mc["deming"]["slope"]
        b_gor = metrics_mc["deming"]["intercept"]
        y_pred_gor = metrics_mc["deming"]["y_pred"]
        se_m_gor = metrics_mc["deming"]["se_slope"]
        se_b_gor = metrics_mc["deming"]["se_intercept"]
        rmse_gor = metrics_mc["deming"]["rmse"]
        r2_gor = metrics_mc["deming"]["r2"]
        X_t = metrics_mc["deming"]["X_t"]
        Y_t = metrics_mc["deming"]["Y_t"]
        
        m_mom = metrics_mc["mom"]["slope"]
        b_mom = metrics_mc["mom"]["intercept"]
        y_pred_mom = metrics_mc["mom"]["y_pred"]
        se_m_mom = metrics_mc["mom"]["se_slope"]
        se_b_mom = metrics_mc["mom"]["se_intercept"]
        rmse_mom = metrics_mc["mom"]["rmse"]
        r2_mom = metrics_mc["mom"]["r2"]
        
        # --- GOR Propuesto (Das et al.) ---
        x_mean = model_mc.mean_w
        Sxx = (n - 1) * model_mc.s2_w
        Syy = (n - 1) * model_mc.s2_y
        Y_t_mean = np.mean(Y_t)
        if Sxx != 0:
            m_prop = np.sum((X - x_mean) * (Y_t - Y_t_mean)) / Sxx
        else:
            m_prop = 0.0
        b_prop = Y_t_mean - m_prop * x_mean
        y_pred_prop = m_prop * X + b_prop
        
        sse_prop = np.sum((Y - y_pred_prop)**2)
        s2_e_prop = sse_prop / (n - 2) if n > 2 else 0.0
        se_m_prop = np.sqrt(s2_e_prop / Sxx) if Sxx != 0 and n > 2 else 0.0
        se_b_prop = se_m_prop * np.sqrt(np.sum(X**2) / n) if n > 0 else 0.0
        rmse_prop = np.sqrt(sse_prop / n)
        r2_prop = 1.0 - (sse_prop / Syy) if Syy != 0 else 0.0
        
        st.subheader(t["mc_reg_results_title"])
        results_data = {
            t["table_col_method"]: [t["methods_names"]["SLR"], t["methods_names"]["GOR Conv"], t["methods_names"]["GOR Prop"], t["methods_names"]["MoM"]],
            t["table_col_slope"]: [m_slr, m_gor, m_prop, m_mom],
            t["table_col_intercept"]: [c_slr, b_gor, b_prop, b_mom],
            t["table_col_se_m"]: [se_m_slr, se_m_gor, se_m_prop, se_m_mom],
            t["table_col_se_c"]: [se_c_slr, se_b_gor, se_b_prop, se_b_mom],
            t["table_col_rmse"]: [rmse_slr, rmse_gor, rmse_prop, rmse_mom],
            t["table_col_r2"]: [r2_slr, r2_gor, r2_prop, r2_mom]
        }
        df_results = pd.DataFrame(results_data)
        st.dataframe(df_results.style.format({
            t["table_col_slope"]: "{:.6f}",
            t["table_col_intercept"]: "{:.6f}",
            t["table_col_se_m"]: "{:.6f}",
            t["table_col_se_c"]: "{:.6f}",
            t["table_col_rmse"]: "{:.6f}",
            t["table_col_r2"]: "{:.6f}"
        }), use_container_width=True)
        
        x_min = float(np.min(X))
        x_max = float(np.max(X))
        x_line = np.linspace(x_min, x_max, 100)
        y_line_slr = m_slr * x_line + c_slr
        y_line_gor = m_gor * x_line + b_gor
        y_line_prop = m_prop * x_line + b_prop
        y_line_mom = m_mom * x_line + b_mom
        
        fig = plgo.Figure()
        
        fig.add_trace(plgo.Scatter(
            x=X, y=Y,
            mode='markers',
            name=t["mc_reg_obs_lbl"],
            marker=dict(color='#1f77b4', size=8),
            hovertemplate='Observed MB: %{x:.4f}<br>Observed MW: %{y:.4f}<extra></extra>'
        ))
        
        fig.add_trace(plgo.Scatter(
            x=edited_df["True mbm_bmb"], y=edited_df["True MwM_wMw"],
            mode='markers',
            name=t["mc_reg_true_lbl"],
            marker=dict(color='#7f7f7f', size=6, symbol='x'),
            hovertemplate='True MB: %{x:.4f}<br>True MW: %{y:.4f}<extra></extra>'
        ))
        
        fig.add_trace(plgo.Scatter(
            x=x_line, y=y_line_slr,
            mode='lines',
            name=t["methods_names"]["SLR"],
            line=dict(color='#2ca02c', width=2),
            hovertemplate='X: %{x:.4f}<br>Y: %{y:.4f}<extra></extra>'
        ))
        
        fig.add_trace(plgo.Scatter(
            x=x_line, y=y_line_gor,
            mode='lines',
            name=t["methods_names"]["GOR Conv"],
            line=dict(color='#d62728', width=2),
            hovertemplate='X: %{x:.4f}<br>Y: %{y:.4f}<extra></extra>'
        ))
        
        fig.add_trace(plgo.Scatter(
            x=x_line, y=y_line_prop,
            mode='lines',
            name=t["methods_names"]["GOR Prop"],
            line=dict(color='#9467bd', width=2, dash='dash'),
            hovertemplate='X: %{x:.4f}<br>Y: %{y:.4f}<extra></extra>'
        ))
        
        fig.add_trace(plgo.Scatter(
            x=x_line, y=y_line_mom,
            mode='lines',
            name=t["methods_names"]["MoM"],
            line=dict(color='#ff7f0e', width=2, dash='dashdot'),
            hovertemplate='X: %{x:.4f}<br>Y: %{y:.4f}<extra></extra>'
        ))
        
        fig.update_layout(
            title=t["mc_reg_plot_title"],
            xaxis_title="MB",
            yaxis_title="MW",
            template='plotly_white',
            legend=dict(x=0.01, y=0.99)
        )
        
        st.plotly_chart(fig, use_container_width=True)
    else:
        st.warning(t["info_data_points"])
        
    csv_data = edited_df.to_csv(index=False).encode('utf-8')
    st.download_button(
        label=t["mc_reg_download_lbl"],
        data=csv_data,
        file_name="simulacion_monte_carlo_regresion.csv",
        mime="text/csv",
        key="mc_reg_download_btn"
    )
    
    st.markdown("---")
    st.markdown(
        "<p style='text-align: center; color: gray; font-size: 14px;'>"
        "Desarrollado y mantenido por <b>Alexander Acosta</b> "
        "(<a href='https://github.com/j-alexander-acosta' target='_blank' style='color: #1f77b4; text-decoration: none;'>@j-alexander-acosta</a>)"
        "</p>", 
        unsafe_allow_html=True
    )
    st.stop()

if is_prob:
    t_prob = t["prob_module"]
    
    # Initialize distributions in session state
    state_key = f"dists_{st.session_state.lang}"
    if state_key not in st.session_state:
        st.session_state[state_key] = list(t["default_distributions"])
        
    if "editing_id" not in st.session_state:
        st.session_state.editing_id = None
        
    st.header(t_prob["title"])
    
    # Define tabs
    tab_intro, tab_class, tab_crud, tab_sturges, tab_kde, tab_chi, tab_ks, tab_monte_carlo = st.tabs([
        t_prob["tab_intro"],
        t_prob["tab_class"],
        t_prob["tab_crud"],
        t_prob["tab_sturges"],
        t_prob["tab_kde"],
        t_prob["tab_chi"],
        t_prob["tab_ks"],
        t_prob["tab_monte_carlo"]
    ])
    
    with tab_intro:
        st.subheader(t_prob["intro_title"])
        st.markdown(t_prob["intro_p1"])
        st.markdown(f"### {t_prob['intro_points_title']}")
        st.markdown(t_prob["intro_p2"])
        st.markdown("---")
        
        st.subheader(t_prob["dice_title"])
        
        # We can place N selector here
        dice_n = st.number_input(t_prob["dice_roll_num"], min_value=1, max_value=1000000, value=100, step=100, key="dice_n_input")
        
        if "dice_results" not in st.session_state:
            st.session_state.dice_results = np.random.randint(1, 7, size=100)
            
        col_roll_btn, _ = st.columns([1, 4])
        with col_roll_btn:
            if st.button(t_prob["dice_roll_btn"], key="roll_dice_btn_action"):
                st.session_state.dice_results = np.random.randint(1, 7, size=int(dice_n))
                st.success(t_prob["dice_roll_success"].format(dice_n))
                
        rolls = st.session_state.dice_results
        N_curr = len(rolls)
        counts = np.bincount(rolls, minlength=7)[1:]  # ensure length is 6 for faces 1-6
        freqs = counts / N_curr
        
        # Plotly bar chart
        fig_dice = plgo.Figure()
        fig_dice.add_trace(plgo.Bar(
            x=[1, 2, 3, 4, 5, 6],
            y=freqs,
            name=t_prob["dice_legend_exp"],
            marker_color='#1f77b4',
            opacity=0.75,
            text=[f"{f:.4f}" for f in freqs],
            textposition='auto'
        ))
        fig_dice.add_trace(plgo.Scatter(
            x=[0.5, 6.5],
            y=[1/6, 1/6],
            mode='lines',
            name=t_prob["dice_legend_theo"],
            line=dict(color='red', width=3, dash='dash')
        ))
        fig_dice.update_layout(
            xaxis_title=t_prob["dice_x_axis"],
            yaxis_title=t_prob["dice_y_axis"],
            xaxis=dict(tickmode='linear', tick0=1, dtick=1, range=[0.5, 6.5]),
            yaxis=dict(range=[0, max(max(freqs) + 0.05, 0.25)]),
            legend=dict(x=0.01, y=0.99),
            template='plotly_white',
            margin=dict(l=40, r=40, t=40, b=40)
        )
        st.plotly_chart(fig_dice, use_container_width=True)
        
        st.subheader(t_prob["dice_stats_title"])
        st.markdown(t_prob["dice_stats_desc"])
        dice_df = pd.DataFrame({
            t_prob["dice_x_axis"]: [1, 2, 3, 4, 5, 6],
            t_prob["dice_legend_exp"]: [f"{f:.4f}" for f in freqs],
            t_prob["dice_legend_theo"]: [f"{1/6:.4f}"] * 6,
            "Diferencia / Diff": [f"{abs(f - 1/6):.4f}" for f in freqs]
        })
        st.table(dice_df)
        
    with tab_class:
        st.subheader(t_prob["class_title"])
        
        col_disc, col_cont = st.columns(2)
        with col_disc:
            st.markdown(f"### {t_prob['discrete_section']}")
            st.info(t_prob["discrete_desc"])
            
            # Show discrete distributions from CRUD state
            disc_dists = [d for d in st.session_state[state_key] if d["type"] in ["Discreta", "Discrete", t_prob["crud_type_discrete"]]]
            for d in disc_dists:
                st.markdown(f"**{d['name']}**")
                st.latex(d["formula"])
                
        with col_cont:
            st.markdown(f"### {t_prob['continuous_section']}")
            st.info(t_prob["continuous_desc"])
            
            # Show continuous distributions from CRUD state
            cont_dists = [d for d in st.session_state[state_key] if d["type"] in ["Continua", "Continuous", t_prob["crud_type_continuous"]]]
            for d in cont_dists:
                st.markdown(f"**{d['name']}**")
                st.latex(d["formula"])
                
        st.markdown("---")
        
        # Gaussian Bell Curve Visualizer Parameters
        st.sidebar.markdown(f"### {t_prob['norm_params_header']}")
        mu = st.sidebar.slider(t_prob["norm_mean_label"], min_value=-10.0, max_value=10.0, value=0.0, step=0.1)
        sigma = st.sidebar.slider(t_prob["norm_sd_label"], min_value=0.1, max_value=5.0, value=1.0, step=0.1)
        
        shade_option = st.sidebar.selectbox(t_prob["norm_shade_label"], [
            t_prob["norm_shade_none"],
            t_prob["norm_shade_1s"],
            t_prob["norm_shade_2s"],
            t_prob["norm_shade_3s"],
            t_prob["norm_shade_custom"]
        ])
        
        a_shade, b_shade = None, None
        if shade_option == t_prob["norm_shade_1s"]:
            a_shade, b_shade = mu - sigma, mu + sigma
        elif shade_option == t_prob["norm_shade_2s"]:
            a_shade, b_shade = mu - 2*sigma, mu + 2*sigma
        elif shade_option == t_prob["norm_shade_3s"]:
            a_shade, b_shade = mu - 3*sigma, mu + 3*sigma
        elif shade_option == t_prob["norm_shade_custom"]:
            col_in1, col_in2 = st.sidebar.columns(2)
            with col_in1:
                custom_a = st.number_input(t_prob["norm_custom_a"], value=float(mu - sigma), step=0.5, key="norm_cust_a_input")
            with col_in2:
                custom_b = st.number_input(t_prob["norm_custom_b"], value=float(mu + sigma), step=0.5, key="norm_cust_b_input")
            a_shade, b_shade = custom_a, custom_b
            
        x_vals = np.linspace(mu - 4*sigma, mu + 4*sigma, 500)
        y_vals = stats.norm.pdf(x_vals, mu, sigma)
        
        fig_norm = plgo.Figure()
        fig_norm.add_trace(plgo.Scatter(
            x=x_vals, y=y_vals,
            mode='lines',
            name='f(X)',
            line=dict(color='#2ca02c', width=3)
        ))
        
        if a_shade is not None and b_shade is not None:
            area_prob = stats.norm.cdf(b_shade, mu, sigma) - stats.norm.cdf(a_shade, mu, sigma)
            x_shade = np.linspace(a_shade, b_shade, 200)
            y_shade = stats.norm.pdf(x_shade, mu, sigma)
            
            fig_norm.add_trace(plgo.Scatter(
                x=np.concatenate([[a_shade], x_shade, [b_shade]]),
                y=np.concatenate([[0], y_shade, [0]]),
                fill='tozeroy',
                fillcolor='rgba(44, 160, 44, 0.3)',
                line=dict(color='rgba(255,255,255,0)'),
                name=f'P({a_shade:.2f} <= X <= {b_shade:.2f})',
                hoverinfo='skip'
            ))
            
        fig_norm.update_layout(
            xaxis_title=t_prob["norm_axis_x"],
            yaxis_title=t_prob["norm_axis_y"],
            template='plotly_white',
            margin=dict(l=40, r=40, t=40, b=40)
        )
        
        st.subheader(t_prob["norm_visualizer_title"])
        st.plotly_chart(fig_norm, use_container_width=True)
        
        if a_shade is not None and b_shade is not None:
            st.info(t_prob["norm_prob_text"].format(f"{a_shade:.2f}", f"{b_shade:.2f}", area_prob))
            
    with tab_crud:
        st.subheader(t_prob["crud_section_title"])
        st.markdown(t_prob["crud_section_desc"])
        
        # Render each distribution card
        for idx, dist in enumerate(st.session_state[state_key]):
            with st.expander(f"{'📊' if dist['type'] in ['Discreta', 'Discrete', t_prob['crud_type_discrete']] else '📈'} {dist['name']} ({dist['type']})"):
                st.markdown(f"**{t_prob['crud_desc_label']}:** {dist['desc']}")
                st.markdown(f"**{t_prob['crud_formula_label']}:**")
                st.latex(dist["formula"])
                
                col_chars, col_ex = st.columns(2)
                with col_chars:
                    st.markdown(f"**{t_prob['crud_chars_label']}**")
                    chars_list = dist["chars"].split("\n")
                    for item in chars_list:
                        if item.strip():
                            st.markdown(f"- {item.strip()}")
                with col_ex:
                    st.markdown(f"**{t_prob['crud_examples_label']}**")
                    ex_list = dist["examples"].split("\n")
                    for item in ex_list:
                        if item.strip():
                            st.markdown(f"- {item.strip()}")
                            
        st.markdown("---")
        with st.expander(f"➕ {t_prob['crud_add_header']}", expanded=False):
            with st.form(key="form_add_dist_action"):
                add_name = st.text_input(t_prob["crud_name_label"])
                add_type = st.selectbox(t_prob["crud_type_label"], [t_prob["crud_type_discrete"], t_prob["crud_type_continuous"]])
                add_formula = st.text_input(t_prob["crud_formula_label"], placeholder="f(x) = ...")
                add_desc = st.text_area(t_prob["crud_desc_label"])
                add_chars = st.text_area(t_prob["crud_chars_label"])
                add_examples = st.text_area(t_prob["crud_examples_label"])
                
                submit_add = st.form_submit_button(t_prob["crud_btn_add"])
                if submit_add:
                    if add_name:
                        new_id = add_name.lower().replace(" ", "_")
                        st.session_state[state_key].append({
                            "id": new_id,
                            "name": add_name,
                            "type": add_type,
                            "formula": add_formula,
                            "desc": add_desc,
                            "chars": add_chars,
                            "examples": add_examples
                        })
                        st.success(t_prob["crud_msg_added"].format(add_name))
                        st.rerun()
                    else:
                        st.error("Por favor ingresa un nombre para la distribución.")
                        
        if st.button(t_prob["crud_btn_restore"], key="restore_defaults_action_btn"):
            st.session_state[state_key] = list(t["default_distributions"])
            st.session_state.editing_id = None
            st.success(t_prob["crud_msg_restored"])
            st.rerun()
            
    with tab_sturges:
        st.subheader(t_prob["sturges_section_title"])
        st.markdown(t_prob["sturges_section_desc"])
        
        default_raw_data = "1.2, 1.5, 1.7, 1.8, 2.0, 2.1, 2.3, 2.5, 2.7, 2.9, 3.0, 3.2, 3.5, 3.8, 4.0, 4.3, 4.7, 5.2, 6.0, 8.5"
        data_input = st.text_area(t_prob["sturges_input_label"], value=default_raw_data, height=100, key="sturges_raw_data_input")
        
        try:
            nums = [float(x.strip()) for x in data_input.split(",") if x.strip()]
            if len(nums) < 3:
                st.warning(t_prob["sturges_invalid_warn"])
                st.stop()
        except ValueError:
            st.warning(t_prob["sturges_invalid_warn"])
            st.stop()
            
        n = len(nums)
        val_min = min(nums)
        val_max = max(nums)
        val_range = val_max - val_min
        k = int(np.ceil(1 + 3.322 * np.log10(n)))
        w = val_range / k if k > 0 else 0.0
        
        st.subheader(t_prob["sturges_results_header"])
        col_st1, col_st2, col_st3 = st.columns(3)
        with col_st1:
            st.metric(t_prob["sturges_lbl_n"], n)
            st.metric(t_prob["sturges_lbl_min"], f"{val_min:.4f}")
        with col_st2:
            st.metric(t_prob["sturges_lbl_max"], f"{val_max:.4f}")
            st.metric(t_prob["sturges_lbl_range"], f"{val_range:.4f}")
        with col_st3:
            st.metric(t_prob["sturges_lbl_bins"], k)
            st.metric(t_prob["sturges_lbl_width"], f"{w:.4f}")
            
        bins_edges = [val_min + i * w for i in range(k + 1)]
        bins_edges[-1] = val_max + 1e-9
        
        counts_sturges, _ = np.histogram(nums, bins=bins_edges)
        counts_sturges = counts_sturges.astype(float)
        midpoints = np.array([(bins_edges[i] + bins_edges[i+1]) / 2 for i in range(k)], dtype=float)
        
        # Apply edits from st.session_state to counts and midpoints before calculating other columns!
        editor_key = "editable_freq_table_editor"
        if editor_key in st.session_state and st.session_state[editor_key]:
            edits = st.session_state[editor_key].get("edited_rows", {})
            freq_col = t_prob["sturges_col_freq"]
            midpoint_col = t_prob["sturges_col_midpoint"]
            for row_idx_str, col_edits in edits.items():
                try:
                    row_idx = int(row_idx_str)
                    if 0 <= row_idx < k:
                        if freq_col in col_edits:
                            counts_sturges[row_idx] = float(col_edits[freq_col])
                        if midpoint_col in col_edits:
                            midpoints[row_idx] = float(col_edits[midpoint_col])
                except ValueError:
                    pass
                    
        n_edited = np.sum(counts_sturges)
        cum_counts = np.cumsum(counts_sturges)
        
        if n_edited > 0:
            rel_freqs = counts_sturges / n_edited
            cum_rel_freqs = np.cumsum(rel_freqs)
            grouped_mean = np.sum(counts_sturges * midpoints) / n_edited
            sq_diffs = (midpoints - grouped_mean) ** 2
            weighted_sq_diffs = counts_sturges * sq_diffs
            
            # Recalculate variance and standard deviation
            sum_f_x_mean2 = np.sum(weighted_sq_diffs)
            grouped_variance = sum_f_x_mean2 / (n_edited - 1) if n_edited > 1 else 0.0
            grouped_std = np.sqrt(grouped_variance)
        else:
            rel_freqs = np.zeros_like(counts_sturges)
            cum_rel_freqs = np.zeros_like(counts_sturges)
            grouped_mean = 0.0
            sq_diffs = np.zeros_like(midpoints)
            weighted_sq_diffs = np.zeros_like(counts_sturges)
            grouped_variance = 0.0
            grouped_std = 0.0
            
        intervals_str = [f"[{bins_edges[i]:.4f}, {bins_edges[i+1]:.4f})" for i in range(k-1)]
        intervals_str.append(f"[{bins_edges[k-1]:.4f}, {val_max:.4f}]")
        
        freq_df = pd.DataFrame({
            t_prob["sturges_col_interval"]: intervals_str,
            t_prob["sturges_col_midpoint"]: midpoints,
            t_prob["sturges_col_freq"]: counts_sturges,
            t_prob["sturges_col_cum_freq"]: cum_counts,
            t_prob["sturges_col_rel_freq"]: rel_freqs,
            t_prob["sturges_col_cum_rel_freq"]: cum_rel_freqs,
            t_prob["sturges_col_prod"]: midpoints * counts_sturges,
            t_prob["sturges_col_sq_diff"]: sq_diffs,
            t_prob["sturges_col_weighted_sq_diff"]: weighted_sq_diffs
        })
        
        st.subheader(t_prob["sturges_table_title"])
        
        col_configs = {
            t_prob["sturges_col_interval"]: st.column_config.TextColumn(disabled=True),
            t_prob["sturges_col_cum_freq"]: st.column_config.NumberColumn(disabled=True),
            t_prob["sturges_col_rel_freq"]: st.column_config.NumberColumn(disabled=True, format="%.4f"),
            t_prob["sturges_col_cum_rel_freq"]: st.column_config.NumberColumn(disabled=True, format="%.4f"),
            t_prob["sturges_col_prod"]: st.column_config.NumberColumn(disabled=True, format="%.4f"),
            t_prob["sturges_col_sq_diff"]: st.column_config.NumberColumn(disabled=True, format="%.4f"),
            t_prob["sturges_col_weighted_sq_diff"]: st.column_config.NumberColumn(disabled=True, format="%.4f"),
        }
        
        edited_freq_df = st.data_editor(
            freq_df, 
            use_container_width=True, 
            num_rows="fixed", 
            key=editor_key,
            column_config=col_configs
        )
        
        # We read again from edited_freq_df to ensure any edits done in the *current* frame are processed for the chart/metrics immediately
        edited_freqs = pd.to_numeric(edited_freq_df[t_prob["sturges_col_freq"]]).values
        edited_midpoints = pd.to_numeric(edited_freq_df[t_prob["sturges_col_midpoint"]]).values
        
        n_edited = np.sum(edited_freqs)
        if n_edited > 0:
            mean_edited = np.sum(edited_freqs * edited_midpoints) / n_edited
            sq_diffs_edited = (edited_midpoints - mean_edited) ** 2
            weighted_sq_diffs_edited = edited_freqs * sq_diffs_edited
            
            # Update columns
            edited_freq_df[t_prob["sturges_col_prod"]] = edited_midpoints * edited_freqs
            edited_freq_df[t_prob["sturges_col_sq_diff"]] = sq_diffs_edited
            edited_freq_df[t_prob["sturges_col_weighted_sq_diff"]] = weighted_sq_diffs_edited
            edited_freq_df[t_prob["sturges_col_cum_freq"]] = np.cumsum(edited_freqs)
            rel_freqs_edited = edited_freqs / n_edited
            edited_freq_df[t_prob["sturges_col_rel_freq"]] = rel_freqs_edited
            edited_freq_df[t_prob["sturges_col_cum_rel_freq"]] = np.cumsum(rel_freqs_edited)
            
            sum_f_x_mean2 = np.sum(weighted_sq_diffs_edited)
            grouped_variance = sum_f_x_mean2 / (n_edited - 1) if n_edited > 1 else 0.0
            grouped_std = np.sqrt(grouped_variance)
        else:
            grouped_variance = 0.0
            grouped_std = 0.0
            
        st.markdown(f"### {t_prob['sturges_stats_header']}")
        col_res1, col_res2, col_res3 = st.columns(3)
        with col_res1:
            st.metric(t_prob["sturges_lbl_sum_freq"], f"{n_edited:.2f}")
        with col_res2:
            st.metric(t_prob["sturges_lbl_grouped_variance"], f"{grouped_variance:.4f}")
        with col_res3:
            st.metric(t_prob["sturges_lbl_grouped_std"], f"{grouped_std:.4f}")
            
        x_intervals = edited_freq_df[t_prob["sturges_col_interval"]].values
        y_counts = pd.to_numeric(edited_freq_df[t_prob["sturges_col_freq"]]).values
        
        fig_hist = plgo.Figure()
        fig_hist.add_trace(plgo.Bar(
            x=x_intervals,
            y=y_counts,
            name=t_prob["sturges_col_freq"],
            marker_color='#d62728',
            opacity=0.8,
            text=y_counts,
            textposition='auto'
        ))
        fig_hist.update_layout(
            title=t_prob["sturges_plot_title"],
            xaxis_title=t_prob["sturges_plot_x"],
            yaxis_title=t_prob["sturges_plot_y"],
            template='plotly_white',
            margin=dict(l=40, r=40, t=40, b=40)
        )
        st.plotly_chart(fig_hist, use_container_width=True)

    with tab_kde:
        st.subheader(t_prob["kde_section_title"])
        st.markdown(t_prob["kde_section_desc"])
        
        default_raw_data = "1.2, 1.5, 1.7, 1.8, 2.0, 2.1, 2.3, 2.5, 2.7, 2.9, 3.0, 3.2, 3.5, 3.8, 4.0, 4.3, 4.7, 5.2, 6.0, 8.5"
        kde_data_input = st.text_area(t_prob["kde_input_label"], value=default_raw_data, height=100, key="kde_raw_data_input")
        
        try:
            kde_nums = np.array([float(x.strip()) for x in kde_data_input.split(",") if x.strip()])
            if len(kde_nums) < 3:
                st.warning(t_prob["sturges_invalid_warn"])
                st.stop()
        except ValueError:
            st.warning(t_prob["sturges_invalid_warn"])
            st.stop()
            
        col_kde1, col_kde2 = st.columns(2)
        with col_kde1:
            kernel_choice = st.selectbox(t_prob["kde_kernel_label"], ["Gaussian", "Epanechnikov", "Uniform", "Triangular"])
        with col_kde2:
            bw_method = st.radio(t_prob["kde_bandwidth_method"], [
                t_prob["kde_bandwidth_silverman"], 
                t_prob["kde_bandwidth_scott"], 
                t_prob["kde_bandwidth_manual"]
            ])
            
        std_dev = np.std(kde_nums, ddof=1) if len(kde_nums) > 1 else 1.0
        n_points = len(kde_nums)
        
        if bw_method == t_prob["kde_bandwidth_silverman"]:
            h = 1.06 * std_dev * (n_points ** (-0.2))
            st.metric(label=f"{t_prob['kde_bandwidth_label']} ({t_prob['kde_bandwidth_silverman']})", value=f"{h:.4f}")
        elif bw_method == t_prob["kde_bandwidth_scott"]:
            h = 3.5 * std_dev * (n_points ** (-1/3))
            st.metric(label=f"{t_prob['kde_bandwidth_label']} ({t_prob['kde_bandwidth_scott']})", value=f"{h:.4f}")
        else:
            h = st.slider(t_prob["kde_bandwidth_label"], min_value=0.05, max_value=5.0, value=0.5, step=0.05, key="kde_h_manual_slider")
            
        # Kernel functions
        if kernel_choice == "Gaussian":
            K = lambda u: np.exp(-0.5 * u**2) / np.sqrt(2 * np.pi)
        elif kernel_choice == "Epanechnikov":
            K = lambda u: np.where(np.abs(u) <= 1, 0.75 * (1 - u**2), 0.0)
        elif kernel_choice == "Uniform":
            K = lambda u: np.where(np.abs(u) <= 1, 0.5, 0.0)
        elif kernel_choice == "Triangular":
            K = lambda u: np.where(np.abs(u) <= 1, 1 - np.abs(u), 0.0)
            
        # Calculate density for each observed point
        densities = []
        for xi in kde_nums:
            u = (kde_nums - xi) / h
            dens_val = np.sum(K(u)) / (n_points * h)
            densities.append(dens_val)
            
        # Create columns for layout: Left for the detailed table, Right for reference point, formulas, and extra stats.
        col_table, col_formulas = st.columns([3, 2])
        
        with col_formulas:
            st.markdown(f"### {t_prob['kde_ref_x_label']}")
            x_ref = st.selectbox(
                t_prob["kde_ref_x_label"],
                options=kde_nums,
                index=0,
                format_func=lambda val: f"{val:.4f}",
                label_visibility="collapsed",
                key="kde_x_ref_selectbox"
            )
            
            st.markdown("### Fórmulas / Formulas")
            st.latex(r"u = \frac{x - x_i}{h}")
            
            if kernel_choice == "Gaussian":
                st.latex(r"K(u) = \frac{1}{\sqrt{2\pi}} e^{-u^2 / 2}")
            elif kernel_choice == "Epanechnikov":
                st.latex(r"K(u) = \frac{3}{4}(1 - u^2) \quad \text{for } |u| \le 1")
            elif kernel_choice == "Uniform":
                st.latex(r"K(u) = \frac{1}{2} \quad \text{for } |u| \le 1")
            elif kernel_choice == "Triangular":
                st.latex(r"K(u) = 1 - |u| \quad \text{for } |u| \le 1")
                
            st.latex(r"f(x) = \frac{1}{n \cdot h} \sum_{i=1}^{n} K(u_i)")
            
            # Compute sum of K(u) and the final density at the selected reference x
            u_for_ref = (x_ref - kde_nums) / h
            k_for_ref = K(u_for_ref)
            sum_k_ref = np.sum(k_for_ref)
            f_x_ref = sum_k_ref / (n_points * h)
            
            col_met1, col_met2 = st.columns(2)
            with col_met1:
                st.metric(label=t_prob["kde_sum_k_label"], value=f"{sum_k_ref:.7f}")
            with col_met2:
                st.metric(label=t_prob["kde_val_f_x_label"], value=f"{f_x_ref:.7f}")
                
        # Compute u and K using the selected x_ref for the table
        u_vals = (x_ref - kde_nums) / h
        k_vals = K(u_vals)
        
        kde_df = pd.DataFrame({
            t_prob["kde_col_index"]: range(1, n_points + 1),
            t_prob["kde_col_value"]: kde_nums,
            "u": u_vals,
            "K": k_vals,
            t_prob["kde_col_density"]: densities,
            t_prob["kde_col_std"]: [std_dev] + [None] * (n_points - 1),
            t_prob["kde_col_n"]: [n_points] + [None] * (n_points - 1),
            t_prob["kde_col_h"]: [h] + [None] * (n_points - 1)
        })
        
        with col_table:
            st.subheader(t_prob["kde_table_title"])
            st.dataframe(
                kde_df,
                use_container_width=True,
                column_config={
                    t_prob["kde_col_index"]: st.column_config.NumberColumn(format="%d"),
                    t_prob["kde_col_value"]: st.column_config.NumberColumn(format="%.4f"),
                    "u": st.column_config.NumberColumn(format="%.9f"),
                    "K": st.column_config.NumberColumn(format="%.7f"),
                    t_prob["kde_col_density"]: st.column_config.NumberColumn(format="%.7f"),
                    t_prob["kde_col_std"]: st.column_config.NumberColumn(format="%.7f"),
                    t_prob["kde_col_n"]: st.column_config.NumberColumn(format="%d"),
                    t_prob["kde_col_h"]: st.column_config.NumberColumn(format="%.7f")
                },
                hide_index=True
            )
        
        # Grid for plotting
        x_grid = np.linspace(min(kde_nums) - 3 * h, max(kde_nums) + 3 * h, 500)
        
        # Calculate individual kernels
        individual_kernels = []
        for xi in kde_nums:
            u = (x_grid - xi) / h
            k_val = K(u) / (n_points * h)
            individual_kernels.append(k_val)
            
        # Combined KDE
        kde_curve = np.sum(individual_kernels, axis=0)
        
        # Plotly KDE chart
        fig_kde = plgo.Figure()
        
        # Add Density Histogram
        fig_kde.add_trace(plgo.Histogram(
            x=kde_nums,
            histnorm='probability density',
            name=t_prob["kde_hist_label"],
            marker=dict(color='rgba(100, 100, 100, 0.15)', line=dict(color='rgba(100, 100, 100, 0.3)', width=1)),
            opacity=0.6,
            nbinsx=10
        ))
        
        # Add individual kernels
        for i, xi in enumerate(kde_nums):
            fig_kde.add_trace(plgo.Scatter(
                x=x_grid,
                y=individual_kernels[i],
                mode='lines',
                line=dict(color='rgba(255, 127, 14, 0.3)', width=1.5, dash='dash'),
                name=t_prob["kde_individual_label"] if i == 0 else "",
                showlegend=True if i == 0 else False,
                legendgroup="individual"
            ))
            
        # Add combined KDE curve
        fig_kde.add_trace(plgo.Scatter(
            x=x_grid,
            y=kde_curve,
            mode='lines',
            name=t_prob["kde_combined_label"],
            line=dict(color='#1f77b4', width=3)
        ))
        
        # Add rug plot
        fig_kde.add_trace(plgo.Scatter(
            x=kde_nums,
            y=[-0.005] * len(kde_nums),
            mode='markers',
            marker=dict(symbol='line-ns-open', size=10, color='black', line=dict(width=2)),
            name=t_prob["kde_rug_label"]
        ))
        
        fig_kde.update_layout(
            title=t_prob["kde_plot_title"],
            xaxis_title=t_prob["kde_plot_x"],
            yaxis_title=t_prob["kde_plot_y"],
            template='plotly_white',
            margin=dict(l=40, r=40, t=40, b=40)
        )
        st.plotly_chart(fig_kde, use_container_width=True)

    with tab_chi:
        st.subheader(t_prob["chi_title"])
        st.markdown(t_prob["chi_desc"])
        
        default_raw_data = "63.123, 64.654, 65.234, 65.567, 65.789, 65.432, 65.876, 66.123, 66.456, 66.789, 66.234, 66.567, 66.89, 67.123, 67.456, 67.789, 67.234, 67.567, 67.89, 67.543, 67.876, 67.098, 67.321, 67.654, 67.987, 67.432, 67.765, 68.123, 68.456, 68.789, 68.234, 68.567, 68.89, 68.543, 68.876, 68.098, 69.123, 69.456, 69.789, 69.234, 69.567, 69.89, 69.543, 70.123, 70.456, 70.789, 71.456, 71.789, 71.234, 71.123"
        default_intervals = "63, 65, 67, 69, 71, 73"
        
        col_input1, col_input2 = st.columns(2)
        with col_input1:
            chi_data_input = st.text_area(t_prob["chi_data_input"], value=default_raw_data, height=120, key="chi_raw_data_input")
        with col_input2:
            chi_intervals_input = st.text_input(t_prob["chi_intervals_input"], value=default_intervals, key="chi_intervals_input")
            chi_alpha = st.slider(t_prob["chi_alpha_label"], min_value=0.01, max_value=0.20, value=0.05, step=0.01, key="chi_alpha_slider")
            
        try:
            chi_nums = np.array([float(x.strip()) for x in chi_data_input.split(",") if x.strip()])
            bin_edges = np.array(sorted([float(x.strip()) for x in chi_intervals_input.split(",") if x.strip()]))
            if len(chi_nums) < 5:
                st.warning(t_prob["sturges_invalid_warn"])
                st.stop()
            if len(bin_edges) < 3:
                st.warning("⚠️ Ingresa al menos 3 límites de intervalos (por ejemplo: 63, 66, 68).")
                st.stop()
        except ValueError:
            st.warning("⚠️ Asegúrate de ingresar números separados por comas.")
            st.stop()
            
        n_points = len(chi_nums)
        mean_val = np.mean(chi_nums)
        std_val = np.std(chi_nums, ddof=1) if n_points > 1 else 1.0
        k = len(bin_edges) - 1
        
        # Calculate Observed Frequencies and Normal Probabilities
        lows = []
        upps = []
        intervals_col = []
        frequencies_col = []
        z_lows = []
        p_lows = []
        p_lows_adj = []
        p_upps = []
        p_intervals = []
        expecteds = []
        chi_squares = []
        
        for i in range(k):
            low = bin_edges[i]
            upp = bin_edges[i+1]
            lows.append(low)
            upps.append(upp)
            
            # Label
            intervals_col.append(f"{low:.2f} - {upp:.2f}")
            
            # Observed counts in [low, upp)
            # Last bin includes upper boundary: [low, upp]
            if i == k - 1:
                freq = np.sum((chi_nums >= low) & (chi_nums <= upp))
            else:
                freq = np.sum((chi_nums >= low) & (chi_nums < upp))
            frequencies_col.append(int(freq))
            
            # Z-scores and probabilities
            z_l = (low - mean_val) / std_val
            z_u = (upp - mean_val) / std_val
            z_lows.append(z_l)
            
            p_l = stats.norm.cdf(z_l)
            p_u = stats.norm.cdf(z_u)
            p_lows.append(p_l)
            p_upps.append(p_u)
            
            # Adjusted low probability: first is 0, rest are previous row's P(HEIGHT<U)
            if i == 0:
                p_l_adj = 0.0
            else:
                p_l_adj = p_upps[i-1]
            p_lows_adj.append(p_l_adj)
            
            # Probability in interval: P(HEIGHT<U) - P(HEIGHT<L)
            p_int = p_u - p_l
            p_intervals.append(p_int)
            
            # Expected value: n * p_int
            exp_val = n_points * p_int
            expecteds.append(exp_val)
            
            # Chi-square term: (O - E)^2 / E
            if exp_val > 0:
                chi_sq = ((freq - exp_val) ** 2) / exp_val
            else:
                chi_sq = 0.0
            chi_squares.append(chi_sq)
            
        sum_freq = np.sum(frequencies_col)
        sum_expected = np.sum(expecteds)
        sum_chi_square = np.sum(chi_squares)
        
        # Build pandas dataframe
        chi_df = pd.DataFrame({
            "LOW": lows,
            "UPP": upps,
            "Interval": intervals_col,
            "Frequency": frequencies_col,
            "Mean": [mean_val] + [None] * (k - 1),
            "Standard Deviation": [std_val] + [None] * (k - 1),
            "Z": z_lows,
            "P(HEIGHT<L)": p_lows,
            "P(HEIGHT<L) adj": p_lows_adj,
            "P(HEIGHT<U)": p_upps,
            "P(L<HEIGHT<=U)": p_intervals,
            "expected Value": expecteds,
            "Chi Square": chi_squares
        })
        
        st.subheader(t_prob["chi_table_title"])
        
        # Display the dataframe with custom formatting
        st.dataframe(
            chi_df,
            use_container_width=True,
            column_config={
                "LOW": st.column_config.NumberColumn(format="%.1f"),
                "UPP": st.column_config.NumberColumn(format="%.1f"),
                "Interval": st.column_config.TextColumn(),
                "Frequency": st.column_config.NumberColumn(format="%d"),
                "Mean": st.column_config.NumberColumn(format="%.5f"),
                "Standard Deviation": st.column_config.NumberColumn(format="%.9f"),
                "Z": st.column_config.NumberColumn(format="%.9f"),
                "P(HEIGHT<L)": st.column_config.NumberColumn(format="%.9f"),
                "P(HEIGHT<L) adj": st.column_config.NumberColumn(format="%.9f"),
                "P(HEIGHT<U)": st.column_config.NumberColumn(format="%.9f"),
                "P(L<HEIGHT<=U)": st.column_config.NumberColumn(format="%.9f"),
                "expected Value": st.column_config.NumberColumn(format="%.8f"),
                "Chi Square": st.column_config.NumberColumn(format="%.7f")
            },
            hide_index=True
        )
        
        # Degrees of Freedom and Hypothesis Testing
        df = k - 3  # Bins - Parameters - 1 (since k bins, 2 parameters estimated, df = k - 2 - 1 = k - 3)
        
        col_res_l, col_res_r = st.columns([1, 1])
        with col_res_l:
            st.subheader(t_prob["chi_results_title"])
            
            st.metric(t_prob["chi_mean_label"], f"{mean_val:.5f}")
            st.metric(t_prob["chi_std_label"], f"{std_val:.9f}")
            
            if df > 0:
                # Critical Chi2 value at alpha
                chi_crit = stats.chi2.ppf(1.0 - chi_alpha, df)
                p_val = stats.chi2.sf(sum_chi_square, df)
                
                st.metric(t_prob["chi_df_label"], f"{df}")
                st.metric(t_prob["chi_stat_label"], f"{sum_chi_square:.7f}")
                st.metric(t_prob["chi_crit_label"], f"{chi_crit:.4f}")
                st.metric(t_prob["chi_p_label"], f"{p_val:.6f}")
                
                # Hypothesis test conclusion
                st.markdown("---")
                st.markdown(f"**{t_prob['chi_h0_text']}**")
                st.markdown(f"**{t_prob['chi_h1_text']}**")
                
                if sum_chi_square >= chi_crit:
                    st.error(t_prob["chi_reject_msg"])
                else:
                    st.success(t_prob["chi_accept_msg"])
            else:
                st.info("⚠️ Los grados de libertad son menores o iguales a 0 (gl = Bins - 3). Por favor, define más intervalos para calcular el estadístico de prueba.")
                
        with col_res_r:
            st.subheader("Visualización de Frecuencias")
            
            # Double bar chart
            fig_chi = plgo.Figure()
            fig_chi.add_trace(plgo.Bar(
                x=intervals_col,
                y=frequencies_col,
                name="Observado / Observed",
                marker_color='#1f77b4',
                text=frequencies_col,
                textposition='auto'
            ))
            fig_chi.add_trace(plgo.Bar(
                x=intervals_col,
                y=expecteds,
                name="Esperado / Expected",
                marker_color='#ff7f0e',
                text=[f"{e:.2f}" for e in expecteds],
                textposition='auto'
            ))
            fig_chi.update_layout(
                xaxis_title="Intervalos",
                yaxis_title="Frecuencias",
                barmode='group',
                legend=dict(x=0.01, y=0.99),
                template='plotly_white',
                margin=dict(l=40, r=40, t=40, b=40)
            )
            st.plotly_chart(fig_chi, use_container_width=True)

    with tab_ks:
        st.subheader(t_prob["ks_title"])
        st.markdown(t_prob["ks_desc"])
        
        default_ks_data = pd.DataFrame({
            "x": [1.2, 1.6, 1.8, 1.9, 1.9, 2.0, 2.2, 2.6, 3.0, 3.5, 4.0, 4.8, 5.6, 6.6, 7.6]
        })
        
        col_ks_input1, col_ks_input2 = st.columns([1, 1])
        with col_ks_input1:
            st.markdown(f"**{t_prob['ks_data_input']}**")
            ks_df_input = st.data_editor(
                default_ks_data, 
                num_rows="dynamic", 
                use_container_width=True, 
                key="ks_data_editor_table"
            )
        with col_ks_input2:
            ks_alpha = st.slider(
                t_prob["ks_alpha_label"],
                min_value=0.01,
                max_value=0.20,
                value=0.05,
                step=0.01,
                key="ks_alpha_slider"
            )
            
            if ks_df_input is not None and not ks_df_input.empty:
                try:
                    ks_nums = pd.to_numeric(ks_df_input["x"], errors="coerce").dropna().values
                except Exception:
                    st.warning("⚠️ Asegúrate de ingresar números válidos.")
                    st.stop()
            else:
                ks_nums = np.array([])
                
            if len(ks_nums) < 3:
                st.warning("⚠️ Ingresa al menos 3 valores numéricos en la tabla.")
                st.stop()
                
            ks_nums = np.sort(ks_nums)
            n_ks = len(ks_nums)
            
            mean_ks = np.mean(ks_nums)
            std_ks = np.std(ks_nums, ddof=1) if n_ks > 1 else 1.0
            
            if abs(ks_alpha - 0.01) < 1e-4:
                c_alpha = 1.63
            elif abs(ks_alpha - 0.02) < 1e-4:
                c_alpha = 1.52
            elif abs(ks_alpha - 0.05) < 1e-4:
                c_alpha = 1.36
            elif abs(ks_alpha - 0.10) < 1e-4:
                c_alpha = 1.22
            elif abs(ks_alpha - 0.15) < 1e-4:
                c_alpha = 1.14
            elif abs(ks_alpha - 0.20) < 1e-4:
                c_alpha = 1.07
            else:
                c_alpha = np.sqrt(-0.5 * np.log(ks_alpha / 2.0))
                
            d_crit_asymp = c_alpha / np.sqrt(n_ks)
            
            st.markdown(f"**{t_prob['ks_crit_formula_label'].format(ks_alpha, f'{c_alpha:.2f}')}**")
            st.latex(rf"D_c = \frac{{{c_alpha:.2f}}}{{\sqrt{{{n_ks}}}}} = {d_crit_asymp:.4f}")
            
        cummulative_col = np.arange(1, n_ks + 1)
        obs_cdf_col = cummulative_col / n_ks
        z_col = (ks_nums - mean_ks) / std_ks
        fx_col = stats.norm.cdf(z_col)
        diff_col = np.abs(obs_cdf_col - fx_col)
        
        ks_calc_df = pd.DataFrame({
            "x": ks_nums,
            "fre": [1] * n_ks,
            "cummulative": cummulative_col,
            "obs cdf": obs_cdf_col,
            "mean": [mean_ks] * n_ks,
            "stdev": [std_ks] * n_ks,
            "zcore": z_col,
            "FX": fx_col,
            "DIFFERENCE": diff_col
        })
        
        st.subheader(t_prob["ks_table_title"])
        
        st.dataframe(
            ks_calc_df,
            use_container_width=True,
            column_config={
                "x": st.column_config.NumberColumn(format="%.2f"),
                "fre": st.column_config.NumberColumn(format="%d"),
                "cummulative": st.column_config.NumberColumn(format="%d"),
                "obs cdf": st.column_config.NumberColumn(format="%.6f"),
                "mean": st.column_config.NumberColumn(format="%.6f"),
                "stdev": st.column_config.NumberColumn(format="%.6f"),
                "zcore": st.column_config.NumberColumn(format="%.5f"),
                "FX": st.column_config.NumberColumn(format="%.6f"),
                "DIFFERENCE": st.column_config.NumberColumn(format="%.6f")
            },
            hide_index=True
        )
        
        d_max = np.max(diff_col)
        
        col_ks_res1, col_ks_res2 = st.columns([1, 1])
        with col_ks_res1:
            st.subheader(t_prob["ks_results_title"])
            st.metric(t_prob["ks_mean_label"], f"{mean_ks:.5f}")
            st.metric(t_prob["ks_std_label"], f"{std_ks:.6f}")
            st.metric(t_prob["ks_stat_label"], f"{d_max:.6f}")
            st.metric(t_prob["ks_crit_label"], f"{d_crit_asymp:.4f}")
            
            st.markdown("---")
            st.markdown(f"**{t_prob['ks_h0_text']}**")
            st.markdown(f"**{t_prob['ks_h1_text']}**")
            
            if d_max >= d_crit_asymp:
                st.error(t_prob["ks_reject_msg"])
            else:
                st.success(t_prob["ks_accept_msg"])
                
        with col_ks_res2:
            st.subheader("ECDF vs. CDF Teórica")
            
            fig_ks = plgo.Figure()
            
            x_ecdf = np.concatenate(([ks_nums[0] - 0.5], ks_nums, [ks_nums[-1] + 0.5]))
            y_ecdf = np.concatenate(([0.0], obs_cdf_col, [1.0]))
            
            fig_ks.add_trace(plgo.Scatter(
                x=x_ecdf,
                y=y_ecdf,
                mode='lines+markers',
                name="ECDF (Empírica / Empirical)",
                line=dict(color='#1f77b4', width=3, shape='hv'),
                marker=dict(size=6)
            ))
            
            x_smooth = np.linspace(ks_nums[0] - 1.0, ks_nums[-1] + 1.0, 200)
            y_theoretical = stats.norm.cdf((x_smooth - mean_ks) / std_ks)
            
            fig_ks.add_trace(plgo.Scatter(
                x=x_smooth,
                y=y_theoretical,
                mode='lines',
                name="CDF Teórica (Normal / Theoretical)",
                line=dict(color='#ff7f0e', width=2, dash='dash')
            ))
            
            fig_ks.update_layout(
                xaxis_title="Valores (X)",
                yaxis_title="Probabilidad Acumulada / Cumulative Probability",
                template='plotly_white',
                margin=dict(l=40, r=40, t=40, b=40),
                legend=dict(x=0.01, y=0.99)
            )
            
            st.plotly_chart(fig_ks, use_container_width=True)

    with tab_monte_carlo:
        st.subheader(t_prob["mc_title"])
        st.markdown(t_prob["mc_desc"])
        
        col_mc_conf, col_mc_func = st.columns([1, 1])
        
        with col_mc_conf:
            st.markdown(f"**{t_prob['mc_dist_label']}**")
            dist_options = {
                "es": ["Normal (Gaussiana)", "Uniforme", "Exponencial", "Poisson", "Binomial"],
                "en": ["Normal", "Uniform", "Exponential", "Poisson", "Binomial"]
            }
            lang = st.session_state.lang if st.session_state.lang in dist_options else "en"
            selected_dist_label = st.selectbox(t_prob["mc_dist_label"], dist_options[lang], key="mc_dist_select", label_visibility="collapsed")
            idx = dist_options[lang].index(selected_dist_label)
            dist_ids = ["normal", "uniform", "exponential", "poisson", "binomial"]
            dist_id = dist_ids[idx]
            
            # Dynamic parameters
            if dist_id == "normal":
                mc_mu = st.number_input(t_prob["mc_normal_mu"], value=0.0, step=0.1, key="mc_mu")
                mc_sigma = st.number_input(t_prob["mc_normal_sigma"], value=1.0, min_value=0.001, step=0.1, key="mc_sigma")
            elif dist_id == "uniform":
                mc_a = st.number_input(t_prob["mc_uniform_a"], value=0.0, step=0.1, key="mc_a")
                mc_b = st.number_input(t_prob["mc_uniform_b"], value=1.0, step=0.1, key="mc_b")
                if mc_b <= mc_a:
                    st.error(t_prob["mc_uniform_error"])
                    st.stop()
            elif dist_id == "exponential":
                mc_beta = st.number_input(t_prob["mc_exponential_beta"], value=1.0, min_value=0.001, step=0.1, key="mc_beta")
            elif dist_id == "poisson":
                mc_lam = st.number_input(t_prob["mc_poisson_lambda"], value=3.0, min_value=0.001, step=0.1, key="mc_lam")
            elif dist_id == "binomial":
                mc_n_trials = st.number_input(t_prob["mc_binomial_n"], value=10, min_value=1, step=1, key="mc_n_trials")
                mc_prob = st.slider(t_prob["mc_binomial_p"], min_value=0.0, max_value=1.0, value=0.5, step=0.01, key="mc_p")
                
            mc_n_sims = st.number_input(t_prob["mc_num_sims"], min_value=10, max_value=500000, value=10000, step=1000, key="mc_nsims")

        with col_mc_func:
            st.markdown(f"**{t_prob['mc_presets']}**")
            preset_options = ["g(x) = x", "g(x) = x^2", "g(x) = exp(x)", "g(x) = sin(x)", t_prob["mc_custom_option"]]
            preset_sel = st.selectbox(t_prob["mc_presets"], preset_options, key="mc_preset", label_visibility="collapsed")
            
            if preset_sel == "g(x) = x":
                func_expr_default = "x"
            elif preset_sel == "g(x) = x^2":
                func_expr_default = "x**2"
            elif preset_sel == "g(x) = exp(x)":
                func_expr_default = "exp(x)"
            elif preset_sel == "g(x) = sin(x)":
                func_expr_default = "sin(x)"
            else:
                func_expr_default = "x"
                
            st.markdown(f"**{t_prob['mc_func_label']}**")
            mc_expr = st.text_input(t_prob["mc_func_label"], value=func_expr_default, help=t_prob["mc_func_desc"], key="mc_expr_input", label_visibility="collapsed")

        # Sample generation
        np.random.seed(42)
        N = int(mc_n_sims)
        
        if dist_id == "normal":
            x_samples = np.random.normal(loc=mc_mu, scale=mc_sigma, size=N)
            true_mean_val = mc_mu
            true_mean_lbl = f"μ = {mc_mu:.4f}"
        elif dist_id == "uniform":
            x_samples = np.random.uniform(low=mc_a, high=mc_b, size=N)
            true_mean_val = (mc_a + mc_b) / 2
            true_mean_lbl = f"(a+b)/2 = {true_mean_val:.4f}"
        elif dist_id == "exponential":
            x_samples = np.random.exponential(scale=mc_beta, size=N)
            true_mean_val = mc_beta
            true_mean_lbl = f"β = {mc_beta:.4f}"
        elif dist_id == "poisson":
            x_samples = np.random.poisson(lam=mc_lam, size=N)
            true_mean_val = mc_lam
            true_mean_lbl = f"λ = {mc_lam:.4f}"
        elif dist_id == "binomial":
            x_samples = np.random.binomial(n=int(mc_n_trials), p=mc_prob, size=N)
            true_mean_val = int(mc_n_trials) * mc_prob
            true_mean_lbl = f"n*p = {true_mean_val:.4f}"

        # Evaluate g(x)
        safe_dict = {
            "x": x_samples,
            "np": np,
            "sin": np.sin,
            "cos": np.cos,
            "tan": np.tan,
            "exp": np.exp,
            "log": np.log,
            "log10": np.log10,
            "sqrt": np.sqrt,
            "abs": np.abs,
            "pi": np.pi,
            "e": np.e
        }
        
        try:
            expr_cleaned = mc_expr.replace("^", "**").strip()
            y_samples = eval(expr_cleaned, {"__builtins__": None}, safe_dict)
            if not isinstance(y_samples, np.ndarray):
                y_samples = np.full_like(x_samples, float(y_samples))
        except Exception as e:
            st.error(t_prob["mc_invalid_func"] + f"\nError: {e}")
            st.stop()
            
        # Calculation of stats
        y_mean = np.mean(y_samples)
        y_std = np.std(y_samples, ddof=1) if N > 1 else 0.0
        se = y_std / np.sqrt(N)
        me = 1.96 * se
        ci_lower = y_mean - me
        ci_upper = y_mean + me
        
        st.markdown("---")
        st.subheader(t_prob["mc_stats_title"])
        
        col_m1, col_m2, col_m3, col_m4 = st.columns(4)
        with col_m1:
            st.metric(t_prob["mc_est_mean"], f"{y_mean:.6f}")
        with col_m2:
            if mc_expr.strip().lower() in ["x", "g(x) = x"]:
                st.metric(t_prob["mc_true_mean"], true_mean_lbl)
            else:
                st.metric(t_prob["mc_true_mean"], "N/A")
        with col_m3:
            st.metric(t_prob["mc_std_dev"], f"{y_std:.6f}")
        with col_m4:
            st.metric(t_prob["mc_std_error"], f"{se:.6f}")
            
        st.info(f"**{t_prob['mc_ci']}** `[{ci_lower:.6f}, {ci_upper:.6f}]`")
        
        # Interactive plots
        col_g1, col_g2 = st.columns(2)
        
        with col_g1:
            st.subheader(t_prob["mc_plot_x_title"])
            fig_x = plgo.Figure()
            # Samples histogram
            fig_x.add_trace(plgo.Histogram(
                x=x_samples,
                nbinsx=50,
                histnorm='probability density',
                name=t_prob["mc_plot_samples"],
                marker_color='#1f77b4',
                opacity=0.6
            ))
            
            # Theoretical density if continuous
            if dist_id in ["normal", "uniform", "exponential"]:
                x_eval = np.linspace(np.min(x_samples), np.max(x_samples), 200)
                if dist_id == "normal":
                    y_eval = stats.norm.pdf(x_eval, loc=mc_mu, scale=mc_sigma)
                elif dist_id == "uniform":
                    y_eval = stats.uniform.pdf(x_eval, loc=mc_a, scale=mc_b - mc_a)
                elif dist_id == "exponential":
                    y_eval = stats.expon.pdf(x_eval, scale=mc_beta)
                fig_x.add_trace(plgo.Scatter(
                    x=x_eval,
                    y=y_eval,
                    mode='lines',
                    name=t_prob["mc_plot_theo"],
                    line=dict(color='red', width=3)
                ))
            elif dist_id in ["poisson", "binomial"]:
                x_eval = np.unique(x_samples)
                if dist_id == "poisson":
                    y_eval = stats.poisson.pmf(x_eval, mu=mc_lam)
                elif dist_id == "binomial":
                    y_eval = stats.binom.pmf(x_eval, n=int(mc_n_trials), p=mc_prob)
                fig_x.add_trace(plgo.Bar(
                    x=x_eval,
                    y=y_eval,
                    name=t_prob["mc_plot_theo_pmf"],
                    marker_color='red',
                    opacity=0.8,
                    width=0.3
                ))
                
            fig_x.update_layout(
                xaxis_title="X",
                yaxis_title=t_prob["mc_plot_density"],
                template='plotly_white',
                margin=dict(l=30, r=30, t=30, b=30),
                legend=dict(x=0.01, y=0.99)
            )
            st.plotly_chart(fig_x, use_container_width=True)

        with col_g2:
            st.subheader(t_prob["mc_plot_y_title"])
            fig_y = plgo.Figure()
            # Transformed samples histogram
            fig_y.add_trace(plgo.Histogram(
                x=y_samples,
                nbinsx=50,
                histnorm='probability density',
                name="Y = g(X)",
                marker_color='#ff7f0e',
                opacity=0.6
            ))
            
            # Estimated mean
            hist_y_vals, _ = np.histogram(y_samples, bins=50, density=True)
            max_y_density = np.max(hist_y_vals) if len(hist_y_vals) > 0 else 1.0
            
            fig_y.add_trace(plgo.Scatter(
                x=[y_mean, y_mean],
                y=[0, 1.1 * max_y_density],
                mode='lines',
                name=t_prob["mc_plot_est_mean"],
                line=dict(color='blue', width=3, dash='dash')
            ))
            
            # Confidence interval boundaries
            fig_y.add_trace(plgo.Scatter(
                x=[ci_lower, ci_lower],
                y=[0, 1.1 * max_y_density],
                mode='lines',
                name=t_prob["mc_plot_ci_lower"],
                line=dict(color='purple', width=1.5, dash='dot')
            ))
            fig_y.add_trace(plgo.Scatter(
                x=[ci_upper, ci_upper],
                y=[0, 1.1 * max_y_density],
                mode='lines',
                name=t_prob["mc_plot_ci_upper"],
                line=dict(color='purple', width=1.5, dash='dot')
            ))
            
            fig_y.update_layout(
                xaxis_title="Y = g(X)",
                yaxis_title=t_prob["mc_plot_density"],
                template='plotly_white',
                margin=dict(l=30, r=30, t=30, b=30),
                legend=dict(x=0.01, y=0.99)
            )
            st.plotly_chart(fig_y, use_container_width=True)

        st.markdown("---")
        st.subheader(t_prob["mc_math_title"])
        
        # Latex details
        st.markdown(t_prob["mc_math_p1"])
        st.latex(r"\mathbb{E}[g(X)] = \int_{-\infty}^{\infty} g(x)f(x)\,dx \quad \text{o} \quad \sum_{x} g(x)P(X=x)")
        
        st.markdown(t_prob["mc_math_p2"])
        st.latex(r"\bar{Y} = \frac{1}{N} \sum_{i=1}^{N} g(X_i)")
        
        st.markdown(t_prob["mc_math_p3"])
        st.latex(r"\text{SE} = \frac{S_Y}{\sqrt{N}} = \sqrt{\frac{1}{N(N-1)} \sum_{i=1}^{N} (g(X_i) - \bar{Y})^2}")
        
        st.markdown(t_prob["mc_math_p4"])
        st.latex(r"\text{IC}_{95\%} = \left[ \bar{Y} - 1.96 \cdot \text{SE}, \,\, \bar{Y} + 1.96 \cdot \text{SE} \right]")
        
        # CSV export
        df_export = pd.DataFrame({
            "Index": np.arange(1, N + 1),
            "X_Sample": x_samples,
            "Y_Transformed": y_samples
        })
        csv_data = df_export.to_csv(index=False).encode('utf-8')
        
        st.download_button(
            label=t_prob["mc_download_lbl"],
            data=csv_data,
            file_name="simulacion_monte_carlo.csv",
            mime="text/csv",
            key="mc_download_btn"
        )

    # Render Footer and stop execution
    st.markdown("---")
    st.markdown(
        "<p style='text-align: center; color: gray; font-size: 14px;'>"
        "Desarrollado y mantenido por <b>Alexander Acosta</b> "
        "(<a href='https://github.com/j-alexander-acosta' target='_blank' style='color: #1f77b4; text-decoration: none;'>@j-alexander-acosta</a>)"
        "</p>", 
        unsafe_allow_html=True
    )
    st.stop()

st.sidebar.header(t["sb_header_data"])
data_source = st.sidebar.radio(t["sb_data_source"], (t["sb_manual"], t["sb_upload"]))

df = None

if not is_mlr:
    if data_source == t["sb_manual"]:
        st.sidebar.markdown(t["sb_manual_desc"])
        default_data = pd.DataFrame({
            "X": [4.4, 4.5, 4.6, 4.7, 4.8, 4.9, 5.0, 5.1, 5.2, 5.3, 5.4, 5.5, 5.6, 5.7, 5.8, 5.9, 6.0, 6.1, 6.2],
            "Y": [4.7, 4.6, 5.1, 5.2, 5.0, 5.4, 5.3, 5.5, 5.4, 5.6, 5.8, 5.9, 5.7, 6.1, 6.0, 6.2, 6.4, 6.3, 6.6]
        })
        df = st.sidebar.data_editor(default_data, num_rows="dynamic", use_container_width=True)
    else:
        uploaded_file = st.sidebar.file_uploader(t["sb_upload_desc"], type=["csv", "xlsx"])
        if uploaded_file is not None:
            try:
                if uploaded_file.name.endswith('.csv'):
                    df_raw = pd.read_csv(uploaded_file)
                else:
                    df_raw = pd.read_excel(uploaded_file)
                
                # Detectar errores de medición y guardarlos en session_state ( Carroll & Ruppert )
                if "MB_ERROR" in df_raw.columns:
                    st.session_state.detected_sigma_u = float(np.mean(df_raw["MB_ERROR"]) / 2.0)
                if "MW_ERROR" in df_raw.columns:
                    st.session_state.detected_sigma_e = float(np.mean(df_raw["MW_ERROR"]) / 2.0)
                
                st.sidebar.markdown(t["sb_select_cols"])
                x_col = st.sidebar.selectbox(t["sb_col_x"], df_raw.columns)
                y_col = st.sidebar.selectbox(t["sb_col_y"], df_raw.columns)
                
                df = df_raw[[x_col, y_col]].rename(columns={x_col: "X", y_col: "Y"}).dropna()
                
            except Exception as e:
                st.sidebar.error(t["sb_err_file"].format(e))
else:
    if data_source == t["sb_manual"]:
        st.sidebar.markdown(t["sb_manual_desc"])
        default_data_mlr = pd.DataFrame({
            "logPOA": [1.91991, 1.91472, 1.56733, 1.93615, 1.84532, 1.46172, 1.73993, 1.34949, 1.23044, 1.73635, 1.58530, 1.85273, 1.68352, 2.03520, 1.60728],
            "M": [4.78507, 4.78507, 4.78507, 4.78507, 4.78507, 4.78507, 4.78507, 5.81963, 5.81963, 5.74574, 5.74574, 5.74574, 5.74574, 5.74574, 5.74574],
            "M^2": [22.8969, 22.8969, 22.8969, 22.8969, 22.8969, 22.8969, 22.8969, 33.8681, 33.8681, 33.0135, 33.0135, 33.0135, 33.0135, 33.0135, 33.0135],
            "logR": [1.79358, 1.81018, 1.83163, 1.85999, 1.90601, 1.91687, 1.92090, 2.24137, 2.40310, 2.07246, 1.93409, 2.07355, 2.10917, 2.07282, 2.17219],
            "R": [62.1692, 64.5929, 67.8788, 72.4423, 80.5395, 82.5788, 83.3484, 174.328, 252.991, 118.156, 85.9195, 118.454, 128.578, 118.255, 148.659]
        })
        df = st.sidebar.data_editor(default_data_mlr, num_rows="dynamic", use_container_width=True)
    else:
        uploaded_file = st.sidebar.file_uploader(t["sb_upload_desc"], type=["csv", "xlsx"])
        if uploaded_file is not None:
            try:
                if uploaded_file.name.endswith('.csv'):
                    df = pd.read_csv(uploaded_file)
                else:
                    df = pd.read_excel(uploaded_file)
            except Exception as e:
                st.sidebar.error(t["sb_err_file"].format(e))

if df is not None and len(df) >= 2:
    if is_mlr:
        st.sidebar.markdown("---")
        st.sidebar.header(t["sb_select_cols"])
        y_col = st.sidebar.selectbox(t["sb_col_y"], df.columns, index=0)
        x_cols = st.sidebar.multiselect(t["sb_col_x_multiple"], [c for c in df.columns if c != y_col], default=[c for c in df.columns if c != y_col])
        
        if not x_cols:
            st.warning(t["mlr_err_no_x"])
            st.markdown("---")
            st.markdown(
                "<p style='text-align: center; color: gray; font-size: 14px;'>"
                "Desarrollado y mantenido por <b>Alexander Acosta</b> "
                "(<a href='https://github.com/j-alexander-acosta' target='_blank' style='color: #1f77b4; text-decoration: none;'>@j-alexander-acosta</a>)"
                "</p>", 
                unsafe_allow_html=True
            )
            st.stop()
            
        df_clean = df[[y_col] + x_cols].dropna()
        if len(df_clean) < len(x_cols) + 2:
            st.warning(t["mlr_err_insufficient_data"].format(len(x_cols) + 2, len(x_cols)))
            st.markdown("---")
            st.markdown(
                "<p style='text-align: center; color: gray; font-size: 14px;'>"
                "Desarrollado y mantenido por <b>Alexander Acosta</b> "
                "(<a href='https://github.com/j-alexander-acosta' target='_blank' style='color: #1f77b4; text-decoration: none;'>@j-alexander-acosta</a>)"
                "</p>", 
                unsafe_allow_html=True
            )
            st.stop()
            
        # Math Solver
        Y = df_clean[y_col].values
        X_vals = df_clean[x_cols].values
        N = len(Y)
        K = len(x_cols)
        
        X_mat = np.column_stack([np.ones(N), X_vals])
        
        try:
            XtX = np.dot(X_mat.T, X_mat)
            XtX_inv = np.linalg.inv(XtX)
            XtY = np.dot(X_mat.T, Y)
            beta = np.dot(XtX_inv, XtY)
            
            Y_pred = np.dot(X_mat, beta)
            residuals = Y - Y_pred
            
            Y_mean = np.mean(Y)
            SST = np.sum((Y - Y_mean)**2)
            SSE = np.sum(residuals**2)
            SSR = SST - SSE
            
            df_reg = K
            df_resid = N - K - 1
            df_total = N - 1
            
            MSR = SSR / df_reg if df_reg > 0 else 0
            MSE = SSE / df_resid if df_resid > 0 else 0
            
            F_stat = MSR / MSE if MSE > 0 else 0
            signif_f = stats.f.sf(F_stat, df_reg, df_resid) if df_resid > 0 else 1.0
            
            r2 = 1.0 - (SSE / SST) if SST > 0 else 0.0
            r2_adj = 1.0 - ((SSE / df_resid) / (SST / df_total)) if df_resid > 0 and df_total > 0 else 0.0
            se_reg = np.sqrt(MSE)
            
            cov_beta = MSE * XtX_inv
            se_beta = np.sqrt(np.diagonal(cov_beta))
            t_stats = beta / se_beta
            p_values = 2.0 * stats.t.sf(np.abs(t_stats), df_resid) if df_resid > 0 else np.ones(K+1)
            
            t_crit = stats.t.ppf(0.975, df_resid) if df_resid > 0 else 0.0
            ci_lower = beta - t_crit * se_beta
            ci_upper = beta + t_crit * se_beta
            
            multiple_r = np.sqrt(r2) if r2 >= 0 else 0.0
            Y_sorted = np.sort(Y)
            percentiles = (np.arange(1, N + 1) - 0.5) / N * 100
        except np.linalg.LinAlgError:
            st.error("Error: Colinealidad detectada en las variables independientes. Por favor, remueve alguna variable correlacionada.")
            st.stop()
            
        # UI
        st.header(t["mlr_title"])
        tab_stats, tab_kmean, tab_kvalue, tab_elbow, tab_silhouette, tab_dbscan, tab_formulas, tab_res_analysis, tab_prob_output, tab_normal, tab_plots = st.tabs([
            t["mlr_tab_stats"],
            t["mlr_tab_kmean"],
            t["mlr_tab_kvalue"],
            t["mlr_tab_elbow"],
            t["mlr_tab_silhouette"],
            t["mlr_tab_dbscan"],
            t["mlr_tab_formulas"],
            t["mlr_tab_res_analysis"],
            t["mlr_tab_prob_output"],
            t["mlr_tab_normal"],
            t["mlr_tab_plots"]
        ])
        
        with tab_stats:
            st.subheader(t["mlr_stats_title"])
            stats_df = pd.DataFrame({
                t["table_col_method"]: [
                    t["mlr_multiple_r"],
                    t["mlr_r2"],
                    t["mlr_r2_adj"],
                    t["mlr_se"],
                    t["mlr_obs"]
                ],
                "Valor": [
                    f"{multiple_r:.8f}",
                    f"{r2:.8f}",
                    f"{r2_adj:.8f}",
                    f"{se_reg:.8f}",
                    f"{N}"
                ]
            })
            st.table(stats_df)
            
            st.subheader(t["mlr_anova_title"])
            anova_df = pd.DataFrame({
                "": [t["mlr_row_reg"], t["mlr_row_resid"], t["mlr_row_total"]],
                t["mlr_anova_df"]: [df_reg, df_resid, df_total],
                t["mlr_anova_ss"]: [f"{SSR:.6f}", f"{SSE:.6f}", f"{SST:.6f}"],
                t["mlr_anova_ms"]: [f"{MSR:.6f}", f"{MSE:.6f}", ""],
                t["mlr_anova_f"]: [f"{F_stat:.6f}" if df_reg > 0 else "", "", ""],
                t["mlr_anova_sig_f"]: [f"{signif_f:.3e}" if df_reg > 0 else "", "", ""]
            })
            st.table(anova_df)
            
            st.subheader(t["mlr_coefs_title"])
            coef_names = [t["mlr_intercept"]] + list(x_cols)
            coefs_df = pd.DataFrame({
                "": coef_names,
                t["mlr_col_coef"]: [f"{b:.8f}" for b in beta],
                t["mlr_col_se"]: [f"{se:.8f}" for se in se_beta],
                t["mlr_col_t"]: [f"{t_val:.8f}" for t_val in t_stats],
                t["mlr_col_p"]: [f"{p:.3e}" for p in p_values],
                t["mlr_col_low95"]: [f"{low:.8f}" for low in ci_lower],
                t["mlr_col_upp95"]: [f"{upp:.8f}" for upp in ci_upper]
            })
            st.table(coefs_df)
            
        with tab_formulas:
            st.markdown(r"""
### Estadísticas de la Regresión

| Estadística | Fórmula Visual (Cálculo) |
| :--- | :--- |
| **Coeficiente de correlación múltiple** | $R = \sqrt{R^2}$ |
| **Coeficiente de determinación ($R^2$)** | $R^2 = \frac{SSR}{SST} = \frac{\sum(\hat{y}_i - \bar{y})^2}{\sum(y_i - \bar{y})^2}$ |
| **$R^2$ ajustado** | $R^2_{adj} = 1 - (1 - R^2) \frac{(n - 1)}{(n - k - 1)}$ |
| **Error típico** | $S_e = \sqrt{\frac{SSE}{(n - k - 1)}} = \sqrt{\frac{\sum(y_i - \hat{y}_i)^2}{(n - k - 1)}}$ |
| **Observaciones** | $n = \text{Total Observaciones}$ |

<br>

### Análisis de Varianza (ANOVA)

| Fuente | Grados de libertad ($df$) | Suma de Cuadrados ($SS$) | Promedio de los Cuadrados ($MS$) | $F$ |
| :--- | :--- | :--- | :--- | :--- |
| **Regresión** | $k$ | $SSR = \sum(\hat{y}_i - \bar{y})^2$ | $MSR = \frac{SSR}{k}$ | $F = \frac{MSR}{MSE}$ |
| **Residuos** | $n - k - 1$ | $SSE = \sum(y_i - \hat{y}_i)^2$ | $MSE = \frac{SSE}{n - k - 1}$ | |
| **Total** | $n - 1$ | $SST = \sum(y_i - \bar{y})^2$ | | |

<br>

### Parámetros y Coeficientes

| Estadística | Fórmula Visual (Cálculo) |
| :--- | :--- |
| **Coeficientes ($\hat{\beta}$)** | $\hat{\beta} = (X^T X)^{-1} X^T Y$ |
| **Error típico ($SE(\hat{\beta}_j)$)** | $SE(\hat{\beta}_j) = \sqrt{MSE \cdot [(X^T X)^{-1}]_{jj}}$ |
| **Estadístico $t$** | $t = \frac{\hat{\beta}_j}{SE(\hat{\beta}_j)}$ |
| **Probabilidad (p-valor)** | $p = 2 \cdot (1 - P(T \le \lvert t \rvert))$ |
| **Intervalo de Confianza (95%)** | $\hat{\beta}_j \pm t_{\alpha/2, \text{df}} \cdot SE(\hat{\beta}_j)$ |

<br>

**Leyenda:**
- **SSR, SST, SSE:** Suma de Cuadrados (Regresión, Total, Error/Residuo)
- **MSR, MSE:** Promedio de los Cuadrados (Regresión, Residuo)
- **n:** total de observaciones
- **k:** número de predictores (variables independientes)
- **$X$:** Matriz de variables independientes (incluyendo columna de unos)
- **$Y$:** Vector de la variable dependiente
- **$(X^T X)^{-1}$:** Matriz inversa de $X^T X$
- **$[ \cdot ]_{jj}$:** Elemento de la diagonal principal de la matriz
- **$t_{\alpha/2, \text{df}}$:** Valor crítico $t$ de Student con $\alpha = 0.05$
""", unsafe_allow_html=True)
            
            
        with tab_res_analysis:
            st.subheader(t["mlr_res_analysis_title"])
            
            res_df = pd.DataFrame({
                t["mlr_observation"]: np.arange(1, N + 1),
                t["mlr_predicted_value"].format(y_col): Y_pred,
                t["mlr_residual_val"]: residuals
            })
            st.dataframe(
                res_df.style.format({
                    t["mlr_predicted_value"].format(y_col): "{:.8f}",
                    t["mlr_residual_val"]: "{:.8f}"
                }),
                use_container_width=True,
                hide_index=True
            )
            
        with tab_prob_output:
            st.subheader(t["mlr_tab_prob_output"])
            
            prob_df = pd.DataFrame({
                t["mlr_prob_observation"]: np.arange(1, N + 1),
                t["mlr_prob_percentile"]: percentiles,
                t["mlr_prob_ordered_y"].format(y_col): Y_sorted
            })
            st.dataframe(
                prob_df.style.format({
                    t["mlr_prob_percentile"]: "{:.8f}",
                    t["mlr_prob_ordered_y"].format(y_col): "{:.8f}"
                }),
                use_container_width=True,
                hide_index=True
            )
            
        with tab_normal:
            fig_norm = plgo.Figure()
            fig_norm.add_trace(plgo.Scatter(
                x=percentiles, y=Y_sorted, mode='markers+lines',
                marker=dict(color='#1f77b4', size=8),
                line=dict(color='#1f77b4', width=1.5),
                hovertemplate='Percentil: %{x:.2f}%<br>Y: %{y:.4f}<extra></extra>'
            ))
            fig_norm.update_layout(
                title=t["mlr_normal_plot_title"],
                xaxis_title=t["mlr_normal_x_label"],
                yaxis_title=t["mlr_normal_y_label"],
                template='plotly_white',
                height=450
            )
            st.plotly_chart(fig_norm, use_container_width=True)
            
        with tab_plots:
            for j, x_col in enumerate(x_cols):
                st.markdown(f"### Variable X: **{x_col}**")
                col_p1, col_p2 = st.columns(2)
                x_vals_col = df_clean[x_col].values
                
                with col_p1:
                    fig_res = plgo.Figure()
                    fig_res.add_trace(plgo.Scatter(
                        x=x_vals_col, y=residuals, mode='markers',
                        marker=dict(color='#1f77b4', size=8),
                        hovertemplate='X: %{x}<br>Residuo: %{y:.4f}<extra></extra>'
                    ))
                    fig_res.add_hline(y=0, line_dash="dash", line_color="red")
                    fig_res.update_layout(
                        title=t["mlr_res_plot_title"].format(x_col),
                        xaxis_title=x_col,
                        yaxis_title=t["mlr_res_y_label"],
                        template='plotly_white',
                        height=350
                    )
                    st.plotly_chart(fig_res, use_container_width=True)
                    
                with col_p2:
                    fig_fit = plgo.Figure()
                    fig_fit.add_trace(plgo.Scatter(
                        x=x_vals_col, y=Y, mode='markers', name=t["mlr_observed"],
                        marker=dict(color='#1f77b4', size=8),
                        hovertemplate='X: %{x}<br>Obs: %{y:.4f}<extra></extra>'
                    ))
                    sort_idx = np.argsort(x_vals_col)
                    fig_fit.add_trace(plgo.Scatter(
                        x=x_vals_col[sort_idx], y=Y_pred[sort_idx], mode='lines+markers', name=t["mlr_predicted"],
                        line=dict(color='orange', width=2),
                        marker=dict(color='orange', size=6),
                        hovertemplate='X: %{x}<br>Pred: %{y:.4f}<extra></extra>'
                    ))
                    fig_fit.update_layout(
                        title=t["mlr_fit_plot_title"].format(x_col),
                        xaxis_title=x_col,
                        yaxis_title=y_col,
                        template='plotly_white',
                        height=350
                    )
                    st.plotly_chart(fig_fit, use_container_width=True)
            
        with tab_kmean:
            st.subheader(t["km_title"])
            st.write(t["km_intro"])
            
            col_f1, col_f2, col_f3 = st.columns(3)
            with col_f1:
                st.markdown(f"#### {t['km_formula_dist_title']}")
                st.latex(r"d(P, C) = \sqrt{(x_p - x_c)^2 + (y_p - y_c)^2}")
                st.write(t["km_formula_dist"])
            with col_f2:
                st.markdown(f"#### {t['km_formula_assign_title']}")
                st.latex(r"\text{Cluster}(P) = \arg\min_{i} d(P, C_i)")
                st.write(t["km_formula_assign"])
            with col_f3:
                st.markdown(f"#### {t['km_formula_update_title']}")
                st.latex(r"C_i^{(new)} = \frac{1}{|S_i|} \sum_{P \in S_i} P")
                st.write(t["km_formula_update"])

            st.markdown("---")
            st.subheader(t["km_input_data_title"])
            
            col_ed1, col_ed2 = st.columns(2)
            with col_ed1:
                st.markdown(f"**{t['km_data_points_label']}**")
                df_points_default = pd.DataFrame({
                    t["km_col_label"]: ["A1", "A2", "A3", "B1", "B2", "B3", "C1", "C2"],
                    "X": [2.0, 2.0, 8.0, 5.0, 7.0, 6.0, 1.0, 4.0],
                    "Y": [10.0, 5.0, 4.0, 8.0, 5.0, 4.0, 2.0, 9.0]
                })
                df_points = st.data_editor(df_points_default, num_rows="dynamic", use_container_width=True, key="km_df_points")
                st.session_state["km_df_points_df"] = df_points
            with col_ed2:
                st.markdown(f"**{t['km_centroids_label']}**")
                df_centroids_default = pd.DataFrame({
                    t["km_col_label"]: ["A1", "B1", "C1"],
                    "X": [2.0, 6.0, 1.5],
                    "Y": [10.0, 6.0, 3.5]
                })
                df_centroids = st.data_editor(df_centroids_default, num_rows="dynamic", use_container_width=True, key="km_df_centroids")

            if df_points is not None and df_centroids is not None:
                points_clean = df_points.dropna(subset=["X", "Y"])
                centroids_clean = df_centroids.dropna(subset=["X", "Y"])
                
                if len(points_clean) > 0 and len(centroids_clean) > 0:
                    import plotly.express as px
                    
                    # Calculate Euclidean distances
                    dist_matrix = {}
                    for idx, row_c in centroids_clean.iterrows():
                        c_label = str(row_c[t["km_col_label"]])
                        c_x = float(row_c["X"])
                        c_y = float(row_c["Y"])
                        
                        distances = []
                        for idx_p, row_p in points_clean.iterrows():
                            p_x = float(row_p["X"])
                            p_y = float(row_p["Y"])
                            dist = np.sqrt((p_x - c_x)**2 + (p_y - c_y)**2)
                            distances.append(dist)
                        
                        dist_matrix[c_label] = distances
                    
                    # Assign initial cluster (closest centroid)
                    closest_clusters = []
                    for i in range(len(points_clean)):
                        min_dist = float('inf')
                        best_cluster_idx = 1
                        for idx_c, (c_label, distances) in enumerate(dist_matrix.items()):
                            if distances[i] < min_dist:
                                min_dist = distances[i]
                                best_cluster_idx = idx_c + 1
                        closest_clusters.append(best_cluster_idx)
                    
                    # Calculate New Centroids (average of points assigned to each cluster)
                    new_centroids_list = []
                    for idx_c, c_label in enumerate(dist_matrix.keys()):
                        cluster_num = idx_c + 1
                        assigned_points_x = []
                        assigned_points_y = []
                        for i, p_idx in enumerate(points_clean.index):
                            if closest_clusters[i] == cluster_num:
                                assigned_points_x.append(points_clean.loc[p_idx, "X"])
                                assigned_points_y.append(points_clean.loc[p_idx, "Y"])
                        
                        if len(assigned_points_x) > 0:
                            new_x = np.mean(assigned_points_x)
                            new_y = np.mean(assigned_points_y)
                        else:
                            c_row = centroids_clean.iloc[idx_c]
                            new_x = float(c_row["X"])
                            new_y = float(c_row["Y"])
                        new_centroids_list.append((c_label, new_x, new_y))
                    
                    # Calculate New Cluster Assignment (using the updated new centroids)
                    new_closest_clusters = []
                    for idx_p, row_p in points_clean.iterrows():
                        p_x = float(row_p["X"])
                        p_y = float(row_p["Y"])
                        min_dist = float('inf')
                        best_cluster_idx = 1
                        for idx_c, (c_label, new_x, new_y) in enumerate(new_centroids_list):
                            dist = np.sqrt((p_x - new_x)**2 + (p_y - new_y)**2)
                            if dist < min_dist:
                                min_dist = dist
                                best_cluster_idx = idx_c + 1
                        new_closest_clusters.append(best_cluster_idx)

                    # Build results dataframe
                    results_df = pd.DataFrame()
                    results_df[t["km_col_label"]] = points_clean[t["km_col_label"]].values
                    results_df["X"] = points_clean["X"].values
                    results_df["Y"] = points_clean["Y"].values
                    
                    format_dict = {}
                    for c_label, distances in dist_matrix.items():
                        col_name = t["km_col_distance_to"].format(c_label)
                        results_df[col_name] = distances
                        format_dict[col_name] = "{:.2f}"
                        
                    results_df[t["km_col_cluster"]] = closest_clusters
                    results_df[t["km_col_new_cluster"]] = new_closest_clusters
                    
                    format_dict["X"] = "{:.2f}"
                    format_dict["Y"] = "{:.2f}"
                    
                    # Render Main Table
                    st.markdown("---")
                    st.subheader(t["km_table_title"])
                    st.dataframe(
                        results_df.style.format(format_dict),
                        use_container_width=True,
                        hide_index=True
                    )
                    
                    # Render New Centroids & Plotly
                    st.markdown("---")
                    col_nc1, col_nc2 = st.columns([1, 2])
                    
                    with col_nc1:
                        st.subheader(t["km_new_centroids_title"])
                        new_centroids_df = pd.DataFrame(new_centroids_list, columns=[t["km_col_label"], "X", "Y"])
                        st.table(new_centroids_df.style.format({"X": "{:.2f}", "Y": "{:.2f}"}))
                        
                    with col_nc2:
                        st.subheader(t["km_plot_title"])
                        fig_km = plgo.Figure()
                        
                        unique_clusters = sorted(list(set(closest_clusters)))
                        colors = px.colors.qualitative.Plotly
                        
                        for idx_cl, cl in enumerate(unique_clusters):
                            cl_mask = [c == cl for c in closest_clusters]
                            cl_points = points_clean[cl_mask]
                            fig_km.add_trace(plgo.Scatter(
                                x=cl_points["X"],
                                y=cl_points["Y"],
                                mode='markers',
                                name=f"{t['km_col_cluster']} {cl}",
                                marker=dict(size=10, color=colors[idx_cl % len(colors)]),
                                text=cl_points[t["km_col_label"]],
                                hovertemplate="<b>%{text}</b><br>X: %{x}<br>Y: %{y}<extra></extra>"
                            ))
                            
                        # Add initial centroids
                        fig_km.add_trace(plgo.Scatter(
                            x=centroids_clean["X"],
                            y=centroids_clean["Y"],
                            mode='markers+text',
                            name=t["km_legend_initial_centroids"],
                            marker=dict(size=14, symbol='x', color='black', line=dict(width=2)),
                            text=centroids_clean[t["km_col_label"]],
                            textposition="top center",
                            hovertemplate="<b>Inicial: %{text}</b><br>X: %{x}<br>Y: %{y}<extra></extra>"
                        ))
                        
                        # Add updated centroids
                        new_c_x = [nc[1] for nc in new_centroids_list]
                        new_c_y = [nc[2] for nc in new_centroids_list]
                        new_c_labels = [nc[0] for nc in new_centroids_list]
                        
                        fig_km.add_trace(plgo.Scatter(
                            x=new_c_x,
                            y=new_c_y,
                            mode='markers+text',
                            name=t["km_legend_new_centroids"],
                            marker=dict(size=14, symbol='star', color='red'),
                            text=new_c_labels,
                            textposition="bottom center",
                            hovertemplate="<b>Nuevo: %{text}</b><br>X: %{x}<br>Y: %{y}<extra></extra>"
                        ))
                        
                        fig_km.update_layout(
                            xaxis_title="X",
                            yaxis_title="Y",
                            template='plotly_white',
                            height=400,
                            margin=dict(l=20, r=20, t=30, b=20),
                            legend=dict(
                                yanchor="top",
                                y=0.99,
                                xanchor="left",
                                x=0.01
                            )
                        )
                        st.plotly_chart(fig_km, use_container_width=True)

        with tab_kvalue:
            st.subheader(t["kv_title"])
            st.write(t["kv_intro"])
            
            points_data = None
            if "km_df_points_df" in st.session_state:
                points_data = st.session_state["km_df_points_df"]
            elif "df_points" in locals() and df_points is not None:
                points_data = df_points
                
            if points_data is not None:
                points_clean = points_data.dropna(subset=["X", "Y"])
                
                if len(points_clean) >= 3:
                    from sklearn.cluster import KMeans
                    from sklearn.metrics import silhouette_score
                    
                    # Ensure numeric data types
                    points_clean = points_clean.copy()
                    points_clean["X"] = pd.to_numeric(points_clean["X"], errors='coerce')
                    points_clean["Y"] = pd.to_numeric(points_clean["Y"], errors='coerce')
                    points_clean = points_clean.dropna(subset=["X", "Y"])
                    
                    X_clust = points_clean[["X", "Y"]].values
                    N_points = len(X_clust)
                    
                    k_max = min(8, N_points - 1)
                    k_values = list(range(1, k_max + 1))
                    
                    wcss = []
                    sil_scores = []
                    
                    for k in k_values:
                        try:
                            kmeans_model = KMeans(n_clusters=k, random_state=42, n_init=10)
                            kmeans_model.fit(X_clust)
                            wcss.append(kmeans_model.inertia_)
                            
                            if k > 1:
                                if len(set(kmeans_model.labels_)) > 1:
                                    score = silhouette_score(X_clust, kmeans_model.labels_)
                                    sil_scores.append(score)
                                else:
                                    sil_scores.append(0.0)
                            else:
                                sil_scores.append(0.0)
                        except Exception:
                            wcss.append(0.0)
                            sil_scores.append(0.0)
                            
                    metrics_data = {
                        t["kv_col_k"]: k_values,
                        t["kv_col_wcss"]: [f"{w:.4f}" for w in wcss]
                    }
                    if len(k_values) > 1:
                        metrics_data[t["kv_col_silhouette"]] = [f"{s:.4f}" if k > 1 else "N/A" for k, s in zip(k_values, sil_scores)]
                        
                    metrics_df = pd.DataFrame(metrics_data)
                    
                    col_kv_left, col_kv_right = st.columns([1, 1])
                    
                    with col_kv_left:
                        st.markdown(f"### 🎯 {t['kv_elbow_title']}")
                        st.write(t["kv_elbow_desc"])
                        
                        st.markdown(f"### 👤 {t['kv_silhouette_title']}")
                        st.write(t["kv_silhouette_desc"])
                        
                        if len(sil_scores) > 1:
                            valid_sil_scores = sil_scores[1:]
                            valid_k_values = k_values[1:]
                            if valid_sil_scores:
                                max_idx = np.argmax(valid_sil_scores)
                                opt_k = valid_k_values[max_idx]
                                st.markdown(f"### {t['kv_optimal_title']}")
                                st.info(t["kv_optimal_desc"].format(opt_k))
                                
                        st.markdown(f"#### {t['kv_table_title']}")
                        st.dataframe(metrics_df, use_container_width=True, hide_index=True)
                        
                    with col_kv_right:
                        st.markdown(f"### {t['kv_plot_title']}")
                        fig_kv = plgo.Figure()
                        
                        fig_kv.add_trace(plgo.Scatter(
                            x=k_values,
                            y=wcss,
                            mode='lines+markers',
                            name=t["kv_col_wcss"],
                            marker=dict(color='#1f77b4', size=8),
                            line=dict(color='#1f77b4', width=2),
                            yaxis="y1"
                        ))
                        
                        if len(k_values) > 1:
                            fig_kv.add_trace(plgo.Scatter(
                                x=k_values[1:],
                                y=sil_scores[1:],
                                mode='lines+markers',
                                name=t["kv_col_silhouette"],
                                marker=dict(color='orange', size=8),
                                line=dict(color='orange', width=2),
                                yaxis="y2"
                            ))
                            
                        fig_kv.update_layout(
                            xaxis=dict(title=t["kv_col_k"], tickmode='linear', tick0=1, dtick=1),
                            yaxis=dict(
                                title=dict(
                                    text=t["kv_col_wcss"],
                                    font=dict(color="#1f77b4")
                                ),
                                tickfont=dict(color="#1f77b4")
                            ),
                            yaxis2=dict(
                                title=dict(
                                    text=t["kv_col_silhouette"],
                                    font=dict(color="orange")
                                ),
                                tickfont=dict(color="orange"),
                                overlaying="y",
                                side="right"
                            ),
                            template='plotly_white',
                            height=500,
                            margin=dict(l=20, r=20, t=30, b=20),
                            legend=dict(
                                yanchor="top",
                                y=0.99,
                                xanchor="left",
                                x=0.01
                            )
                        )
                        st.plotly_chart(fig_kv, use_container_width=True)
                else:
                    st.warning(t["kv_no_points_warn"])
            else:
                st.warning(t["kv_no_points_warn"])
            
        with tab_elbow:
            st.subheader(t["el_title"])
            st.write(t["el_intro"])
            
            points_data = None
            if "km_df_points_df" in st.session_state:
                points_data = st.session_state["km_df_points_df"]
            elif "df_points" in locals() and df_points is not None:
                points_data = df_points
                
            if points_data is not None:
                points_clean = points_data.dropna(subset=["X", "Y"])
                
                if len(points_clean) >= 3:
                    from sklearn.cluster import KMeans
                    
                    # Ensure numeric data types
                    points_clean = points_clean.copy()
                    points_clean["X"] = pd.to_numeric(points_clean["X"], errors='coerce')
                    points_clean["Y"] = pd.to_numeric(points_clean["Y"], errors='coerce')
                    points_clean = points_clean.dropna(subset=["X", "Y"])
                    
                    X_clust = points_clean[["X", "Y"]].values
                    N_points = len(X_clust)
                    
                    # Calculate TSS (Total Sum of Squares)
                    mean_x = np.mean(X_clust[:, 0])
                    mean_y = np.mean(X_clust[:, 1])
                    tss_val = np.sum((X_clust[:, 0] - mean_x)**2 + (X_clust[:, 1] - mean_y)**2)
                    
                    k_max = min(8, N_points - 1)
                    k_values = list(range(1, k_max + 1))
                    
                    wcss = []
                    bcss = []
                    var_explained = []
                    
                    for k in k_values:
                        try:
                            kmeans_model = KMeans(n_clusters=k, random_state=42, n_init=10)
                            kmeans_model.fit(X_clust)
                            w_val = kmeans_model.inertia_
                            wcss.append(w_val)
                            
                            b_val = max(0.0, tss_val - w_val)
                            bcss.append(b_val)
                            
                            ve_val = (b_val / tss_val * 100.0) if tss_val > 0 else 0.0
                            var_explained.append(ve_val)
                        except Exception:
                            wcss.append(0.0)
                            bcss.append(0.0)
                            var_explained.append(0.0)
                            
                    # Build metrics dataframe
                    metrics_data = {
                        t["el_col_k"]: k_values,
                        t["el_col_wcss"]: [f"{w:.4f}" for w in wcss],
                        t["el_col_bcss"]: [f"{b:.4f}" for b in bcss],
                        t["el_col_tss"]: [f"{tss_val:.4f}"] * len(k_values),
                        t["el_col_var"]: [f"{v:.2f}%" for v in var_explained]
                    }
                    metrics_df = pd.DataFrame(metrics_data)
                    
                    # Layout: 2 Columns (Formulas vs Table)
                    col_el_left, col_el_right = st.columns([1, 1])
                    
                    with col_el_left:
                        st.markdown(f"### 📐 {t['el_formula_wcss_title']}")
                        st.latex(r"WCSS = \sum_{j=1}^{K} \sum_{i \in S_j} \| x_i - \mu_j \|^2")
                        st.write(t["el_formula_wcss_desc"])
                        
                        st.markdown(f"### 🌐 {t['el_formula_tss_title']}")
                        st.latex(r"TSS = \sum_{i=1}^{N} \| x_i - \bar{x} \|^2")
                        st.write(t["el_formula_tss_desc"])
                        
                    with col_el_right:
                        st.markdown(f"### 🔀 {t['el_formula_bcss_title']}")
                        st.latex(r"BCSS = TSS - WCSS = \sum_{j=1}^{K} |S_j| \| \mu_j - \bar{x} \|^2")
                        st.write(t["el_formula_bcss_desc"])
                        
                        st.markdown(f"### 📈 {t['el_formula_var_title']}")
                        st.latex(r"\eta^2 = \frac{BCSS}{TSS} \times 100\%")
                        st.write(t["el_formula_var_desc"])
                        
                    st.markdown("---")
                    col_bottom_left, col_bottom_right = st.columns([1, 1])
                    
                    with col_bottom_left:
                        st.subheader(t["el_table_title"])
                        st.dataframe(metrics_df, use_container_width=True, hide_index=True)
                        
                    with col_bottom_right:
                        st.subheader(t["el_plot_title"])
                        fig_el = plgo.Figure()
                        
                        fig_el.add_trace(plgo.Scatter(
                            x=k_values,
                            y=wcss,
                            mode='lines+markers',
                            name="WCSS",
                            marker=dict(color='#1f77b4', size=8),
                            line=dict(color='#1f77b4', width=2)
                        ))
                        
                        fig_el.update_layout(
                            xaxis=dict(title=t["el_plot_x"], tickmode='linear', tick0=1, dtick=1),
                            yaxis=dict(
                                title=dict(
                                    text=t["el_plot_y"],
                                    font=dict(color="#1f77b4")
                                ),
                                tickfont=dict(color="#1f77b4")
                            ),
                            template='plotly_white',
                            height=380,
                            margin=dict(l=20, r=20, t=30, b=20),
                            showlegend=False
                        )
                        st.plotly_chart(fig_el, use_container_width=True)
                else:
                    st.warning(t["kv_no_points_warn"])
            else:
                st.warning(t["kv_no_points_warn"])
            
        with tab_silhouette:
            st.subheader(t["sil_title"])
            st.write(t["sil_intro"])
            st.markdown(t["sil_interpretation"])
            
            st.markdown("---")
            st.latex(r"s(i) = \frac{b(i) - a(i)}{\max\{a(i), b(i)\}}")
            st.markdown("""
            * **$a(i)$**: Distancia media entre el punto $i$ y todos los demás puntos en el **mismo** cluster (Cohesión intra-cluster).
            * **$b(i)$**: Distancia media entre el punto $i$ y todos los puntos en el **cluster vecino más cercano** (Separación inter-cluster).
            """)
            st.info("💡 **Nota:** Puedes observar la aplicación práctica del cálculo del Coeficiente de Silueta en la pestaña **K-value**, donde se utiliza para determinar el número óptimo de clusters (K).")
            
        with tab_dbscan:
            st.subheader(t["db_title"])
            st.write(t["db_intro"])
            
            st.markdown(f"### {t['db_params_title']}")
            st.markdown(t["db_eps"])
            st.markdown(t["db_minpts"])
            st.markdown(t["db_advantages"])
            
            st.markdown("---")
            st.markdown("### 1. Entrada de Datos DBSCAN")
            st.write("Inserta o modifica las coordenadas (X e Y) de los puntos:")
            
            if "db_df_points" not in st.session_state:
                st.session_state.db_df_points = pd.DataFrame({
                    "Punto": [f"P{i}" for i in range(1, 13)],
                    "X": [3.0, 4.0, 5.0, 6.0, 7.0, 6.0, 7.0, 8.0, 3.0, 2.0, 3.0, 2.0],
                    "Y": [7.0, 6.0, 5.0, 4.0, 3.0, 2.0, 2.0, 4.0, 3.0, 6.0, 5.0, 4.0]
                })
                
            edited_db_df = st.data_editor(
                st.session_state.db_df_points,
                num_rows="dynamic",
                key="db_data_editor",
                use_container_width=True,
                on_change=update_session_df,
                args=("db_df_points", "db_data_editor")
            )
            
            points_clean = edited_db_df.dropna(subset=["X", "Y"])
            
            if len(points_clean) >= 3:
                from sklearn.cluster import DBSCAN
                from sklearn.metrics import pairwise_distances
                
                points_clean = points_clean.copy()
                points_clean["X"] = pd.to_numeric(points_clean["X"], errors='coerce')
                points_clean["Y"] = pd.to_numeric(points_clean["Y"], errors='coerce')
                points_clean = points_clean.dropna(subset=["X", "Y"])
                
                X_clust = points_clean[["X", "Y"]].values
                point_labels = points_clean["Punto"].values
                
                st.markdown("### 2. Parámetros y Ejecución")
                col_eps, col_minpts, col_btn = st.columns([1, 1, 1])
                with col_eps:
                    eps_val = st.number_input("Epsilon (eps)", min_value=0.1, max_value=100.0, value=2.0, step=0.5, key="db_eps_input")
                with col_minpts:
                    minpts_val = st.number_input("MinPts", min_value=2, max_value=20, value=4, step=1, key="db_minpts_input")
                with col_btn:
                    st.write("")
                    st.write("")
                    run_dbscan = st.button(t["db_btn_run"], key="btn_run_dbscan")
                
                if run_dbscan:
                    # 1. DBSCAN Model
                    dbscan_model = DBSCAN(eps=eps_val, min_samples=minpts_val)
                    labels = dbscan_model.fit_predict(X_clust)
                    points_clean["Cluster"] = labels
                    
                    # 2. Compute Distances and Neighborhoods
                    dist_matrix = pairwise_distances(X_clust)
                    neighborhoods = dist_matrix <= eps_val
                    density_counts = np.sum(neighborhoods, axis=1)
                    
                    # 3. Create Full Detailed Table
                    full_df = points_clean[["Punto", "X", "Y"]].copy()
                    
                    # Add distance columns
                    for i, p_label in enumerate(point_labels):
                        full_df[f"Dist a {p_label}"] = dist_matrix[:, i]
                        
                    # Add properties
                    full_df["Densidad (puntos <= eps)"] = density_counts
                    full_df["Core?"] = np.where(density_counts >= minpts_val, "CORE", "NO CORE")
                    
                    point_types = []
                    for idx, label in enumerate(labels):
                        if label == -1:
                            point_types.append("NOISE")
                        elif density_counts[idx] >= minpts_val:
                            point_types.append("CORE")
                        else:
                            point_types.append("BORDER")
                            
                    full_df["Tipo (DBSCAN)"] = point_types
                    full_df["Cluster ID"] = labels
                    
                    st.markdown("### 3. Matriz de Distancias y Resultados Detallados")
                    st.dataframe(full_df, use_container_width=True, hide_index=True)
                    
                    st.subheader(t["db_plot_title"])
                    fig_db = plgo.Figure()
                    
                    unique_labels = set(labels)
                    for label in unique_labels:
                        cluster_points = points_clean[points_clean["Cluster"] == label]
                        if label == -1:
                            fig_db.add_trace(plgo.Scatter(
                                x=cluster_points["X"],
                                y=cluster_points["Y"],
                                mode="markers",
                                name=t["db_noise_legend"],
                                marker=dict(color="black", size=8, symbol="x")
                            ))
                        else:
                            fig_db.add_trace(plgo.Scatter(
                                x=cluster_points["X"],
                                y=cluster_points["Y"],
                                mode="markers",
                                name=f"{t['db_cluster_legend']} {label}",
                                marker=dict(size=10)
                            ))
                            
                    fig_db.update_layout(
                        xaxis_title="X",
                        yaxis_title="Y",
                        template="plotly_white",
                        height=500,
                        margin=dict(l=20, r=20, t=30, b=20)
                    )
                    st.plotly_chart(fig_db, use_container_width=True)
            else:
                st.warning(t["db_no_points_warn"])
                
        st.markdown("---")
        st.markdown(
            "<p style='text-align: center; color: gray; font-size: 14px;'>"
            "Desarrollado y mantenido por <b>Alexander Acosta</b> "
            "(<a href='https://github.com/j-alexander-acosta' target='_blank' style='color: #1f77b4; text-decoration: none;'>@j-alexander-acosta</a>)"
            "</p>", 
            unsafe_allow_html=True
        )
        st.stop()

    st.sidebar.markdown("---")
    st.sidebar.header(t["sb_header_params"])
    
    # Obtener valores predeterminados o detectados para errores de medición
    val_sig_u = st.session_state.get("detected_sigma_u", 0.2000)
    val_sig_e = st.session_state.get("detected_sigma_e", 0.1500)
    
    sigma_u = st.sidebar.number_input(
        t["sb_sigma_u"],
        value=val_sig_u,
        step=0.05,
        format="%.4f",
        key="sig_u_input"
    )
    sigma_e = st.sidebar.number_input(
        t["sb_sigma_e"],
        value=val_sig_e,
        step=0.05,
        format="%.4f",
        key="sig_e_input"
    )
    
    eta = (sigma_e ** 2) / (sigma_u ** 2) if sigma_u != 0.0 else 1.0
    st.sidebar.markdown(f"**Relación de varianzas (Calculado η):** `{eta:.4f}`" if st.session_state.lang == 'es' else f"**Variance ratio (Calculated η):** `{eta:.4f}`")

    st.sidebar.markdown("---")
    st.sidebar.header(t["sb_header_nl"])
    
    nl_options = [
        t["nl_models"]["exp"],
        t["nl_models"]["log"],
        t["nl_models"]["pot"],
        t["nl_models"]["quad"]
    ]
    nl_model = st.sidebar.selectbox(
        t["sb_select_nl"],
        nl_options
    )
    
    nl_mapping = {
        t["nl_models"]["exp"]: "Exponential",
        t["nl_models"]["log"]: "Logarithmic",
        t["nl_models"]["pot"]: "Power",
        t["nl_models"]["quad"]: "Quadratic"
    }
    selected_nl_key = nl_mapping[nl_model]
    
    nl_short_names = {
        "Exponential": t["nl_models"]["exp"].split(' ')[0],
        "Logarithmic": t["nl_models"]["log"].split(' ')[0],
        "Power": t["nl_models"]["pot"].split(' ')[0],
        "Quadratic": t["nl_models"]["quad"].split(' ')[0]
    }
    selected_nl_short = nl_short_names[selected_nl_key]
    
    a_val = 0.0
    if selected_nl_key == "Exponential":
        Y_temp = df["Y"].values
        min_y = float(np.min(Y_temp))
        default_a = 5.0 if min_y > 5.0 else float(np.round(min_y - 1.0, 2))
        a_val = st.sidebar.number_input(t["sb_asymptote"], value=default_a, step=0.1, format="%.4f")
        if a_val >= min_y:
            st.sidebar.error(t["sb_asymptote_err"])




    # --- SECCIÓN DE CÁLCULOS GLOBALES ---
    # Cálculo estricto de dimensiones (n)
    n = len(df)
    X = df["X"].values
    Y = df["Y"].values
    
    x_mean = np.mean(X)
    y_mean = np.mean(Y)
    
    # Sumas de cuadrados y productos cruzados
    Sxx = np.sum((X - x_mean)**2)
    Syy = np.sum((Y - y_mean)**2)
    Sxy = np.sum((X - x_mean)*(Y - y_mean))

    # Calcular usando la nueva clase helper para variables con error
    model_global = LinearErrorsInVariables(X, Y, sigma_u, sigma_e)
    metrics_global = model_global.compute_all_metrics()
    
    # ==================================================
    # MÓDULO 1: SLR (Standard Linear Regression / OLS)
    # ==================================================
    m_slr = metrics_global["ols"]["slope"]
    c_slr = metrics_global["ols"]["intercept"]
    y_pred_slr = metrics_global["ols"]["y_pred"]
    sse_slr = np.sum((Y - y_pred_slr)**2)
    s2_e_slr = sse_slr / (n - 2) if n > 2 else 0
    se_m_slr = metrics_global["ols"]["se_slope"]
    se_c_slr = metrics_global["ols"]["se_intercept"]
    rmse_slr = metrics_global["ols"]["rmse"]
    r2_slr = metrics_global["ols"]["r2"]

    # ==================================================
    # MÓDULO 2: GOR Convencional
    # ==================================================
    m_gor = metrics_global["deming"]["slope"]
    b_gor = metrics_global["deming"]["intercept"]
    y_pred_gor = metrics_global["deming"]["y_pred"]
    X_t = metrics_global["deming"]["X_t"]
    Y_t = metrics_global["deming"]["Y_t"]
    s2_e_gor = metrics_global["deming"]["s2_e_ort"]
    sigma_gor = metrics_global["deming"]["sigma_ort"]
    sse_gor = np.sum((Y - y_pred_gor)**2)
    s2_e_pseudo_gor = sse_gor / (n - 2) if n > 2 else 0
    se_m_gor = metrics_global["deming"]["se_slope"]
    se_b_gor = metrics_global["deming"]["se_intercept"]
    rmse_gor = metrics_global["deming"]["rmse"]
    r2_gor = metrics_global["deming"]["r2"]

    # ==================================================
    # MÓDULO 3: GOR Propuesto (Ranjit Das et al.)
    # ==================================================
    Y_t_mean = np.mean(Y_t)
    if Sxx != 0:
        m_prop = np.sum((X - x_mean) * (Y_t - Y_t_mean)) / Sxx
    else:
        m_prop = 0
    b_prop = Y_t_mean - m_prop * x_mean
    y_pred_prop = m_prop * X + b_prop

    sse_prop = np.sum((Y - y_pred_prop)**2)
    s2_e_prop = sse_prop / (n - 2) if n > 2 else 0
    se_m_prop = np.sqrt(s2_e_prop / Sxx) if Sxx != 0 and n > 2 else 0
    se_b_prop = se_m_prop * np.sqrt(np.sum(X**2) / n) if n > 0 else 0
    rmse_prop = np.sqrt(sse_prop / n)
    r2_prop = 1 - (sse_prop / Syy) if Syy != 0 else 0

    # ==================================================
    # MÓDULO ADICIONAL: Método de Momentos (MoM)
    # ==================================================
    m_mom = metrics_global["mom"]["slope"]
    b_mom = metrics_global["mom"]["intercept"]
    y_pred_mom = metrics_global["mom"]["y_pred"]
    se_m_mom = metrics_global["mom"]["se_slope"]
    se_b_mom = metrics_global["mom"]["se_intercept"]
    rmse_mom = metrics_global["mom"]["rmse"]
    r2_mom = metrics_global["mom"]["r2"]
    lambda_ratio = metrics_global["mom"]["lambda"]

    # ==================================================
    # MÓDULO 4: Regresión No Lineal (Cálculos Globales para Todos los Modelos)
    # ==================================================
    
    # 4.1 Modelo Exponencial con Asíntota (Y = a + b * e^(cx))
    exp_asymp_valid = True
    exp_asymp_error_msg = ""
    a_nl_exp, b_nl_exp, c_nl_exp = 0.0, 0.0, 0.0
    A_nl_exp = 0.0
    y_minus_a = None
    y_trans = None
    sum_x = 0.0
    sum_x2 = 0.0
    sum_y_trans = 0.0
    sum_x_y_trans = 0.0
    r2_linearized_exp = 0.0
    y_pred_nl_exp = None
    sse_nl_exp = None
    rmse_nl_exp = None
    r2_nl_exp = None
    nl_equation_exp = ""

    try:
        a_nl_exp = a_val
        min_y = np.min(Y)
        if a_nl_exp >= min_y:
            exp_asymp_valid = False
            if st.session_state.lang == 'es':
                exp_asymp_error_msg = f"La asíntota a ({a_nl_exp:.4f}) debe ser estrictamente menor que el valor mínimo de Y ({min_y:.4f}) para calcular ln(y - a)."
            else:
                exp_asymp_error_msg = f"Asymptote a ({a_nl_exp:.4f}) must be strictly less than the minimum value of Y ({min_y:.4f}) to calculate ln(y - a)."
        else:
            y_minus_a = Y - a_nl_exp
            y_trans = np.log(y_minus_a)
            
            sum_x = np.sum(X)
            sum_x2 = np.sum(X**2)
            sum_y_trans = np.sum(y_trans)
            sum_x_y_trans = np.sum(X * y_trans)
            
            den = n * sum_x2 - sum_x**2
            if den != 0:
                c_nl_exp = (n * sum_x_y_trans - sum_x * sum_y_trans) / den
                A_nl_exp = (sum_y_trans - c_nl_exp * sum_x) / n
            else:
                c_nl_exp = 0.0
                A_nl_exp = np.mean(y_trans)
            
            b_nl_exp = np.exp(A_nl_exp)
            y_pred_nl_exp = a_nl_exp + b_nl_exp * np.exp(c_nl_exp * X)
            
            y_trans_mean = np.mean(y_trans)
            S_y_trans_y_trans = np.sum((y_trans - y_trans_mean)**2)
            y_trans_pred = c_nl_exp * X + A_nl_exp
            sse_linearized_exp = np.sum((y_trans - y_trans_pred)**2)
            r2_linearized_exp = 1.0 - (sse_linearized_exp / S_y_trans_y_trans) if S_y_trans_y_trans != 0 else 1.0
            
            sse_nl_exp = np.sum((Y - y_pred_nl_exp)**2)
            rmse_nl_exp = np.sqrt(sse_nl_exp / n)
            r2_nl_exp = 1 - (sse_nl_exp / Syy) if Syy != 0 else 0
            
            sign_b = "+" if b_nl_exp >= 0 else "-"
            nl_equation_exp = rf"y = {a_nl_exp:.4f} {sign_b} {abs(b_nl_exp):.4f}e^{{{c_nl_exp:.4f}x}}"
    except Exception as e:
        exp_asymp_valid = False
        if st.session_state.lang == 'es':
            exp_asymp_error_msg = f"Error en los cálculos del modelo Exponencial con Asíntota: {e}"
        else:
            exp_asymp_error_msg = f"Error in Exponential with Asymptote model calculations: {e}"

    # 4.2 Modelo Logarítmico (Y = a + b * ln(x))
    log_valid = True
    log_error_msg = ""
    a_nl_log, b_nl_log = 0.0, 0.0
    y_pred_nl_log = None
    sse_nl_log = None
    rmse_nl_log = None
    r2_nl_log = None
    nl_equation_log = ""

    if np.any(X <= 0):
        log_valid = False
        if st.session_state.lang == 'es':
            log_error_msg = "El modelo Logarítmico requiere que todos los valores de X sean estrictamente mayores a cero (X > 0) para calcular ln(X)."
        else:
            log_error_msg = "Logarithmic model requires all X values to be strictly greater than zero (X > 0) to calculate ln(X)."
    else:
        try:
            ln_X = np.log(X)
            ln_X_mean = np.mean(ln_X)
            S_lnX_lnX = np.sum((ln_X - ln_X_mean)**2)
            S_lnX_Y = np.sum((ln_X - ln_X_mean) * (Y - y_mean))
            b_nl_log = S_lnX_Y / S_lnX_lnX if S_lnX_lnX != 0 else 0
            a_nl_log = y_mean - b_nl_log * ln_X_mean
            y_pred_nl_log = a_nl_log + b_nl_log * np.log(X)
            nl_equation_log = rf"y = {a_nl_log:.4f} + {b_nl_log:.4f}\ln(x)"
            
            sse_nl_log = np.sum((Y - y_pred_nl_log)**2)
            rmse_nl_log = np.sqrt(sse_nl_log / n)
            r2_nl_log = 1 - (sse_nl_log / Syy) if Syy != 0 else 0
        except Exception as e:
            log_valid = False
            if st.session_state.lang == 'es':
                log_error_msg = f"Error en los cálculos del modelo Logarítmico: {e}"
            else:
                log_error_msg = f"Error in Logarithmic model calculations: {e}"

    # 4.3 Modelo Potencial / Power Law (Y = a * x^b) usando ln
    pot_valid = True
    pot_error_msg = ""
    a_nl_pot, b_nl_pot = 0.0, 0.0
    A_nl_pot = 0.0
    y_pred_nl_pot = None
    sse_nl_pot = None
    rmse_nl_pot = None
    r2_nl_pot = None
    nl_equation_pot = ""

    if np.any(X <= 0) or np.any(Y <= 0):
        pot_valid = False
        if st.session_state.lang == 'es':
            pot_error_msg = "El modelo Potencial requiere que todos los valores tanto de X como de Y sean estrictamente mayores a cero (X > 0, Y > 0) para aplicar ln(X) y ln(Y)."
        else:
            pot_error_msg = "Power model requires all values of both X and Y to be strictly greater than zero (X > 0, Y > 0) to apply ln(X) and ln(Y)."
    else:
        try:
            ln_X = np.log(X)
            ln_Y = np.log(Y)
            ln_X_mean = np.mean(ln_X)
            ln_Y_mean = np.mean(ln_Y)
            S_lnX_lnX = np.sum((ln_X - ln_X_mean)**2)
            S_lnX_lnY = np.sum((ln_X - ln_X_mean) * (ln_Y - ln_Y_mean))
            b_nl_pot = S_lnX_lnY / S_lnX_lnX if S_lnX_lnX != 0 else 0
            A_nl_pot = ln_Y_mean - b_nl_pot * ln_X_mean
            a_nl_pot = np.exp(A_nl_pot)
            y_pred_nl_pot = a_nl_pot * (X ** b_nl_pot)
            nl_equation_pot = rf"y = {a_nl_pot:.4f}x^{{{b_nl_pot:.4f}}}"
            
            sse_nl_pot = np.sum((Y - y_pred_nl_pot)**2)
            rmse_nl_pot = np.sqrt(sse_nl_pot / n)
            r2_nl_pot = 1 - (sse_nl_pot / Syy) if Syy != 0 else 0
        except Exception as e:
            pot_valid = False
            if st.session_state.lang == 'es':
                pot_error_msg = f"Error en los cálculos del modelo Potencial: {e}"
            else:
                pot_error_msg = f"Error in Power model calculations: {e}"

    # 4.4 Modelo Cuadrático (Y = ax^2 + bx + c)
    quad_valid = True
    quad_error_msg = ""
    a_nl_quad, b_nl_quad, c_nl_quad = 0.0, 0.0, 0.0
    y_pred_nl_quad = None
    sse_nl_quad = None
    rmse_nl_quad = None
    r2_nl_quad = None
    nl_equation_quad = ""

    if n < 3:
        quad_valid = False
        if st.session_state.lang == 'es':
            quad_error_msg = "El modelo Cuadrático requiere al menos 3 puntos de datos para calcular un ajuste único. Por favor, agrega más puntos en la barra lateral."
        else:
            quad_error_msg = "Quadratic model requires at least 3 data points to calculate a unique fit. Please add more points in the sidebar."
    else:
        try:
            coefs = np.polyfit(X, Y, 2)
            a_nl_quad, b_nl_quad, c_nl_quad = coefs[0], coefs[1], coefs[2]
            y_pred_nl_quad = a_nl_quad * (X**2) + b_nl_quad * X + c_nl_quad
            nl_equation_quad = rf"y = {a_nl_quad:.4f}x^2 + {b_nl_quad:.4f}x + {c_nl_quad:.4f}"
            
            sse_nl_quad = np.sum((Y - y_pred_nl_quad)**2)
            rmse_nl_quad = np.sqrt(sse_nl_quad / n)
            r2_nl_quad = 1 - (sse_nl_quad / Syy) if Syy != 0 else 0
        except Exception as e:
            quad_valid = False
            if st.session_state.lang == 'es':
                quad_error_msg = f"Error en los cálculos del modelo Cuadrático: {e}"
            else:
                quad_error_msg = f"Error in Quadratic model calculations: {e}"

    # Asignar variables del modelo no lineal seleccionado en la barra lateral (para Tab 4)
    nl_valid = True
    nl_error_msg = ""
    y_pred_nl = None
    nl_equation = ""
    a_nl, b_nl, c_nl = 0.0, 0.0, 0.0
    A_nl = 0.0
    sse_nl = None
    rmse_nl = None
    r2_nl = None

    if selected_nl_key == "Exponential":
        nl_valid = exp_asymp_valid
        nl_error_msg = exp_asymp_error_msg
        a_nl, b_nl, c_nl = a_nl_exp, b_nl_exp, c_nl_exp
        A_nl = A_nl_exp
        y_pred_nl = y_pred_nl_exp
        nl_equation = nl_equation_exp
        sse_nl, rmse_nl, r2_nl = sse_nl_exp, rmse_nl_exp, r2_nl_exp
        # variables específicas para la tabla de sumatorias en tab 4
        if exp_asymp_valid:
            y_minus_a = Y - a_nl_exp
            y_trans = np.log(y_minus_a)
            sum_x = np.sum(X)
            sum_x2 = np.sum(X**2)
            sum_y_trans = np.sum(y_trans)
            sum_x_y_trans = np.sum(X * y_trans)
            r2_linearized = r2_linearized_exp
    elif selected_nl_key == "Logarithmic":
        nl_valid = log_valid
        nl_error_msg = log_error_msg
        a_nl, b_nl = a_nl_log, b_nl_log
        y_pred_nl = y_pred_nl_log
        nl_equation = nl_equation_log
        sse_nl, rmse_nl, r2_nl = sse_nl_log, rmse_nl_log, r2_nl_log
    elif selected_nl_key == "Power":
        nl_valid = pot_valid
        nl_error_msg = pot_error_msg
        a_nl, b_nl = a_nl_pot, b_nl_pot
        A_nl = A_nl_pot
        y_pred_nl = y_pred_nl_pot
        nl_equation = nl_equation_pot
        sse_nl, rmse_nl, r2_nl = sse_nl_pot, rmse_nl_pot, r2_nl_pot
    elif selected_nl_key == "Quadratic":
        nl_valid = quad_valid
        nl_error_msg = quad_error_msg
        a_nl, b_nl, c_nl = a_nl_quad, b_nl_quad, c_nl_quad
        y_pred_nl = y_pred_nl_quad
        nl_equation = nl_equation_quad
        sse_nl, rmse_nl, r2_nl = sse_nl_quad, rmse_nl_quad, r2_nl_quad

    # ==================================================
    # MÓDULO 5: Power Law Model (Duplicate de Potencial para Tab 5)
    # ==================================================
    power_valid = pot_valid
    power_error_msg = pot_error_msg
    a_power, b_power = a_nl_pot, b_nl_pot
    A_power = A_nl_pot
    y_pred_power = y_pred_nl_pot
    power_equation = nl_equation_pot
    sse_power = sse_nl_pot
    rmse_power = rmse_nl_pot
    r2_power = r2_nl_pot
    x_trans_power = np.log(X) if pot_valid else None
    y_trans_power = np.log(Y) if pot_valid else None
    sum_x_trans_power = np.sum(x_trans_power) if pot_valid else 0.0
    sum_x2_trans_power = np.sum(x_trans_power**2) if pot_valid else 0.0
    sum_y_trans_power = np.sum(y_trans_power) if pot_valid else 0.0
    sum_xy_trans_power = np.sum(x_trans_power * y_trans_power) if pot_valid else 0.0
    r2_linearized_power = 0.0
    if pot_valid:
        y_trans_power_mean = np.mean(y_trans_power)
        S_y_trans_power_y_trans_power = np.sum((y_trans_power - y_trans_power_mean)**2)
        y_trans_power_pred = b_power * x_trans_power + A_power
        sse_linearized_power = np.sum((y_trans_power - y_trans_power_pred)**2)
        r2_linearized_power = 1.0 - (sse_linearized_power / S_y_trans_power_y_trans_power) if S_y_trans_power_y_trans_power != 0 else 1.0

    # ==================================================
    # MÓDULO 6: Función Potencia (base 10) (Cálculos Globales)
    # ==================================================
    fun_pot_valid = True
    fun_pot_error_msg = ""
    a_correct, b_correct = 0.0, 0.0
    a_err, b_err = 0.0, 0.0
    m_log, c_log = 0.0, 0.0
    r2_lin_pot10 = 0.0
    X_log, Y_log = None, None
    X_log_mean, Y_log_mean = 0.0, 0.0
    S_xx_log, S_xy_log, S_yy_log = 0.0, 0.0, 0.0
    y_pred_pot10_correct = None
    y_pred_pot10_incorrect = None
    sse_pot10_correct, rmse_pot10_correct, r2_pot10_correct = None, None, None
    sse_pot10_incorrect, rmse_pot10_incorrect, r2_pot10_incorrect = None, None, None

    if np.any(X <= 0) or np.any(Y <= 0):
        fun_pot_valid = False
        if st.session_state.lang == 'es':
            fun_pot_error_msg = "⚠️ El modelo Función Potencia requiere que todos los valores de X e Y sean estrictamente mayores a cero (>0) para aplicar el logaritmo base 10 (log10)."
        else:
            fun_pot_error_msg = "⚠️ The Power Function model requires all X and Y values to be strictly greater than zero (>0) to apply base 10 logarithm (log10)."
    else:
        try:
            X_log = np.log10(X)
            Y_log = np.log10(Y)
            X_log_mean = np.mean(X_log)
            Y_log_mean = np.mean(Y_log)

            S_xx_log = np.sum((X_log - X_log_mean)**2)
            S_xy_log = np.sum((X_log - X_log_mean) * (Y_log - Y_log_mean))
            S_yy_log = np.sum((Y_log - Y_log_mean)**2)

            m_log = S_xy_log / S_xx_log if S_xx_log != 0 else 0.0
            c_log = Y_log_mean - m_log * X_log_mean

            r2_lin_pot10 = (S_xy_log**2) / (S_xx_log * S_yy_log) if (S_xx_log * S_yy_log) != 0 else 0.0

            a_correct = 10**c_log
            b_correct = m_log

            a_err = 10**c_log
            b_err = 10**m_log

            y_pred_pot10_correct = a_correct * (X ** b_correct)
            sse_pot10_correct = np.sum((Y - y_pred_pot10_correct)**2)
            rmse_pot10_correct = np.sqrt(sse_pot10_correct / n)
            r2_pot10_correct = 1 - (sse_pot10_correct / Syy) if Syy != 0 else 0

            y_pred_pot10_incorrect = a_err * (X ** b_err)
            sse_pot10_incorrect = np.sum((Y - y_pred_pot10_incorrect)**2)
            rmse_pot10_incorrect = np.sqrt(sse_pot10_incorrect / n)
            r2_pot10_incorrect = 1 - (sse_pot10_incorrect / Syy) if Syy != 0 else 0
        except Exception as e:
            fun_pot_valid = False
            if st.session_state.lang == 'es':
                fun_pot_error_msg = f"Error en los cálculos de Función Potencia (base 10): {e}"
            else:
                fun_pot_error_msg = f"Error in Power Function (base 10) calculations: {e}"

    # ==================================================
    # MÓDULO 7: Función Exponencial Simple (ln) (Cálculos Globales)
    # ==================================================
    fun_exp_valid = True
    fun_exp_error_msg = ""
    b_coeff, ln_a = 0.0, 0.0
    r2_lin_exp = 0.0
    a_coeff = 0.0
    Y_log_exp = None
    X_mean_exp, Y_log_mean_exp = 0.0, 0.0
    S_xx_exp, S_xy_log_exp, S_yy_log_exp = 0.0, 0.0, 0.0
    y_pred_exp_simple = None
    sse_exp_simple, rmse_exp_simple, r2_exp_simple = None, None, None

    if np.any(Y <= 0):
        fun_exp_valid = False
        if st.session_state.lang == 'es':
            fun_exp_error_msg = "⚠️ El modelo Exponencial requiere que todos los valores de Y sean estrictamente mayores a cero (>0) para aplicar el logaritmo natural (ln)."
        else:
            fun_exp_error_msg = "⚠️ The Exponential model requires all Y values to be strictly greater than zero (>0) to apply natural logarithm (ln)."
    else:
        try:
            Y_log_exp = np.log(Y)
            X_mean_exp = np.mean(X)
            Y_log_mean_exp = np.mean(Y_log_exp)

            S_xx_exp = np.sum((X - X_mean_exp)**2)
            S_xy_log_exp = np.sum((X - X_mean_exp) * (Y_log_exp - Y_log_mean_exp))
            S_yy_log_exp = np.sum((Y_log_exp - Y_log_mean_exp)**2)

            b_coeff = S_xy_log_exp / S_xx_exp if S_xx_exp != 0 else 0.0
            ln_a = Y_log_mean_exp - b_coeff * X_mean_exp

            r2_lin_exp = (S_xy_log_exp**2) / (S_xx_exp * S_yy_log_exp) if (S_xx_exp * S_yy_log_exp) != 0 else 0.0
            a_coeff = np.exp(ln_a)

            y_pred_exp_simple = a_coeff * np.exp(b_coeff * X)
            sse_exp_simple = np.sum((Y - y_pred_exp_simple)**2)
            rmse_exp_simple = np.sqrt(sse_exp_simple / n)
            r2_exp_simple = 1 - (sse_exp_simple / Syy) if Syy != 0 else 0
        except Exception as e:
            fun_exp_valid = False
            if st.session_state.lang == 'es':
                fun_exp_error_msg = f"Error en los cálculos de Función Exponencial: {e}"
            else:
                fun_exp_error_msg = f"Error in Exponential Function calculations: {e}"

    # Construir DataFrame de Resultados
    df_results = df.copy()

    df_results["Y_est (SLR)"] = y_pred_slr
    df_results["Residuo (SLR)"] = Y - y_pred_slr

    df_results["Y_est (GOR Conv)"] = y_pred_gor
    df_results["Residuo (GOR Conv)"] = Y - y_pred_gor
    df_results["X_t (GOR)"] = X_t
    df_results["Y_t (GOR)"] = Y_t

    df_results["Y_est (GOR Prop)"] = y_pred_prop
    df_results["Residuo (GOR Prop)"] = Y - y_pred_prop

    df_results["Y_est (MoM)"] = y_pred_mom
    df_results["Residuo (MoM)"] = Y - y_pred_mom

    if exp_asymp_valid:
        df_results["Y_est (No Lin. Exponencial Asintota)"] = y_pred_nl_exp
        df_results["Residuo (No Lin. Exponencial Asintota)"] = Y - y_pred_nl_exp

    if log_valid:
        df_results["Y_est (No Lin. Logaritmico)"] = y_pred_nl_log
        df_results["Residuo (No Lin. Logaritmico)"] = Y - y_pred_nl_log

    if pot_valid:
        df_results["Y_est (No Lin. Potencial / Power Law)"] = y_pred_nl_pot
        df_results["Residuo (No Lin. Potencial / Power Law)"] = Y - y_pred_nl_pot

    if quad_valid:
        df_results["Y_est (No Lin. Cuadratico)"] = y_pred_nl_quad
        df_results["Residuo (No Lin. Cuadratico)"] = Y - y_pred_nl_quad

    if fun_pot_valid:
        df_results["Y_est (Fun. Potencia log10 - Correcto)"] = y_pred_pot10_correct
        df_results["Residuo (Fun. Potencia log10 - Correcto)"] = Y - y_pred_pot10_correct
        df_results["Y_est (Fun. Potencia log10 - Incorrecto)"] = y_pred_pot10_incorrect
        df_results["Residuo (Fun. Potencia log10 - Incorrecto)"] = Y - y_pred_pot10_incorrect

    if fun_exp_valid:
        df_results["Y_est (Fun. Exponencial Simple ln)"] = y_pred_exp_simple
        df_results["Residuo (Fun. Exponencial Simple ln)"] = Y - y_pred_exp_simple




    # --- DASHBOARD COMPARATIVO CONSOLIDADO ---
    st.header(t["db_title"])
    
    col1, col2 = st.columns([1.2, 1])
    
    with col1:
        # Gráfico Dinámico con Plotly
        fig_plotly = plgo.Figure()

        # Puntos Reales (Negro)
        fig_plotly.add_trace(plgo.Scatter(
            x=X, y=Y, mode='markers', name=t["chart_obs"],
            marker=dict(color='black', size=8),
            hovertemplate='X: %{x}<br>Y: %{y}<extra></extra>'
        ))

        # Líneas de Tendencia (Continuas y nítidas)
        x_line = np.linspace(min(X), max(X), 100)
        
        fig_plotly.add_trace(plgo.Scatter(
            x=x_line, y=m_slr * x_line + c_slr, mode='lines', name=t["chart_slr"],
            line=dict(color='blue', width=2),
            hovertemplate='X: %{x:.2f}<br>Y (SLR): %{y:.2f}<extra></extra>'
        ))
        
        fig_plotly.add_trace(plgo.Scatter(
            x=x_line, y=m_gor * x_line + b_gor, mode='lines', name=t["chart_gor_conv"],
            line=dict(color='orange', width=2),
            hovertemplate='X: %{x:.2f}<br>Y (GOR): %{y:.2f}<extra></extra>'
        ))
        
        fig_plotly.add_trace(plgo.Scatter(
            x=x_line, y=m_prop * x_line + b_prop, mode='lines', name=t["chart_gor_prop"],
            line=dict(color='green', width=2),
            hovertemplate='X: %{x:.2f}<br>Y (Prop): %{y:.2f}<extra></extra>'
        ))

        fig_plotly.add_trace(plgo.Scatter(
            x=x_line, y=m_mom * x_line + b_mom, mode='lines', name=t["methods_names"]["MoM"],
            line=dict(color='magenta', width=2, dash='dashdot'),
            hovertemplate='X: %{x:.2f}<br>Y (MoM): %{y:.2f}<extra></extra>'
        ))

        if nl_valid:
            x_line_nl = np.linspace(min(X), max(X), 200)
            if selected_nl_key == "Exponential":
                y_line_nl = a_nl + b_nl * np.exp(c_nl * x_line_nl)
            elif selected_nl_key == "Logarithmic":
                x_line_safe = np.maximum(x_line_nl, 1e-9)
                y_line_nl = a_nl + b_nl * np.log(x_line_safe)
            elif selected_nl_key == "Power":
                x_line_safe = np.maximum(x_line_nl, 1e-9)
                y_line_nl = a_nl * (x_line_safe ** b_nl)
            elif selected_nl_key == "Quadratic":
                y_line_nl = a_nl * (x_line_nl ** 2) + b_nl * x_line_nl + c_nl
            
            fig_plotly.add_trace(plgo.Scatter(
                x=x_line_nl, y=y_line_nl, mode='lines', name=t["chart_nlin"].format(selected_nl_short),
                line=dict(color='purple', width=2.5, dash='dashdot'),
                hovertemplate='X: %{x:.2f}<br>Y (No Lin.): %{y:.2f}<extra></extra>'
            ))

        if power_valid:
            x_line_power = np.linspace(min(X), max(X), 200)
            x_line_power_safe = np.maximum(x_line_power, 1e-9)
            y_line_power = a_power * (x_line_power_safe ** b_power)
            
            fig_plotly.add_trace(plgo.Scatter(
                x=x_line_power, y=y_line_power, mode='lines', name=t["chart_power"],
                line=dict(color='coral', width=2.5, dash='dash'),
                hovertemplate='X: %{x:.2f}<br>Y (Power Law): %{y:.2f}<extra></extra>'
            ))

        fig_plotly.update_layout(
            title=t["chart_title"],
            xaxis_title=t["chart_x"],
            yaxis_title=t["chart_y"],
            hovermode='closest',
            template='plotly_white',
            legend=dict(yanchor="top", y=0.99, xanchor="left", x=0.01)
        )
        st.plotly_chart(fig_plotly, use_container_width=True)

    with col2:
        # Tabla Comparativa de Parámetros
        comp_data = {
            t["table_col_method"]: [t["methods_names"]["SLR"], t["methods_names"]["GOR Conv"], t["methods_names"]["GOR Prop"], t["methods_names"]["MoM"]],
            t["table_col_slope"]: [m_slr, m_gor, m_prop, m_mom],
            t["table_col_intercept"]: [c_slr, b_gor, b_prop, b_mom],
            t["table_col_se_m"]: [se_m_slr, se_m_gor, se_m_prop, se_m_mom],
            t["table_col_se_c"]: [se_c_slr, se_b_gor, se_b_prop, se_b_mom],
            t["table_col_rmse"]: [rmse_slr, rmse_gor, rmse_prop, rmse_mom],
            t["table_col_r2"]: [r2_slr, r2_gor, r2_prop, r2_mom]
        }
        if exp_asymp_valid:
            comp_data[t["table_col_method"]].append(t["methods_names"]["No Lin. Exponencial Asintota"])
            comp_data[t["table_col_slope"]].append(np.nan)
            comp_data[t["table_col_intercept"]].append(np.nan)
            comp_data[t["table_col_se_m"]].append(np.nan)
            comp_data[t["table_col_se_c"]].append(np.nan)
            comp_data[t["table_col_rmse"]].append(rmse_nl_exp)
            comp_data[t["table_col_r2"]].append(r2_nl_exp)

        if log_valid:
            comp_data[t["table_col_method"]].append(t["methods_names"]["No Lin. Logarítmico"])
            comp_data[t["table_col_slope"]].append(np.nan)
            comp_data[t["table_col_intercept"]].append(np.nan)
            comp_data[t["table_col_se_m"]].append(np.nan)
            comp_data[t["table_col_se_c"]].append(np.nan)
            comp_data[t["table_col_rmse"]].append(rmse_nl_log)
            comp_data[t["table_col_r2"]].append(r2_nl_log)

        if pot_valid:
            comp_data[t["table_col_method"]].append(t["methods_names"]["No Lin. Potencial / Power Law"])
            comp_data[t["table_col_slope"]].append(np.nan)
            comp_data[t["table_col_intercept"]].append(np.nan)
            comp_data[t["table_col_se_m"]].append(np.nan)
            comp_data[t["table_col_se_c"]].append(np.nan)
            comp_data[t["table_col_rmse"]].append(rmse_nl_pot)
            comp_data[t["table_col_r2"]].append(r2_nl_pot)

        if quad_valid:
            comp_data[t["table_col_method"]].append(t["methods_names"]["No Lin. Cuadrático"])
            comp_data[t["table_col_slope"]].append(np.nan)
            comp_data[t["table_col_intercept"]].append(np.nan)
            comp_data[t["table_col_se_m"]].append(np.nan)
            comp_data[t["table_col_se_c"]].append(np.nan)
            comp_data[t["table_col_rmse"]].append(rmse_nl_quad)
            comp_data[t["table_col_r2"]].append(r2_nl_quad)

        if fun_pot_valid:
            comp_data[t["table_col_method"]].append(t["methods_names"]["Fun. Potencia log10 (Correcto)"])
            comp_data[t["table_col_slope"]].append(np.nan)
            comp_data[t["table_col_intercept"]].append(np.nan)
            comp_data[t["table_col_se_m"]].append(np.nan)
            comp_data[t["table_col_se_c"]].append(np.nan)
            comp_data[t["table_col_rmse"]].append(rmse_pot10_correct)
            comp_data[t["table_col_r2"]].append(r2_pot10_correct)

            comp_data[t["table_col_method"]].append(t["methods_names"]["Fun. Potencia log10 (Incorrecto)"])
            comp_data[t["table_col_slope"]].append(np.nan)
            comp_data[t["table_col_intercept"]].append(np.nan)
            comp_data[t["table_col_se_m"]].append(np.nan)
            comp_data[t["table_col_se_c"]].append(np.nan)
            comp_data[t["table_col_rmse"]].append(rmse_pot10_incorrect)
            comp_data[t["table_col_r2"]].append(r2_pot10_incorrect)

        if fun_exp_valid:
            comp_data[t["table_col_method"]].append(t["methods_names"]["Fun. Exponencial Simple ln"])
            comp_data[t["table_col_slope"]].append(np.nan)
            comp_data[t["table_col_intercept"]].append(np.nan)
            comp_data[t["table_col_se_m"]].append(np.nan)
            comp_data[t["table_col_se_c"]].append(np.nan)
            comp_data[t["table_col_rmse"]].append(rmse_exp_simple)
            comp_data[t["table_col_r2"]].append(r2_exp_simple)

        df_comp = pd.DataFrame(comp_data)
        st.markdown(t["table_title"])
        st.dataframe(df_comp.style.format({
            t["table_col_slope"]: "{:.4f}",
            t["table_col_intercept"]: "{:.4f}",
            t["table_col_se_m"]: "{:.4f}",
            t["table_col_se_c"]: "{:.4f}",
            t["table_col_rmse"]: "{:.4f}",
            t["table_col_r2"]: "{:.4f}"
        }, na_rep="-"), use_container_width=True)
        
        diff_pendiente = abs(m_slr - m_prop) / abs(m_slr) * 100 if m_slr != 0 else 0
        st.info(t["insight_text"].format(diff_pendiente))

    st.markdown("---")

    # --- MÓDULOS EDUCATIVOS ---
    st.header(t["edu_header"])
    
    tab1, tab2, tab3, tab_mom, tab4, tab5, tab6, tab7, tab8 = st.tabs(t["tab_titles"])



    with tab1:
        st.subheader(t["t1_title"])
        
        # Concepto e Idea de la Diapositiva 3
        col_concept, col_example = st.columns([1.2, 1])
        with col_concept:
            st.markdown(t["t1_concept_hdr"])
            st.markdown(t["t1_concept"])
            
            st.markdown(t["t1_eq_hdr"])
            st.latex(r"y = mx + c")
            st.markdown(t["t1_params_hdr"])
            st.markdown(t["t1_params"])
        with col_example:
            st.markdown(t["t1_idea_hdr"])
            st.markdown(t["t1_idea"])

        st.markdown(t["math_results"])
        
        col_f1, col_f2 = st.columns(2)
        with col_f1:
            st.markdown("<div class='latex-container'>", unsafe_allow_html=True)
            st.markdown(t["t1_f_slope"])
            st.latex(r"m = \frac{n\sum(XY) - (\sum X)(\sum Y)}{n\sum X^2 - (\sum X)^2} \text{ o } m = \frac{S_{xy}}{S_{xx}}")
            st.latex(rf"\Rightarrow m = {m_slr:.4f}")
            st.markdown("</div>", unsafe_allow_html=True)
            
            st.markdown("<div class='latex-container'>", unsafe_allow_html=True)
            st.markdown(t["t1_f_var"])
            st.latex(r"s_e^2 = \frac{\sum(Y_{obs} - Y_{pred})^2}{n-2}")
            st.latex(rf"\Rightarrow s_e^2 = {s2_e_slr:.4f}")
            st.markdown("</div>", unsafe_allow_html=True)
        
        with col_f2:
            st.markdown("<div class='latex-container'>", unsafe_allow_html=True)
            st.markdown(t["t1_f_intercept"])
            st.latex(r"c = \bar{Y} - m\bar{X}")
            st.latex(rf"\Rightarrow c = {c_slr:.4f}")
            st.markdown("</div>", unsafe_allow_html=True)

            st.markdown("<div class='latex-container'>", unsafe_allow_html=True)
            st.markdown(t["t1_f_se_m"])
            st.latex(r"SE_m = \sqrt{\frac{s_e^2}{\sum(X_i - \bar{X})^2}}")
            st.latex(rf"\Rightarrow SE_m = {se_m_slr:.4f}")
            st.markdown("</div>", unsafe_allow_html=True)

        st.markdown("<div class='latex-container'>", unsafe_allow_html=True)
        st.markdown(t["t1_f_se_c"])
        st.latex(r"SE_c = SE_m \sqrt{\frac{\sum X_i^2}{n}}")
        st.latex(rf"\Rightarrow SE_c = {se_c_slr:.4f}")
        st.markdown("</div>", unsafe_allow_html=True)


    with tab2:
        st.subheader(t["t2_title"])
        st.markdown(t["t2_edu_focus"])
        
        st.markdown(t["math_results"])
        
        col_f1, col_f2 = st.columns(2)
        with col_f1:
            st.markdown("<div class='latex-container'>", unsafe_allow_html=True)
            st.markdown(t["t2_f_eta"])
            st.latex(r"\eta = \frac{\sigma^2_{\varepsilon y}}{\sigma^2_{\varepsilon x}}")
            st.latex(rf"\Rightarrow \eta = {eta:.4f}")
            st.markdown("</div>", unsafe_allow_html=True)

            st.markdown("<div class='latex-container'>", unsafe_allow_html=True)
            st.markdown(t["t2_f_slope"])
            st.latex(r"\hat{\beta}_1 = \frac{(S_{yy} - \eta S_{xx}) + \sqrt{(S_{yy} - \eta S_{xx})^2 + 4 \eta S_{xy}^2}}{2 S_{xy}}")
            st.latex(rf"\Rightarrow \hat{{\beta}}_1 = {m_gor:.4f}")
            st.markdown("</div>", unsafe_allow_html=True)
            
            st.markdown("<div class='latex-container'>", unsafe_allow_html=True)
            st.markdown(t["t2_f_projections"])
            st.latex(r"X_t = \frac{\hat{\beta}_1(Y_{obs} - \hat{\beta}_0) + \eta X_{obs}}{\eta + \hat{\beta}_1^2}")
            st.latex(r"Y_t = \hat{\beta}_0 + \hat{\beta}_1 X_t")
            st.markdown("</div>", unsafe_allow_html=True)

        with col_f2:
            st.markdown("<div class='latex-container'>", unsafe_allow_html=True)
            st.markdown(t["t2_f_intercept"])
            st.latex(r"\hat{\beta}_0 = \bar{Y} - \hat{\beta}_1 \bar{X}")
            st.latex(rf"\Rightarrow \hat{{\beta}}_0 = {b_gor:.4f}")
            st.markdown("</div>", unsafe_allow_html=True)
            
            st.markdown("<div class='latex-container'>", unsafe_allow_html=True)
            st.markdown(t["t2_f_var"])
            st.latex(r"\hat{\sigma}^2 = \frac{1}{n-2} \sum_{i=1}^{n} \frac{(Y_i - \hat{\beta}_0 - \hat{\beta}_1 X_i)^2}{\hat{\beta}_1^2 + \eta}")
            st.latex(rf"\Rightarrow \hat{{\sigma}}^2 = {s2_e_gor:.4f}")
            st.markdown("</div>", unsafe_allow_html=True)
            
            st.markdown("<div class='latex-container'>", unsafe_allow_html=True)
            st.markdown(t["t2_f_se"])
            st.latex(r"\hat{\sigma} = \sqrt{\hat{\sigma}^2}")
            st.latex(rf"\Rightarrow \hat{{\sigma}} = {sigma_gor:.4f}")
            st.markdown("</div>", unsafe_allow_html=True)

    with tab3:
        st.subheader(t["t3_title"])
        st.markdown(t["t3_edu_focus"])
        
        st.markdown(t["math_results"])
        
        col_f1, col_f2 = st.columns(2)
        with col_f1:
            st.markdown("<div class='latex-container'>", unsafe_allow_html=True)
            st.markdown(t["t3_f_slope"])
            st.latex(r"c_1 = \frac{\sum (X_{obs,i} - \bar{X}_{obs})(Y_{t,i} - \bar{Y}_t)}{\sum (X_{obs,i} - \bar{X}_{obs})^2}")
            st.latex(rf"\Rightarrow c_1 = {m_prop:.4f}")
            st.markdown("</div>", unsafe_allow_html=True)
            
            st.markdown("<div class='latex-container'>", unsafe_allow_html=True)
            st.markdown(t["t3_f_eq"])
            st.latex(r"Y_{t\_propuesto} = c_1 X_{obs} + c_2")
            st.markdown(t["t3_f_eq_calc"].format(m_prop, b_prop))
            st.markdown("</div>", unsafe_allow_html=True)

        with col_f2:
            st.markdown("<div class='latex-container'>", unsafe_allow_html=True)
            st.markdown(t["t3_f_intercept"])
            st.latex(r"c_2 = \bar{Y}_t - c_1 \bar{X}_{obs}")
            st.latex(rf"\Rightarrow c_2 = {b_prop:.4f}")
            st.markdown("</div>", unsafe_allow_html=True)

            st.markdown("<div class='latex-container'>", unsafe_allow_html=True)
            st.markdown(t["t3_f_rmse"])
            st.latex(r"RMSE = \sqrt{\frac{\sum (Y_{obs} - Y_{t\_propuesto})^2}{n}}")
            st.latex(rf"\Rightarrow RMSE = {rmse_prop:.4f}")
            st.markdown("</div>", unsafe_allow_html=True)

    with tab_mom:
        st.subheader(t["tmom_title"])
        st.markdown(t["tmom_edu_focus"])
        
        st.markdown(t["math_results"])
        
        col_f1, col_f2 = st.columns(2)
        with col_f1:
            st.markdown("<div class='latex-container'>", unsafe_allow_html=True)
            st.markdown(t["tmom_f_lambda"])
            st.latex(r"\lambda = \frac{s_w^2 - \sigma_u^2}{s_w^2}")
            st.latex(rf"\Rightarrow \lambda = {lambda_ratio:.4f}")
            st.markdown("</div>", unsafe_allow_html=True)

            st.markdown("<div class='latex-container'>", unsafe_allow_html=True)
            st.markdown(t["tmom_f_slope"])
            st.latex(r"\hat{\beta}_1 = \frac{s_{wy}}{s_w^2 - \sigma_u^2}")
            st.latex(rf"\Rightarrow \hat{{\beta}}_1 = {m_mom:.4f}")
            st.markdown("</div>", unsafe_allow_html=True)

        with col_f2:
            st.markdown("<div class='latex-container'>", unsafe_allow_html=True)
            st.markdown(t["tmom_f_intercept"])
            st.latex(r"\hat{\beta}_0 = \bar{Y} - \hat{\beta}_1 \bar{W}")
            st.latex(rf"\Rightarrow \hat{{\beta}}_0 = {b_mom:.4f}")
            st.markdown("</div>", unsafe_allow_html=True)

            st.markdown("<div class='latex-container'>", unsafe_allow_html=True)
            st.markdown(t["tmom_f_se"])
            st.latex(r"SE_{\hat{\beta}_1} = \frac{SE_{\hat{\beta}_1(OLS)}}{\lambda}")
            st.latex(rf"\Rightarrow SE_{{\hat{{\beta}}_1}} = {se_m_mom:.4f}")
            st.markdown("</div>", unsafe_allow_html=True)

    with tab4:
        st.subheader(t["t4_title"])
        
        # Conceptos de las Diapositivas 4 y 5
        st.markdown(t["t4_fund_hdr"])
        col_nl_concept, col_nl_diff = st.columns([1.1, 1.2])
        
        with col_nl_concept:
            st.markdown(t["t4_what_is"])
            st.markdown(t["t4_exp_model"])
            st.latex(r"y = a + be^{cx}")
            st.markdown(t["t4_meaning"])
            st.markdown(t["t4_examples"])
            
        with col_nl_diff:
            st.markdown(t["t4_table_hdr"])
            if st.session_state.lang == 'es':
                diff_df = pd.DataFrame({
                    "Característica": ["Forma (Shape)", "Tasa de Cambio (Rate of change)", "Ecuación (Equation)", "Complejidad (Complexity)", "Ejemplo (Example)"],
                    "Regresión Lineal": ["Línea recta (Straight line)", "Constante (Constant)", "y = mx + c", "Simple", "Salario vs Experiencia"],
                    "Regresión No Lineal": ["Curva (Curve)", "Cambiante (Changing)", "y = a + be^(cx), etc.", "Más compleja (More complex)", "Sistemas de decaimiento/crecimiento"]
                })
            else:
                diff_df = pd.DataFrame({
                    "Feature": ["Shape", "Rate of change", "Equation", "Complexity", "Example"],
                    "Linear Regression": ["Straight line", "Constant", "y = mx + c", "Simple", "Salary vs Experience"],
                    "Non-Linear Regression": ["Curve", "Changing", "y = a + be^(cx), etc.", "More complex", "Decay/growth systems"]
                })
            st.table(diff_df)
            
        st.markdown("---")

        if not nl_valid:
            st.warning(nl_error_msg)
            st.markdown(t["t4_warn_note"])
        else:
            st.markdown(t["t4_edu_focus"].format(selected_nl_short))
            
            st.markdown(t["t4_process_hdr"])
            col_f1, col_f2 = st.columns(2)
            
            if selected_nl_key == "Exponential":
                cols_lin = ["x", "Y", "y-a", "ln(y-a)", "Y est", "Residual"] if st.session_state.lang == 'en' else ["x", "Y", "y-a", "ln(y-a)", "Y estima", "Residuo"]
                df_linearized = pd.DataFrame({
                    cols_lin[0]: X,
                    cols_lin[1]: Y,
                    cols_lin[2]: y_minus_a,
                    cols_lin[3]: y_trans,
                    cols_lin[4]: y_pred_nl,
                    cols_lin[5]: Y - y_pred_nl
                })
                st.markdown(t["t4_table_interm"])
                st.dataframe(df_linearized.style.format({
                    cols_lin[0]: "{:g}",
                    cols_lin[1]: "{:.2f}",
                    cols_lin[2]: "{:.2f}",
                    cols_lin[3]: "{:.6f}",
                    cols_lin[4]: "{:.6f}",
                    cols_lin[5]: "{:.6f}"
                }), use_container_width=True)
                
                # 2. Sumatorias e Parámetros
                col_math1, col_math2 = st.columns([1.1, 1])
                
                with col_math1:
                    st.markdown(t["t4_sums_hdr"])
                    st.markdown(f"""
                    *   {t['t4_n_points'].format(n)}
                    *   {t['t4_sum_x'].format(sum_x)}
                    *   {t['t4_sum_x2'].format(sum_x2)}
                    *   {t['t4_sum_ln_ya'].format(sum_y_trans)}
                    *   {t['t4_sum_x_ln_ya'].format(sum_x_y_trans)}
                    """)
                    
                    st.markdown(t["t4_formulas_params"])
                    st.markdown("<div class='latex-container'>", unsafe_allow_html=True)
                    st.latex(r"c = \frac{n \sum (x \ln(y-a)) - (\sum x)(\sum \ln(y-a))}{n \sum x^2 - (\sum x)^2}")
                    st.latex(rf"\Rightarrow c \text{{ (pendiente)}} = {c_nl:.6f}")
                    st.markdown("</div>", unsafe_allow_html=True)
                    
                    st.markdown("<div class='latex-container'>", unsafe_allow_html=True)
                    st.latex(r"A = \frac{\sum \ln(y-a) - c \sum x}{n}")
                    st.latex(rf"\Rightarrow A \text{{ (intersección)}} = {A_nl:.6f}")
                    st.markdown("</div>", unsafe_allow_html=True)
                    
                    st.markdown("<div class='latex-container'>", unsafe_allow_html=True)
                    st.latex(r"b = e^A")
                    st.latex(rf"\Rightarrow b = e^{{{A_nl:.6f}}} = {b_nl:.6f}")
                    st.markdown("</div>", unsafe_allow_html=True)
                    
                with col_math2:
                    # Gráfico de Linealización
                    fig_lin = plgo.Figure()
                    
                    # Puntos linealizados
                    fig_lin.add_trace(plgo.Scatter(
                        x=X, y=y_trans, mode='markers', name=t["t4_graph_lin_dots"],
                        marker=dict(color='#1f77b4', size=10, symbol='circle'),
                        hovertemplate='x: %{x}<br>ln(y-a): %{y:.6f}<extra></extra>'
                    ))
                    
                    # Recta linealizada
                    x_line = np.linspace(min(X), max(X), 100)
                    y_line_trans = c_nl * x_line + A_nl
                    
                    fig_lin.add_trace(plgo.Scatter(
                        x=x_line, y=y_line_trans, mode='lines', name=t["t4_graph_lin_fit"],
                        line=dict(color='red', width=2, dash='dash'),
                        hovertemplate='x: %{x:.2f}<br>ln(y-a) pred: %{y:.6f}<extra></extra>'
                    ))
                    
                    sign_A = "+" if A_nl >= 0 else "-"
                    abs_A = abs(A_nl)
                    fig_lin.update_layout(
                        title=t["t4_graph_lin_title"].format(c_nl, sign_A, abs_A, r2_linearized),
                        xaxis_title=t["t4_graph_lin_x"],
                        yaxis_title=t["t4_graph_lin_y"],
                        template='plotly_white',
                        showlegend=False,
                        height=350,
                        margin=dict(l=40, r=40, t=60, b=40)
                    )
                    st.plotly_chart(fig_lin, use_container_width=True)
                    
                st.markdown(t["t4_space_orig_hdr"])
                st.markdown(t["t4_eq_adjusted"].format(a_nl, '+' if b_nl >= 0 else '-', abs(b_nl), c_nl))
                    
            elif selected_nl_key == "Logarithmic":
                with col_f1:
                    st.markdown("<div class='latex-container'>", unsafe_allow_html=True)
                    st.markdown(t["t4_eq_orig_lin"])
                    st.latex(r"Y = a + b \ln(X)")
                    if st.session_state.lang == 'es':
                        st.markdown("Definiendo $X' = \\ln(X)$, ajustamos la recta:")
                    else:
                        st.markdown("Defining $X' = \\ln(X)$, we fit the line:")
                    st.latex(r"Y = a + b X'")
                    st.markdown("</div>", unsafe_allow_html=True)
                    
                    st.markdown("<div class='latex-container'>", unsafe_allow_html=True)
                    st.markdown(t["t4_model_coefs"])
                    st.latex(rf"\Rightarrow a = {a_nl:.4f}")
                    st.latex(rf"\Rightarrow b = {b_nl:.4f}")
                    st.markdown("</div>", unsafe_allow_html=True)
                with col_f2:
                    st.markdown("<div class='latex-container'>", unsafe_allow_html=True)
                    st.markdown(t["t1_f_slope"])
                    st.latex(r"b = \frac{\sum (\ln(X_i) - \bar{\ln(X)})(Y_i - \bar{Y})}{\sum (\ln(X_i) - \bar{\ln(X)})^2}")
                    st.latex(rf"\Rightarrow b = {b_nl:.4f}")
                    st.markdown("</div>", unsafe_allow_html=True)
                    
                    st.markdown("<div class='latex-container'>", unsafe_allow_html=True)
                    st.markdown(t["t1_f_intercept"])
                    st.latex(r"a = \bar{Y} - b \bar{\ln(X)}")
                    st.latex(rf"\Rightarrow a = {a_nl:.4f}")
                    st.markdown("</div>", unsafe_allow_html=True)
                    
            elif selected_nl_key == "Power":
                with col_f1:
                    st.markdown("<div class='latex-container'>", unsafe_allow_html=True)
                    st.markdown(t["t4_eq_orig_lin"])
                    st.latex(r"Y = a \cdot X^b \implies \ln(Y) = \ln(a) + b \ln(X)")
                    if st.session_state.lang == 'es':
                        st.markdown("Definiendo $X' = \\ln(X)$, $Y' = \\ln(Y)$ y $A = \\ln(a)$, ajustamos:")
                    else:
                        st.markdown("Defining $X' = \\ln(X)$, $Y' = \\ln(Y)$ and $A = \\ln(a)$, we fit:")
                    st.latex(r"Y' = A + b X'")
                    st.markdown("</div>", unsafe_allow_html=True)
                    
                    st.markdown("<div class='latex-container'>", unsafe_allow_html=True)
                    st.markdown(t["t4_formulas_params"])
                    st.latex(rf"b = {b_nl:.4f}")
                    st.latex(r"a = e^A")
                    st.latex(rf"\Rightarrow a = e^{{{A_nl:.4f}}} = {a_nl:.4f}")
                    st.markdown("</div>", unsafe_allow_html=True)
                with col_f2:
                    st.markdown("<div class='latex-container'>", unsafe_allow_html=True)
                    st.markdown(t["t1_f_slope"])
                    st.latex(r"b = \frac{\sum (\ln(X_i) - \bar{\ln(X)})(\ln(Y_i) - \bar{\ln(Y)})}{\sum (\ln(X_i) - \bar{\ln(X)})^2}")
                    st.latex(rf"\Rightarrow b = {b_nl:.4f}")
                    st.markdown("</div>", unsafe_allow_html=True)
                    
                    st.markdown("<div class='latex-container'>", unsafe_allow_html=True)
                    st.markdown(t["t1_f_intercept"])
                    st.latex(r"A = \bar{\ln(Y)} - b \bar{\ln(X)}")
                    st.latex(rf"\Rightarrow A = {A_nl:.4f}")
                    st.markdown("</div>", unsafe_allow_html=True)
                    
            elif selected_nl_key == "Quadratic":
                with col_f1:
                    st.markdown("<div class='latex-container'>", unsafe_allow_html=True)
                    st.markdown(t["t4_quad_eq"])
                    st.latex(r"Y = a X^2 + b X + c")
                    st.markdown("A pesar de ser curvilíneo, es **lineal en sus parámetros** ($a, b, c$)." if st.session_state.lang == 'es' else "Despite being curvilinear, it is **linear in its parameters** ($a, b, c$).")
                    st.markdown("</div>", unsafe_allow_html=True)
                    
                    st.markdown("<div class='latex-container'>", unsafe_allow_html=True)
                    st.markdown(t["t4_coefs_calc"])
                    st.latex(rf"\Rightarrow a = {a_nl:.4f}")
                    st.latex(rf"\Rightarrow b = {b_nl:.4f}")
                    st.latex(rf"\Rightarrow c = {c_nl:.4f}")
                    st.markdown("</div>", unsafe_allow_html=True)
                with col_f2:
                    st.markdown("<div class='latex-container'>", unsafe_allow_html=True)
                    st.markdown(t["t4_normal_eq_sys"])
                    st.latex(r"\begin{pmatrix} \sum X_i^4 & \sum X_i^3 & \sum X_i^2 \\ \sum X_i^3 & \sum X_i^2 & \sum X_i \\ \sum X_i^2 & \sum X_i & n \end{pmatrix} \begin{pmatrix} a \\ b \\ c \end{pmatrix} = \begin{pmatrix} \sum X_i^2 Y_i \\ \sum X_i Y_i \\ \sum Y_i \end{pmatrix}")
                    st.markdown("Se resuelve directamente por álgebra matricial lineal." if st.session_state.lang == 'es' else "It is solved directly by linear matrix algebra.")
                    st.markdown("</div>", unsafe_allow_html=True)

            st.success(t["t4_eq_adjusted_generic"].format(nl_equation))
            
            # Comparación específica de métricas
            st.markdown(t["t4_metrics_comp"])
            metrics_nl_data = {
                t["t4_metrics_cols"][0]: t["t4_metrics_rows"],
                t["t4_metrics_cols"][1]: [sse_slr, rmse_slr, r2_slr],
                t["t4_metrics_cols"][2].format(selected_nl_short): [sse_nl, rmse_nl, r2_nl]
            }
            df_metrics_nl = pd.DataFrame(metrics_nl_data)
            st.dataframe(df_metrics_nl.style.format({
                t["t4_metrics_cols"][1]: "{:.4f}",
                t["t4_metrics_cols"][2].format(selected_nl_short): "{:.4f}"
            }), use_container_width=True)
            
            # Calculadora de predicción interactiva
            st.markdown("---")
            st.markdown(t["t4_calc_hdr"])
            x_input = st.number_input(t["t4_calc_input"].format(selected_nl_short), value=float(np.mean(X)), format="%.4f")
            
            y_pred_calc = None
            if selected_nl_key == "Exponential":
                y_pred_calc = a_nl + b_nl * np.exp(c_nl * x_input)
                sign_b = "+" if b_nl >= 0 else "-"
                abs_b = abs(b_nl)
                calc_latex = rf"y = {a_nl:.4f} {sign_b} {abs_b:.4f}e^{{{c_nl:.4f} \cdot {x_input:.4f}}} = {y_pred_calc:.4f}"
            elif selected_nl_key == "Logarithmic":
                if x_input <= 0:
                    st.error(t["t4_calc_err_log"])
                else:
                    y_pred_calc = a_nl + b_nl * np.log(x_input)
                    calc_latex = rf"Y = {a_nl:.4f} + {b_nl:.4f} \cdot \ln({x_input:.4f}) = {y_pred_calc:.4f}"
            elif selected_nl_key == "Power":
                if x_input <= 0:
                    st.error(t["t4_calc_err_pot"])
                else:
                    y_pred_calc = a_nl * (x_input ** b_nl)
                    calc_latex = rf"Y = {a_nl:.4f} \cdot {x_input:.4f}^{{{b_nl:.4f}}} = {y_pred_calc:.4f}"
            elif selected_nl_key == "Quadratic":
                y_pred_calc = a_nl * (x_input**2) + b_nl * x_input + c_nl
                calc_latex = rf"Y = {a_nl:.4f} \cdot ({x_input:.4f})^2 + {b_nl:.4f} \cdot ({x_input:.4f}) + ({c_nl:.4f}) = {y_pred_calc:.4f}"
                
            if y_pred_calc is not None:
                st.latex(calc_latex)
                st.markdown(t["t4_calc_res_desc"].format(x_input, y_pred_calc))

    with tab5:
        st.subheader(t["t5_title"])
        if not power_valid:
            st.warning(power_error_msg)
            st.markdown(t["t5_warn_note"])
        else:
            st.markdown(t["t5_edu_focus"])
            
            st.markdown(t["t4_process_hdr"])
            
            # 1. Tabla de datos
            df_power_linearized = pd.DataFrame({
                "x": X,
                "Y": Y,
                "y-a": [np.nan] * len(X),
                "ln(y-a)": [np.nan] * len(X),
                "x*": x_trans_power,
                "y*": y_trans_power,
                "x* * y*": x_trans_power * y_trans_power,
                "(x*)^2": x_trans_power**2,
                "Y est Power": y_pred_power
            })
            st.markdown(t["t4_table_interm"])
            st.dataframe(df_power_linearized.style.format({
                "x": "{:g}",
                "Y": "{:.2f}",
                "y-a": "{:.2f}",
                "ln(y-a)": "{:.6f}",
                "x*": "{:.4f}",
                "y*": "{:.4f}",
                "x* * y*": "{:.4f}",
                "(x*)^2": "{:.4f}",
                "Y est Power": "{:.4f}"
            }, na_rep="-"), use_container_width=True)
            
            # 2. Sumatorias e Parámetros
            col_math1, col_math2 = st.columns([1.1, 1])
            
            with col_math1:
                st.markdown(t["t4_sums_hdr"])
                st.markdown(f"""
                *   {t['t4_n_points'].format(n)}
                *   {t['t5_sum_x_trans'].format(sum_x_trans_power)}
                *   {t['t5_sum_x2_trans'].format(sum_x2_trans_power)}
                *   {t['t5_sum_y_trans'].format(sum_y_trans_power)}
                *   {t['t5_sum_xy_trans'].format(sum_xy_trans_power)}
                """)
                
                st.markdown(t["t4_formulas_params"])
                st.markdown("<div class='latex-container'>", unsafe_allow_html=True)
                st.latex(r"b = \frac{n \sum (x^* y^*) - (\sum x^*)(\sum y^*)}{n \sum (x^*)^2 - (\sum x^*)^2}")
                st.latex(rf"\Rightarrow b \text{{ (exponente)}} = {b_power:.6f}")
                st.markdown("</div>", unsafe_allow_html=True)
                
                st.markdown("<div class='latex-container'>", unsafe_allow_html=True)
                st.latex(r"A = \ln(a) = \frac{\sum y^* - b \sum x^*}{n}")
                st.latex(rf"\Rightarrow A = \ln(a) = {A_power:.6f}")
                st.markdown("</div>", unsafe_allow_html=True)
                
                st.markdown("<div class='latex-container'>", unsafe_allow_html=True)
                st.latex(r"a = e^A")
                st.latex(rf"\Rightarrow a = e^{{{A_power:.6f}}} = {a_power:.6f}")
                st.markdown("</div>", unsafe_allow_html=True)
                
            with col_math2:
                # Gráfico Y est Power
                fig_power = plgo.Figure()
                
                # Puntos reales
                fig_power.add_trace(plgo.Scatter(
                    x=X, y=Y, mode='markers', name=t["chart_obs"],
                    marker=dict(color='black', size=10, symbol='circle'),
                    hovertemplate='x: %{x}<br>Y: %{y:.2f}<extra></extra>'
                ))
                
                # Curva de potencia
                x_curve = np.linspace(min(X), max(X), 100)
                y_curve_power = a_power * (x_curve ** b_power)
                
                fig_power.add_trace(plgo.Scatter(
                    x=x_curve, y=y_curve_power, mode='lines', name=t["t4_graph_lin_fit"],
                    line=dict(color='#1f77b4', width=2),
                    hovertemplate='x: %{x:.2f}<br>Y estima: %{y:.4f}<extra></extra>'
                ))
                
                fig_power.update_layout(
                    title=t["t5_graph_title"].format(a_power, b_power, r2_power),
                    xaxis_title="x",
                    yaxis_title="Y",
                    template='plotly_white',
                    showlegend=False,
                    height=350,
                    margin=dict(l=40, r=40, t=60, b=40)
                )
                st.plotly_chart(fig_power, use_container_width=True)
                
            st.markdown(t["t5_res_hdr"])
            st.success(t["t5_eq_adjusted"].format(a_power, b_power))
            
            # Comparativa de métricas
            st.markdown(t["t5_metrics_comp"])
            metrics_power_data = {
                t["t5_metrics_cols"][0]: t["t4_metrics_rows"],
                t["t5_metrics_cols"][1]: [sse_slr, rmse_slr, r2_slr],
                t["t5_metrics_cols"][2]: [sse_power, rmse_power, r2_power]
            }
            df_metrics_power = pd.DataFrame(metrics_power_data)
            st.dataframe(df_metrics_power.style.format({
                t["t5_metrics_cols"][1]: "{:.4f}",
                t["t5_metrics_cols"][2]: "{:.4f}"
            }), use_container_width=True)
            
            # Calculadora de predicción interactiva
            st.markdown("---")
            st.markdown(t["t4_calc_hdr"])
            x_input = st.number_input(t["t5_calc_input"], value=float(np.mean(X)), format="%.4f")
            
            if x_input <= 0:
                st.error(t["t5_calc_err"])
            else:
                y_pred_calc = a_power * (x_input ** b_power)
                calc_latex = rf"y = {a_power:.4f} \cdot {x_input:.4f}^{{{b_power:.4f}}} = {y_pred_calc:.4f}"
                st.latex(calc_latex)
                st.markdown(t["t5_calc_res"].format(x_input, y_pred_calc))

    with tab6:
        st.subheader(t["t6_title"])
        st.markdown(t["t6_desc"])
        
        X_mod = df["X"].values
        Y_mod = df["Y"].values

        if not fun_pot_valid:
            st.error(fun_pot_error_msg)
            st.info(t["t6_warn_note"])
        else:
            r2_lin = r2_lin_pot10
            df_mod_calc = pd.DataFrame({
                "x": X,
                "y": Y,
                "Log Y": Y_log,
                "Log X": X_log
            })

            st.markdown(t["t6_table_hdr"])
            st.dataframe(df_mod_calc.style.format({
                "x": "{:g}",
                "y": "{:g}",
                "Log Y": "{:.5f}",
                "Log X": "{:.5f}"
            }), use_container_width=True)

            col_coef1, col_coef2 = st.columns(2)
            with col_coef1:
                st.markdown("<div class='latex-container'>", unsafe_allow_html=True)
                st.markdown(t["t6_coefs_lin"])
                st.latex(rf"\log_{{10}}(y) = {m_log:.4f} \cdot \log_{{10}}(x) + {c_log:.4f}")
                st.latex(rf"R^2 = {r2_lin:.4f}")
                st.markdown("</div>", unsafe_allow_html=True)

            with col_coef2:
                st.markdown("<div class='latex-container'>", unsafe_allow_html=True)
                st.markdown(t["t6_coefs_adj"])
                st.markdown(t["t6_coefs_log_a"].format(c_log))
                st.markdown(t["t6_coefs_log_b"].format(m_log))
                st.markdown(t["t6_coefs_a"].format(a_correct))
                st.markdown(t["t6_coefs_b_corr"].format(b_correct))
                st.markdown(t["t6_coefs_b_inc"].format(b_err))
                st.markdown("</div>", unsafe_allow_html=True)

            col_g1, col_g2 = st.columns(2)
            with col_g1:
                fig_lin_power = plgo.Figure()
                fig_lin_power.add_trace(plgo.Scatter(
                    x=X_log, y=Y_log, mode='markers', name=t["t4_graph_lin_dots"],
                    marker=dict(color='#1f77b4', size=10, symbol='circle'),
                    hovertemplate='Log X: %{x:.5f}<br>Log Y: %{y:.5f}<extra></extra>'
                ))
                x_log_line = np.linspace(min(X_log), max(X_log), 100)
                y_log_line = m_log * x_log_line + c_log
                fig_lin_power.add_trace(plgo.Scatter(
                    x=x_log_line, y=y_log_line, mode='lines', name=t["t4_graph_lin_fit"],
                    line=dict(color='#ff7f0e', width=2, dash='dash'),
                    hovertemplate='Log X: %{x:.5f}<br>Log Y Pred: %{y:.5f}<extra></extra>'
                ))
                fig_lin_power.update_layout(
                    title=t["t6_graph_lin_title"].format(m_log, c_log, r2_lin),
                    xaxis_title=t["t6_graph_lin_x"],
                    yaxis_title=t["t6_graph_lin_y"],
                    template='plotly_white',
                    showlegend=False,
                    height=380,
                    margin=dict(l=40, r=40, t=60, b=40)
                )
                st.plotly_chart(fig_lin_power, use_container_width=True)

            with col_g2:
                fig_orig_power = plgo.Figure()
                fig_orig_power.add_trace(plgo.Scatter(
                    x=X_mod, y=Y_mod, mode='markers', name=t["chart_obs"],
                    marker=dict(color='black', size=10, symbol='circle'),
                    hovertemplate='X: %{x}<br>Y: %{y}<extra></extra>'
                ))

                x_orig_line = np.linspace(min(X_mod), max(X_mod), 200)
                y_orig_correct = a_correct * (x_orig_line ** b_correct)
                
                fig_orig_power.add_trace(plgo.Scatter(
                    x=x_orig_line, y=y_orig_correct, mode='lines', name=t["t6_graph_orig_correct"],
                    line=dict(color='green', width=2.5),
                    hovertemplate='X: %{x:.2f}<br>Y (Correcto): %{y:.4f}<extra></extra>'
                ))

                try:
                    y_orig_err = a_err * (x_orig_line ** b_err)
                    if not np.any(np.isinf(y_orig_err)) and np.max(y_orig_err) < 1000000:
                        fig_orig_power.add_trace(plgo.Scatter(
                            x=x_orig_line, y=y_orig_err, mode='lines', name=t["t6_graph_orig_incorrect"],
                            line=dict(color='red', width=2.5, dash='dot'),
                            hovertemplate='X: %{x:.2f}<br>Y (Despeje Erróneo): %{y:.4f}<extra></extra>'
                        ))
                except Exception:
                    pass

                fig_orig_power.update_layout(
                    title=t["t6_graph_orig_title"],
                    xaxis_title="X",
                    yaxis_title="Y",
                    template='plotly_white',
                    legend=dict(yanchor="top", y=0.99, xanchor="left", x=0.01),
                    height=380,
                    margin=dict(l=40, r=40, t=60, b=40)
                )
                st.plotly_chart(fig_orig_power, use_container_width=True)

            st.warning(t["t6_analysis_hdr"])
            st.markdown(t["t6_analysis_body"].format(c_log, c_log, a_correct, m_log, a_correct, b_correct, m_log, b_err, a_err, b_err))

            st.markdown(t["t6_math_formulation"])
            st.markdown(t["t6_math_steps"])

    with tab7:
        st.subheader(t["t7_title"])
        st.markdown(t["t7_desc"])
        
        X_mod = df["X"].values
        Y_mod = df["Y"].values

        if not fun_exp_valid:
            st.error(fun_exp_error_msg)
            st.info(t["t6_warn_note"])
        else:
            r2_lin = r2_lin_exp
            df_mod_calc = pd.DataFrame({
                "x": X,
                "y": Y,
                "Ln Y": Y_log_exp
            })

            st.markdown(t["t7_table_hdr"])
            st.dataframe(df_mod_calc.style.format({
                "x": "{:g}",
                "y": "{:g}",
                "Ln Y": "{:.8f}"
            }), use_container_width=True)

            col_coef1, col_coef2 = st.columns(2)
            with col_coef1:
                st.markdown("<div class='latex-container'>", unsafe_allow_html=True)
                st.markdown(t["t7_coefs_lin"])
                st.latex(rf"\ln(y) = {b_coeff:.4f} \cdot x + {ln_a:.4f}")
                st.latex(rf"R^2 = {r2_lin:.4f}")
                st.markdown("</div>", unsafe_allow_html=True)

            with col_coef2:
                st.markdown("<div class='latex-container'>", unsafe_allow_html=True)
                st.markdown(t["t7_coefs_adj"])
                st.markdown(t["t7_coefs_ln_a"].format(ln_a))
                st.markdown(t["t7_coefs_b"].format(b_coeff))
                st.markdown(t["t7_coefs_a"].format(a_coeff))
                st.markdown(t["t7_eq_final"].format(a_coeff, b_coeff))
                st.markdown("</div>", unsafe_allow_html=True)

            col_g1, col_g2 = st.columns(2)
            with col_g1:
                fig_lin_exp = plgo.Figure()
                fig_lin_exp.add_trace(plgo.Scatter(
                    x=X_mod, y=Y_log, mode='markers', name=t["t4_graph_lin_dots"],
                    marker=dict(color='#1f77b4', size=10, symbol='circle'),
                    hovertemplate='X: %{x}<br>ln(Y): %{y:.5f}<extra></extra>'
                ))
                x_line = np.linspace(min(X_mod), max(X_mod), 100)
                y_line = b_coeff * x_line + ln_a
                fig_lin_exp.add_trace(plgo.Scatter(
                    x=x_line, y=y_line, mode='lines', name=t["t4_graph_lin_fit"],
                    line=dict(color='#ff7f0e', width=2),
                    hovertemplate='X: %{x:.2f}<br>ln(Y) Pred: %{y:.5f}<extra></extra>'
                ))
                fig_lin_exp.update_layout(
                    title=t["t7_graph_lin_title"].format(b_coeff, ln_a, r2_lin),
                    xaxis_title="x",
                    yaxis_title=t["t7_graph_lin_y"],
                    template='plotly_white',
                    showlegend=False,
                    height=380,
                    margin=dict(l=40, r=40, t=60, b=40)
                )
                st.plotly_chart(fig_lin_exp, use_container_width=True)

            with col_g2:
                fig_orig_exp = plgo.Figure()
                fig_orig_exp.add_trace(plgo.Scatter(
                    x=X_mod, y=Y_mod, mode='markers', name=t["chart_obs"],
                    marker=dict(color='black', size=10, symbol='circle'),
                    hovertemplate='X: %{x}<br>Y: %{y}<extra></extra>'
                ))

                x_orig_line = np.linspace(min(X_mod), max(X_mod), 200)
                y_orig_fit = a_coeff * np.exp(b_coeff * x_orig_line)

                fig_orig_exp.add_trace(plgo.Scatter(
                    x=x_orig_line, y=y_orig_fit, mode='lines', name=t["t7_graph_orig_curve"],
                    line=dict(color='green', width=2.5),
                    hovertemplate='X: %{x:.2f}<br>Y estima: %{y:.4f}<extra></extra>'
                ))

                fig_orig_exp.update_layout(
                    title=t["t7_graph_orig_title"],
                    xaxis_title="X",
                    yaxis_title="Y",
                    template='plotly_white',
                    showlegend=False,
                    height=380,
                    margin=dict(l=40, r=40, t=60, b=40)
                )
                st.plotly_chart(fig_orig_exp, use_container_width=True)

            st.info(t["t7_analysis_hdr"])
            st.markdown(t["t7_analysis_body"].format(ln_a, ln_a, a_coeff, b_coeff, a_coeff, b_coeff))

            st.markdown(t["t7_math_formulation"])
            st.markdown(t["t7_math_steps"])

    with tab8:
        st.subheader(t["t8_title_data"])
        st.markdown(t["t8_desc_data"])
        # Translate df_results columns for presentation
        rename_cols = {
            "X": "X",
            "Y": "Y",
            "Y_est (SLR)": t["xls_details_y_est"].format(t["methods_names"]["SLR"]),
            "Residuo (SLR)": t["xls_details_residual"].format(t["methods_names"]["SLR"]),
            "Y_est (GOR Conv)": t["xls_details_y_est"].format(t["methods_names"]["GOR Conv"]),
            "Residuo (GOR Conv)": t["xls_details_residual"].format(t["methods_names"]["GOR Conv"]),
            "X_t (GOR)": "X_t (GOR)",
            "Y_t (GOR)": "Y_t (GOR)",
            "Y_est (GOR Prop)": t["xls_details_y_est"].format(t["methods_names"]["GOR Prop"]),
            "Residuo (GOR Prop)": t["xls_details_residual"].format(t["methods_names"]["GOR Prop"]),
            "Y_est (MoM)": t["xls_details_y_est"].format(t["methods_names"]["MoM"]),
            "Residuo (MoM)": t["xls_details_residual"].format(t["methods_names"]["MoM"]),
        }
        if exp_asymp_valid:
            rename_cols["Y_est (No Lin. Exponencial Asintota)"] = t["xls_details_y_est"].format(t["methods_names"]["No Lin. Exponencial Asintota"])
            rename_cols["Residuo (No Lin. Exponencial Asintota)"] = t["xls_details_residual"].format(t["methods_names"]["No Lin. Exponencial Asintota"])
        if log_valid:
            rename_cols["Y_est (No Lin. Logaritmico)"] = t["xls_details_y_est"].format(t["methods_names"]["No Lin. Logarítmico"])
            rename_cols["Residuo (No Lin. Logaritmico)"] = t["xls_details_residual"].format(t["methods_names"]["No Lin. Logarítmico"])
        if pot_valid:
            rename_cols["Y_est (No Lin. Potencial / Power Law)"] = t["xls_details_y_est"].format(t["methods_names"]["No Lin. Potencial / Power Law"])
            rename_cols["Residuo (No Lin. Potencial / Power Law)"] = t["xls_details_residual"].format(t["methods_names"]["No Lin. Potencial / Power Law"])
        if quad_valid:
            rename_cols["Y_est (No Lin. Cuadratico)"] = t["xls_details_y_est"].format(t["methods_names"]["No Lin. Cuadrático"])
            rename_cols["Residuo (No Lin. Cuadratico)"] = t["xls_details_residual"].format(t["methods_names"]["No Lin. Cuadrático"])
        if fun_pot_valid:
            rename_cols["Y_est (Fun. Potencia log10 - Correcto)"] = t["xls_details_y_est"].format(t["methods_names"]["Fun. Potencia log10 (Correcto)"])
            rename_cols["Residuo (Fun. Potencia log10 - Correcto)"] = t["xls_details_residual"].format(t["methods_names"]["Fun. Potencia log10 (Correcto)"])
            rename_cols["Y_est (Fun. Potencia log10 - Incorrecto)"] = t["xls_details_y_est"].format(t["methods_names"]["Fun. Potencia log10 (Incorrecto)"])
            rename_cols["Residuo (Fun. Potencia log10 - Incorrecto)"] = t["xls_details_residual"].format(t["methods_names"]["Fun. Potencia log10 (Incorrecto)"])
        if fun_exp_valid:
            rename_cols["Y_est (Fun. Exponencial Simple ln)"] = t["xls_details_y_est"].format(t["methods_names"]["Fun. Exponencial Simple ln"])
            rename_cols["Residuo (Fun. Exponencial Simple ln)"] = t["xls_details_residual"].format(t["methods_names"]["Fun. Exponencial Simple ln"])

        df_results_translated = df_results.rename(columns=rename_cols)
        st.dataframe(df_results_translated, use_container_width=True)

        st.markdown("---")
        st.subheader(t["t8_guide_title"])
        st.markdown(t["t8_guide_desc"])
        
        # Build guide table matching models_details
        models_details_guide = []
        models_details_guide.append({
            "name": "SLR",
            "pred": lambda r: f"= {m_slr:.6f} * B{r} + {c_slr:.6f}",
        })
        models_details_guide.append({
            "name": "GOR Conv",
            "pred": lambda r: f"= {m_gor:.6f} * B{r} + {b_gor:.6f}",
        })
        models_details_guide.append({
            "name": "GOR Prop",
            "pred": lambda r: f"= {m_prop:.6f} * B{r} + {b_prop:.6f}",
        })
        models_details_guide.append({
            "name": "MoM",
            "pred": lambda r: f"= {m_mom:.6f} * B{r} + {b_mom:.6f}",
        })
        if exp_asymp_valid:
            models_details_guide.append({
                "name": "No Lin. Exponencial Asintota",
                "pred": lambda r: f"= {a_val:.6f} + {b_nl_exp:.6f} * EXP({c_nl_exp:.6f} * B{r})",
            })
        if log_valid:
            models_details_guide.append({
                "name": "No Lin. Logarítmico",
                "pred": lambda r: f"= {a_nl_log:.6f} + {b_nl_log:.6f} * LN(B{r})",
            })
        if pot_valid:
            models_details_guide.append({
                "name": "No Lin. Potencial / Power Law",
                "pred": lambda r: f"= {a_nl_pot:.6f} * (B{r} ^ {b_nl_pot:.6f})",
            })
        if quad_valid:
            models_details_guide.append({
                "name": "No Lin. Cuadrático",
                "pred": lambda r: f"= {a_nl_quad:.6f} * (B{r} ^ 2) + {b_nl_quad:.6f} * B{r} + {c_nl_quad:.6f}",
            })
        if fun_pot_valid:
            models_details_guide.append({
                "name": "Fun. Potencia log10 (Correcto)",
                "pred": lambda r: f"= {a_correct:.6f} * (B{r} ^ {b_correct:.6f})",
            })
            models_details_guide.append({
                "name": "Fun. Potencia log10 (Incorrecto)",
                "pred": lambda r: f"= {a_err:.6f} * (B{r} ^ {b_err:.6f})",
            })
        if fun_exp_valid:
            models_details_guide.append({
                "name": "Fun. Exponencial Simple ln",
                "pred": lambda r: f"= {a_coeff:.6f} * EXP({b_coeff:.6f} * B{r})",
            })

        from openpyxl.utils import get_column_letter

        col_letters = {}
        c_idx = 4 # column 4 is D
        for m in models_details_guide:
            pred_letter = get_column_letter(c_idx)
            c_idx += 1
            res_letter = get_column_letter(c_idx)
            c_idx += 1
            col_letters[m["name"]] = (pred_letter, res_letter)

        cols = t["t8_guide_table_cols"]
        col_desc = t["t8_guide_table_col_desc"]
        
        guide_data = {
            cols[0]: [],
            cols[1]: [],
            cols[2]: []
        }

        for m in models_details_guide:
            pred_col, res_col = col_letters[m["name"]]
            translated_name = t["methods_names"].get(m["name"], m["name"])
            guide_data[cols[0]].append(translated_name)
            guide_data[cols[1]].append(f"`{m['pred'](6)}` ({col_desc} {pred_col})")
            guide_data[cols[2]].append(f"`= C6 - {pred_col}6` ({col_desc} {res_col})")

        df_guide = pd.DataFrame(guide_data)
        st.table(df_guide)

        st.markdown("---")
        st.subheader(t["t8_export_title"])
        col_out1, col_out2 = st.columns(2)
        
        with col_out1:
            st.markdown(f"**{t['t8_export_excel_lbl']}**")
            
            def generate_excel_with_formulas():
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                from openpyxl.utils import get_column_letter

                wb = openpyxl.Workbook()
                default_sheet = wb.active
                wb.remove(default_sheet)

                # Styles
                title_font = Font(name="Calibri", size=16, bold=True, color="1F497D")
                subtitle_font = Font(name="Calibri", size=11, italic=True, color="595959")
                header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
                bold_font = Font(name="Calibri", size=11, bold=True)
                regular_font = Font(name="Calibri", size=11)
                
                header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
                thin_border_side = Side(border_style="thin", color="D3D3D3")
                thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
                
                align_center = Alignment(horizontal="center", vertical="center")
                align_left = Alignment(horizontal="left", vertical="center")
                align_right = Alignment(horizontal="right", vertical="center")

                # --- SHEET 1: Resumen de Modelos ---
                ws_summary = wb.create_sheet(title=t["xls_sheet_summary"])
                ws_summary.views.sheetView[0].showGridLines = True
                
                ws_summary.cell(row=2, column=2, value=t["xls_summary_title"]).font = title_font
                ws_summary.cell(row=3, column=2, value=t["xls_summary_subtitle"]).font = subtitle_font
                
                headers_summary = t["xls_summary_headers"]
                for col_idx, header in enumerate(headers_summary, start=2):
                    cell = ws_summary.cell(row=5, column=col_idx, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = align_center
                    cell.border = thin_border
                
                methods_list = [
                    {"name": t["methods_names"]["SLR"], "math": "y = m*x + c", "formula": "= m * X + c", "valid": True, "rmse": rmse_slr, "r2": r2_slr},
                    {"name": t["methods_names"]["GOR Conv"], "math": "y = m*x + c", "formula": "= m * X + c", "valid": True, "rmse": rmse_gor, "r2": r2_gor},
                    {"name": t["methods_names"]["GOR Prop"], "math": "y = m*x + c", "formula": "= m * X + c", "valid": True, "rmse": rmse_prop, "r2": r2_prop},
                    {"name": t["methods_names"]["MoM"], "math": "y = m*x + c", "formula": "= m * X + c", "valid": True, "rmse": rmse_mom, "r2": r2_mom},
                    {"name": t["methods_names"]["No Lin. Exponencial Asintota"], "math": "y = a + b*e^(cx)", "formula": "= a + b * EXP(c * X)", "valid": exp_asymp_valid, "rmse": rmse_nl_exp if exp_asymp_valid else np.nan, "r2": r2_nl_exp if exp_asymp_valid else np.nan},
                    {"name": t["methods_names"]["No Lin. Logarítmico"], "math": "y = a + b*ln(x)", "formula": "= a + b * LN(X)", "valid": log_valid, "rmse": rmse_nl_log if log_valid else np.nan, "r2": r2_nl_log if log_valid else np.nan},
                    {"name": t["methods_names"]["No Lin. Potencial / Power Law"], "math": "y = a*x^b", "formula": "= a * (X ^ b)", "valid": pot_valid, "rmse": rmse_nl_pot if pot_valid else np.nan, "r2": r2_nl_pot if pot_valid else np.nan},
                    {"name": t["methods_names"]["No Lin. Cuadrático"], "math": "y = a*x^2 + b*x + c", "formula": "= a * (X ^ 2) + b * X + c", "valid": quad_valid, "rmse": rmse_nl_quad if quad_valid else np.nan, "r2": r2_nl_quad if quad_valid else np.nan},
                    {"name": t["methods_names"]["Fun. Potencia log10 (Correcto)"], "math": "y = a*x^b", "formula": "= a * (X ^ b)", "valid": fun_pot_valid, "rmse": rmse_pot10_correct if fun_pot_valid else np.nan, "r2": r2_pot10_correct if fun_pot_valid else np.nan},
                    {"name": t["methods_names"]["Fun. Potencia log10 (Incorrecto)"], "math": "y = a_err * x^b_err", "formula": "= a_err * (X ^ b_err)", "valid": fun_pot_valid, "rmse": rmse_pot10_incorrect if fun_pot_valid else np.nan, "r2": r2_pot10_incorrect if fun_pot_valid else np.nan},
                    {"name": t["methods_names"]["Fun. Exponencial Simple ln"], "math": "y = a*e^(bx)", "formula": "= a * EXP(b * X)", "valid": fun_exp_valid, "rmse": rmse_exp_simple if fun_exp_valid else np.nan, "r2": r2_exp_simple if fun_exp_valid else np.nan},
                ]
                
                current_row = 6
                for m in methods_list:
                    ws_summary.cell(row=current_row, column=2, value=m["name"]).font = regular_font
                    ws_summary.cell(row=current_row, column=3, value=m["math"]).font = regular_font
                    ws_summary.cell(row=current_row, column=4, value=m["formula"]).font = regular_font
                    
                    if m["valid"]:
                        ws_summary.cell(row=current_row, column=5, value=m["rmse"]).number_format = "0.0000"
                        ws_summary.cell(row=current_row, column=6, value=m["r2"]).number_format = "0.0000"
                        ws_summary.cell(row=current_row, column=7, value=t["xls_summary_status_fit"]).font = regular_font
                    else:
                        ws_summary.cell(row=current_row, column=5, value="-").alignment = align_center
                        ws_summary.cell(row=current_row, column=6, value="-").alignment = align_center
                        ws_summary.cell(row=current_row, column=7, value=t["xls_summary_status_invalid"]).font = Font(name="Calibri", size=11, color="FF0000")
                        
                    for c in range(2, 8):
                        cell = ws_summary.cell(row=current_row, column=c)
                        cell.border = thin_border
                        if c in [5, 6]:
                            cell.alignment = align_right
                        elif c in [2, 3, 4]:
                            cell.alignment = align_left
                        else:
                            cell.alignment = align_center
                    current_row += 1
                    
                # Autosize columns
                for col in ws_summary.columns:
                    max_len = max(len(str(cell.value or '')) for cell in col)
                    col_letter = get_column_letter(col[0].column)
                    if col[0].column >= 2:
                        ws_summary.column_dimensions[col_letter].width = max(max_len + 3, 10)
                
                # --- SHEET 2: Resultados Detallados ---
                ws_details = wb.create_sheet(title=t["xls_sheet_details"])
                ws_details.views.sheetView[0].showGridLines = True
                
                ws_details.cell(row=2, column=2, value=t["xls_details_title"]).font = title_font
                ws_details.cell(row=3, column=2, value=t["xls_details_subtitle"]).font = subtitle_font
                
                models_details = []
                models_details.append({
                    "name": "SLR",
                    "pred": lambda r: f"= {m_slr:.12f} * B{r} + {c_slr:.12f}",
                })
                models_details.append({
                    "name": "GOR Conv",
                    "pred": lambda r: f"= {m_gor:.12f} * B{r} + {b_gor:.12f}",
                })
                models_details.append({
                    "name": "GOR Prop",
                    "pred": lambda r: f"= {m_prop:.12f} * B{r} + {b_prop:.12f}",
                })
                models_details.append({
                    "name": "MoM",
                    "pred": lambda r: f"= {m_mom:.12f} * B{r} + {b_mom:.12f}",
                })
                if exp_asymp_valid:
                    models_details.append({
                        "name": "No Lin. Exponencial Asintota",
                        "pred": lambda r: f"= {a_val:.12f} + {b_nl_exp:.12f} * EXP({c_nl_exp:.12f} * B{r})",
                    })
                if log_valid:
                    models_details.append({
                        "name": "No Lin. Logarítmico",
                        "pred": lambda r: f"= {a_nl_log:.12f} + {b_nl_log:.12f} * LN(B{r})",
                    })
                if pot_valid:
                    models_details.append({
                        "name": "No Lin. Potencial / Power Law",
                        "pred": lambda r: f"= {a_nl_pot:.12f} * (B{r} ^ {b_nl_pot:.12f})",
                    })
                if quad_valid:
                    models_details.append({
                        "name": "No Lin. Cuadrático",
                        "pred": lambda r: f"= {a_nl_quad:.12f} * (B{r} ^ 2) + {b_nl_quad:.12f} * B{r} + {c_nl_quad:.12f}",
                    })
                if fun_pot_valid:
                    models_details.append({
                        "name": "Fun. Potencia log10 (Correcto)",
                        "pred": lambda r: f"= {a_correct:.12f} * (B{r} ^ {b_correct:.12f})",
                    })
                    models_details.append({
                        "name": "Fun. Potencia log10 (Incorrecto)",
                        "pred": lambda r: f"= {a_err:.12f} * (B{r} ^ {b_err:.12f})",
                    })
                if fun_exp_valid:
                    models_details.append({
                        "name": "Fun. Exponencial Simple ln",
                        "pred": lambda r: f"= {a_coeff:.12f} * EXP({b_coeff:.12f} * B{r})",
                    })
                
                # Write Headers
                ws_details.cell(row=5, column=2, value=t["xls_details_col_x"]).font = header_font
                ws_details.cell(row=5, column=2).fill = header_fill
                ws_details.cell(row=5, column=2).border = thin_border
                ws_details.cell(row=5, column=2).alignment = align_center
                
                ws_details.cell(row=5, column=3, value=t["xls_details_col_y"]).font = header_font
                ws_details.cell(row=5, column=3).fill = header_fill
                ws_details.cell(row=5, column=3).border = thin_border
                ws_details.cell(row=5, column=3).alignment = align_center
                
                col_idx = 4
                for m in models_details:
                    cell_est = ws_details.cell(row=5, column=col_idx, value=t["xls_details_y_est"].format(t["methods_names"][m["name"]]))
                    cell_est.font = header_font
                    cell_est.fill = header_fill
                    cell_est.border = thin_border
                    cell_est.alignment = align_center
                    col_idx += 1
                    
                    cell_res = ws_details.cell(row=5, column=col_idx, value=t["xls_details_residual"].format(t["methods_names"][m["name"]]))
                    cell_res.font = header_font
                    cell_res.fill = header_fill
                    cell_res.border = thin_border
                    cell_res.alignment = align_center
                    col_idx += 1

                # Write Data
                start_row = 6
                for i, (x_val, y_val) in enumerate(zip(X, Y)):
                    r = start_row + i
                    ws_details.cell(row=r, column=2, value=float(x_val)).number_format = "0.00"
                    ws_details.cell(row=r, column=3, value=float(y_val)).number_format = "0.00"
                    ws_details.cell(row=r, column=2).border = thin_border
                    ws_details.cell(row=r, column=3).border = thin_border
                    ws_details.cell(row=r, column=2).alignment = align_right
                    ws_details.cell(row=r, column=3).alignment = align_right
                    
                    c_idx = 4
                    for m in models_details:
                        pred_formula = m["pred"](r)
                        pred_letter = get_column_letter(c_idx)
                        
                        cell_pred = ws_details.cell(row=r, column=c_idx, value=pred_formula)
                        cell_pred.number_format = "0.0000"
                        cell_pred.border = thin_border
                        cell_pred.alignment = align_right
                        c_idx += 1
                        
                        res_formula = f"= C{r} - {pred_letter}{r}"
                        cell_res = ws_details.cell(row=r, column=c_idx, value=res_formula)
                        cell_res.number_format = "0.0000"
                        cell_res.border = thin_border
                        cell_res.alignment = align_right
                        c_idx += 1
                        
                for col in ws_details.columns:
                    max_len = max(len(str(cell.value or '')) for cell in col)
                    col_letter = get_column_letter(col[0].column)
                    if col[0].column >= 2:
                        ws_details.column_dimensions[col_letter].width = max(max_len + 3, 12)
                        
                # --- SHEET 3: Parámetros de Ajuste ---
                ws_coefs = wb.create_sheet(title=t["xls_sheet_coefs"])
                ws_coefs.views.sheetView[0].showGridLines = True
                
                ws_coefs.cell(row=2, column=2, value=t["xls_coefs_title"]).font = title_font
                ws_coefs.cell(row=3, column=2, value=t["xls_coefs_subtitle"]).font = subtitle_font
                
                headers_coefs = t["xls_coefs_headers"]
                for col_idx, header in enumerate(headers_coefs, start=2):
                    cell = ws_coefs.cell(row=5, column=col_idx, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = align_center
                    cell.border = thin_border
                    
                coefs_data = []
                coefs_data.append([t["methods_names"]["SLR"], t["xls_param_slope"], m_slr, t["xls_param_intercept"], c_slr, "", ""])
                coefs_data.append([t["methods_names"]["GOR Conv"], t["xls_param_slope"], m_gor, t["xls_param_intercept"], b_gor, "", ""])
                coefs_data.append([t["methods_names"]["GOR Prop"], t["xls_param_slope"], m_prop, t["xls_param_intercept"], b_prop, "", ""])
                coefs_data.append([t["methods_names"]["MoM"], t["xls_param_slope"], m_mom, t["xls_param_intercept"], b_mom, "", ""])
                if exp_asymp_valid:
                    coefs_data.append([t["methods_names"]["No Lin. Exponencial Asintota"], t["xls_param_asymptote"], a_val, t["xls_param_scale_b"], b_nl_exp, t["xls_param_rate_c"], c_nl_exp])
                if log_valid:
                    coefs_data.append([t["methods_names"]["No Lin. Logarítmico"], t["xls_param_intercept_a"], a_nl_log, t["xls_param_scale_b"], b_nl_log, "", ""])
                if pot_valid:
                    coefs_data.append([t["methods_names"]["No Lin. Potencial / Power Law"], t["xls_param_scale_a"], a_nl_pot, t["xls_param_exponent_b"], b_nl_pot, "", ""])
                if quad_valid:
                    coefs_data.append([t["methods_names"]["No Lin. Cuadrático"], t["xls_param_quad_a"], a_nl_quad, t["xls_param_quad_b"], b_nl_quad, t["xls_param_quad_c"], c_nl_quad])
                if fun_pot_valid:
                    coefs_data.append([t["methods_names"]["Fun. Potencia log10 (Correcto)"], t["xls_param_scale_a"], a_correct, t["xls_param_exponent_b"], b_correct, "", ""])
                    coefs_data.append([t["methods_names"]["Fun. Potencia log10 (Incorrecto)"], f"{t['xls_param_scale_a']} (err)", a_err, f"{t['xls_param_exponent_b']} (err)", b_err, "", ""])
                if fun_exp_valid:
                    coefs_data.append([t["methods_names"]["Fun. Exponencial Simple ln"], t["xls_param_scale_a"], a_coeff, t["xls_param_exponent_b"], b_coeff, "", ""])

                current_row = 6
                for r_data in coefs_data:
                    ws_coefs.cell(row=current_row, column=2, value=r_data[0]).font = bold_font
                    
                    ws_coefs.cell(row=current_row, column=3, value=r_data[1]).font = regular_font
                    if isinstance(r_data[2], (int, float)):
                        ws_coefs.cell(row=current_row, column=4, value=r_data[2]).number_format = "0.000000"
                    else:
                        ws_coefs.cell(row=current_row, column=4, value=r_data[2])
                        
                    ws_coefs.cell(row=current_row, column=5, value=r_data[3]).font = regular_font
                    if isinstance(r_data[4], (int, float)):
                        ws_coefs.cell(row=current_row, column=6, value=r_data[4]).number_format = "0.000000"
                    else:
                        ws_coefs.cell(row=current_row, column=6, value=r_data[4])
                        
                    ws_coefs.cell(row=current_row, column=7, value=r_data[5]).font = regular_font
                    if isinstance(r_data[6], (int, float)):
                        ws_coefs.cell(row=current_row, column=8, value=r_data[6]).number_format = "0.000000"
                    else:
                        ws_coefs.cell(row=current_row, column=8, value=r_data[6])
                        
                    for col_idx in range(2, 9):
                        cell = ws_coefs.cell(row=current_row, column=col_idx)
                        cell.border = thin_border
                        if col_idx in [2, 3, 5, 7]:
                            cell.alignment = align_left
                        elif col_idx in [4, 6, 8]:
                            cell.alignment = align_right
                    current_row += 1

                for col in ws_coefs.columns:
                    max_len = max(len(str(cell.value or '')) for cell in col)
                    col_letter = get_column_letter(col[0].column)
                    if col[0].column >= 2:
                        ws_coefs.column_dimensions[col_letter].width = max(max_len + 3, 12)

                output = io.BytesIO()
                wb.save(output)
                return output.getvalue()

            try:
                excel_data = generate_excel_with_formulas()
                st.download_button(
                    label=t["t8_export_excel_btn"],
                    data=excel_data,
                    file_name="Reporte_Comparacion_Regresiones.xlsx" if st.session_state.lang == 'es' else "Regression_Comparison_Report.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            except Exception as e:
                st.error(f"Error generando Excel: {e}")
            
        with col_out2:
            st.markdown(f"**{t['t8_export_pdf_lbl']}**")
            
            def create_pdf():
                t_pdf = t
                
                # Rebuild comparison dataframe for PDF using t_pdf to avoid language conflicts
                comp_data_pdf = {
                    t_pdf["table_col_method"]: [t_pdf["methods_names"]["SLR"], t_pdf["methods_names"]["GOR Conv"], t_pdf["methods_names"]["GOR Prop"], t_pdf["methods_names"]["MoM"]],
                    t_pdf["table_col_slope"]: [m_slr, m_gor, m_prop, m_mom],
                    t_pdf["table_col_intercept"]: [c_slr, b_gor, b_prop, b_mom],
                    t_pdf["table_col_se_m"]: [se_m_slr, se_m_gor, se_m_prop, se_m_mom],
                    t_pdf["table_col_se_c"]: [se_c_slr, se_b_gor, se_b_prop, se_b_mom],
                    t_pdf["table_col_rmse"]: [rmse_slr, rmse_gor, rmse_prop, rmse_mom],
                    t_pdf["table_col_r2"]: [r2_slr, r2_gor, r2_prop, r2_mom]
                }
                if exp_asymp_valid:
                    comp_data_pdf[t_pdf["table_col_method"]].append(t_pdf["methods_names"]["No Lin. Exponencial Asintota"])
                    comp_data_pdf[t_pdf["table_col_slope"]].append(np.nan)
                    comp_data_pdf[t_pdf["table_col_intercept"]].append(np.nan)
                    comp_data_pdf[t_pdf["table_col_se_m"]].append(np.nan)
                    comp_data_pdf[t_pdf["table_col_se_c"]].append(np.nan)
                    comp_data_pdf[t_pdf["table_col_rmse"]].append(rmse_nl_exp)
                    comp_data_pdf[t_pdf["table_col_r2"]].append(r2_nl_exp)

                if log_valid:
                    comp_data_pdf[t_pdf["table_col_method"]].append(t_pdf["methods_names"]["No Lin. Logarítmico"])
                    comp_data_pdf[t_pdf["table_col_slope"]].append(np.nan)
                    comp_data_pdf[t_pdf["table_col_intercept"]].append(np.nan)
                    comp_data_pdf[t_pdf["table_col_se_m"]].append(np.nan)
                    comp_data_pdf[t_pdf["table_col_se_c"]].append(np.nan)
                    comp_data_pdf[t_pdf["table_col_rmse"]].append(rmse_nl_log)
                    comp_data_pdf[t_pdf["table_col_r2"]].append(r2_nl_log)

                if pot_valid:
                    comp_data_pdf[t_pdf["table_col_method"]].append(t_pdf["methods_names"]["No Lin. Potencial / Power Law"])
                    comp_data_pdf[t_pdf["table_col_slope"]].append(np.nan)
                    comp_data_pdf[t_pdf["table_col_intercept"]].append(np.nan)
                    comp_data_pdf[t_pdf["table_col_se_m"]].append(np.nan)
                    comp_data_pdf[t_pdf["table_col_se_c"]].append(np.nan)
                    comp_data_pdf[t_pdf["table_col_rmse"]].append(rmse_nl_pot)
                    comp_data_pdf[t_pdf["table_col_r2"]].append(r2_nl_pot)

                if quad_valid:
                    comp_data_pdf[t_pdf["table_col_method"]].append(t_pdf["methods_names"]["No Lin. Cuadrático"])
                    comp_data_pdf[t_pdf["table_col_slope"]].append(np.nan)
                    comp_data_pdf[t_pdf["table_col_intercept"]].append(np.nan)
                    comp_data_pdf[t_pdf["table_col_se_m"]].append(np.nan)
                    comp_data_pdf[t_pdf["table_col_se_c"]].append(np.nan)
                    comp_data_pdf[t_pdf["table_col_rmse"]].append(rmse_nl_quad)
                    comp_data_pdf[t_pdf["table_col_r2"]].append(r2_nl_quad)

                if fun_pot_valid:
                    comp_data_pdf[t_pdf["table_col_method"]].append(t_pdf["methods_names"]["Fun. Potencia log10 (Correcto)"])
                    comp_data_pdf[t_pdf["table_col_slope"]].append(np.nan)
                    comp_data_pdf[t_pdf["table_col_intercept"]].append(np.nan)
                    comp_data_pdf[t_pdf["table_col_se_m"]].append(np.nan)
                    comp_data_pdf[t_pdf["table_col_se_c"]].append(np.nan)
                    comp_data_pdf[t_pdf["table_col_rmse"]].append(rmse_pot10_correct)
                    comp_data_pdf[t_pdf["table_col_r2"]].append(r2_pot10_correct)

                    comp_data_pdf[t_pdf["table_col_method"]].append(t_pdf["methods_names"]["Fun. Potencia log10 (Incorrecto)"])
                    comp_data_pdf[t_pdf["table_col_slope"]].append(np.nan)
                    comp_data_pdf[t_pdf["table_col_intercept"]].append(np.nan)
                    comp_data_pdf[t_pdf["table_col_se_m"]].append(np.nan)
                    comp_data_pdf[t_pdf["table_col_se_c"]].append(np.nan)
                    comp_data_pdf[t_pdf["table_col_rmse"]].append(rmse_pot10_incorrect)
                    comp_data_pdf[t_pdf["table_col_r2"]].append(r2_pot10_incorrect)

                if fun_exp_valid:
                    comp_data_pdf[t_pdf["table_col_method"]].append(t_pdf["methods_names"]["Fun. Exponencial Simple ln"])
                    comp_data_pdf[t_pdf["table_col_slope"]].append(np.nan)
                    comp_data_pdf[t_pdf["table_col_intercept"]].append(np.nan)
                    comp_data_pdf[t_pdf["table_col_se_m"]].append(np.nan)
                    comp_data_pdf[t_pdf["table_col_se_c"]].append(np.nan)
                    comp_data_pdf[t_pdf["table_col_rmse"]].append(rmse_exp_simple)
                    comp_data_pdf[t_pdf["table_col_r2"]].append(r2_exp_simple)

                df_comp_pdf = pd.DataFrame(comp_data_pdf)

                # Crear imagen de Matplotlib estática en background para el PDF
                fig_pdf, ax_pdf = plt.subplots(figsize=(8, 5))
                ax_pdf.scatter(X, Y, color='black', label=t_pdf["chart_obs"])
                x_vals = np.linspace(min(X), max(X), 100)
                ax_pdf.plot(x_vals, m_slr * x_vals + c_slr, color='blue', label=t_pdf["chart_slr"])
                ax_pdf.plot(x_vals, m_gor * x_vals + b_gor, color='orange', label=t_pdf["chart_gor_conv"])
                ax_pdf.plot(x_vals, m_prop * x_vals + b_prop, color='green', label=t_pdf["chart_gor_prop"])
                ax_pdf.plot(x_vals, m_mom * x_vals + b_mom, color='magenta', linestyle=':', label=t_pdf["methods_names"]["MoM"])
                
                if nl_valid:
                    if selected_nl_key == "Exponential":
                        y_vals_nl = a_nl + b_nl * np.exp(c_nl * x_vals)
                    elif selected_nl_key == "Logarithmic":
                        x_vals_safe = np.maximum(x_vals, 1e-9)
                        y_vals_nl = a_nl + b_nl * np.log(x_vals_safe)
                    elif selected_nl_key == "Power":
                        x_vals_safe = np.maximum(x_vals, 1e-9)
                        y_vals_nl = a_nl * (x_vals_safe ** b_nl)
                    elif selected_nl_key == "Quadratic":
                        y_vals_nl = a_nl * (x_vals ** 2) + b_nl * x_vals + c_nl
                    
                    nl_short_names_pdf = {
                        "Exponential": t_pdf["nl_models"]["exp"].split(' ')[0],
                        "Logarithmic": t_pdf["nl_models"]["log"].split(' ')[0],
                        "Power": t_pdf["nl_models"]["pot"].split(' ')[0],
                        "Quadratic": t_pdf["nl_models"]["quad"].split(' ')[0]
                    }
                    selected_nl_short_pdf = nl_short_names_pdf[selected_nl_key]
                    ax_pdf.plot(x_vals, y_vals_nl, color='purple', linestyle='-.', label=t_pdf["chart_nlin"].format(selected_nl_short_pdf))
                
                if power_valid:
                    x_vals_safe = np.maximum(x_vals, 1e-9)
                    y_vals_power = a_power * (x_vals_safe ** b_power)
                    ax_pdf.plot(x_vals, y_vals_power, color='coral', linestyle='--', label=t_pdf["chart_power"])

                ax_pdf.legend()
                ax_pdf.set_title(t_pdf["pdf_chart_title"])
                
                img_buffer = io.BytesIO()
                fig_pdf.savefig(img_buffer, format='png', bbox_inches='tight', dpi=150)
                plt.close(fig_pdf)
                
                pdf = FPDF()
                pdf.add_page()
                pdf.set_font('Arial', 'B', 16)
                pdf.cell(0, 10, t_pdf["pdf_title"], 0, 1, 'C')
                pdf.ln(5)
                
                import tempfile, os
                with tempfile.NamedTemporaryFile(delete=False, suffix='.png') as tmp_file:
                    tmp_file.write(img_buffer.getvalue())
                    tmp_path = tmp_file.name
                pdf.image(tmp_path, x=15, w=180)
                os.unlink(tmp_path)
                
                pdf.ln(5)
                pdf.set_font('Arial', 'B', 12)
                pdf.cell(0, 10, t_pdf["pdf_sec_summary"], 0, 1)
                pdf.set_font('Arial', '', 10)
                
                for i, row in df_comp_pdf.iterrows():
                    method_name = row[t_pdf["table_col_method"]]
                    if method_name == t_pdf["methods_names"]["SLR"]:
                        eq = f"y = {m_slr:.4f}x + {c_slr:.4f}"
                    elif method_name == t_pdf["methods_names"]["GOR Conv"]:
                        eq = f"y = {m_gor:.4f}x + {b_gor:.4f}"
                    elif method_name == t_pdf["methods_names"]["GOR Prop"]:
                        eq = f"y = {m_prop:.4f}x + {b_prop:.4f}"
                    elif method_name == t_pdf["methods_names"]["MoM"]:
                        eq = f"y = {m_mom:.4f}x + {b_mom:.4f}"
                    elif method_name == t_pdf["methods_names"]["No Lin. Exponencial Asintota"]:
                        eq = nl_equation_exp
                    elif method_name == t_pdf["methods_names"]["No Lin. Logarítmico"]:
                        eq = nl_equation_log
                    elif method_name == t_pdf["methods_names"]["No Lin. Potencial / Power Law"]:
                        eq = nl_equation_pot
                    elif method_name == t_pdf["methods_names"]["No Lin. Cuadrático"]:
                        eq = nl_equation_quad
                    elif method_name == t_pdf["methods_names"]["Fun. Potencia log10 (Correcto)"]:
                        eq = f"y = {a_correct:.4f} * x^{{{b_correct:.4f}}}"
                    elif method_name == t_pdf["methods_names"]["Fun. Potencia log10 (Incorrecto)"]:
                        eq = f"y = {a_err:.4f} * x^{{{b_err:.4f}}}"
                    elif method_name == t_pdf["methods_names"]["Fun. Exponencial Simple ln"]:
                        eq = f"y = {a_coeff:.4f} * e^{{{b_coeff:.4f}x}}"
                    else:
                        eq = ""
                    clean_eq = eq.replace('\\cdot', '*').replace('\\ln', 'ln').replace('\\bar', '').replace('\\hat', '').replace('{', '').replace('}', '')
                    pdf.cell(0, 6, f"{method_name}: {clean_eq} | RMSE: {row[t_pdf['table_col_rmse']]:.4f} | R2: {row[t_pdf['table_col_r2']]:.4f}", 0, 1)
                
                return bytes(pdf.output())

            try:
                pdf_data = create_pdf()
                st.download_button(
                    label=t["t8_export_pdf_btn"],
                    data=pdf_data,
                    file_name="Reporte_Comparacion_Regresiones.pdf" if st.session_state.lang == 'es' else "Regression_Comparison_Report.pdf",
                    mime="application/pdf"
                )
            except Exception as e:
                st.error(t["pdf_err_msg"].format(e))



else:
    st.info("👈 Por favor, ingresa o sube datos con al menos 2 puntos para comenzar el análisis comparativo.")

# --- FOOTER ---
st.markdown("---")
st.markdown(
    "<p style='text-align: center; color: gray; font-size: 14px;'>"
    "Desarrollado y mantenido por <b>Alexander Acosta</b> "
    "(<a href='https://github.com/j-alexander-acosta' target='_blank' style='color: #1f77b4; text-decoration: none;'>@j-alexander-acosta</a>)"
    "</p>", 
    unsafe_allow_html=True
)
